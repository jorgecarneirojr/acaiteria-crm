"""
CRM Simples - Açaiteria Combina Açaí
Aplicação Flask para gerenciamento de clientes e vendas

Desenvolvido por: Grupo 22 - Projeto Integrador UNIVESP
Data: 2026
"""

from flask import (
    Flask,
    g,
    render_template,
    request,
    jsonify,
    send_file,
    session,
    redirect,
)
from flask_restx import Api, Namespace, Resource
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_wtf.csrf import CSRFProtect
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from functools import wraps
import csv
import io
import logging
import math
import os
import re as _re
import secrets
import unicodedata
from pydantic import BaseModel, EmailStr, ValidationError, field_validator
from sqlalchemy.orm import joinedload

from .models import (
    db,
    Usuario,
    Cliente,
    Produto,
    Complemento,
    Venda,
    ItemVenda,
    ItemVendaComplemento,
    Pagamento,
    ConsentimentoHistorico,
    LogAcao,
    TicketSuporte,
    MensagemTicket,
    Fornecedor,
    CompraEstoque,
    ItemCompra,
    CupomDesconto,
    BadgeCliente,
    LancamentoFinanceiro,
    TwoFactorSecret,
    ComboKit,
    ComboKitItem,
    Indicacao,
    Assinatura,
    AssinaturaCliente,
    WebhookConfig,
    Loja,
)

# =============================================================================
# CONFIGURAÇÃO INICIAL
# =============================================================================

from dotenv import load_dotenv

load_dotenv()

# Resolver caminhos absolutos para templates e arquivos estáticos
basedir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
template_dir = os.path.join(basedir, "frontend")
static_dir = os.path.join(basedir, "frontend", "static")

app = Flask(
    __name__,
    template_folder=template_dir,
    static_folder=static_dir,
    static_url_path="/static",
)

# Chave secreta para sessões (obrigatória em produção)
_secret = os.environ.get("SECRET_KEY", "")
if not _secret:
    if os.environ.get("FLASK_ENV") == "production":
        raise RuntimeError("SECRET_KEY obrigatória em produção!")
    # Dev/test: chave fixa para não invalidar sessão a cada restart
    _secret = "dev-only-insecure-key-do-not-use-in-prod"
app.config["SECRET_KEY"] = _secret

# Sessão permanente — dura 7 dias
# (evita logout inesperado entre reinicializações)
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)

# Proteção de cookies de sessão
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
if os.environ.get("FLASK_ENV") == "production":
    app.config["SESSION_COOKIE_SECURE"] = True
    # Render termina TLS no proxy; confiar no header X-Forwarded-Proto
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)


# --- Helpers portáveis para comparação de datas (SQLite + PostgreSQL) ---
def _dia_inicio(d):
    """Retorna datetime no início do dia (00:00:00 UTC) — portável."""
    return datetime.combine(d, datetime.min.time(), tzinfo=timezone.utc)


def _dia_fim(d):
    """Retorna datetime no fim do dia (23:59:59.999999 UTC) — portável."""
    return datetime.combine(d, datetime.max.time(), tzinfo=timezone.utc)


# Desabilitar Swagger em produção
_doc_path = (
    "/api/docs" if os.environ.get("FLASK_ENV") != "production" else False
)

api = Api(
    app,
    version="1.0",
    title="Acaiteria CRM API",
    description=(
        "API REST do CRM — validação Pydantic,"
        " conformidade LGPD, rate-limiting."
    ),
    doc=_doc_path,
    prefix="/api",
)


def _rate_limit_key():
    """Rate-limit por usuario_id (se logado) ou IP."""
    from flask import session as _sess
    uid = _sess.get("usuario_id")
    return f"user:{uid}" if uid else get_remote_address()


limiter = Limiter(
    _rate_limit_key,
    app=app,
    default_limits=["300 per hour"],
    storage_uri="memory://",
)

# --- Cache Layer (#6) — SimpleCache para endpoints de leitura ---
from flask_caching import Cache  # noqa: E402

cache = Cache(app, config={
    "CACHE_TYPE": "SimpleCache",
    "CACHE_DEFAULT_TIMEOUT": 300,
})

# --- 2FA — import pyotp ---
import pyotp  # noqa: E402

# --- WebSocket (#22) — opcional, não quebra se não instalado ---
try:
    from flask_socketio import SocketIO, emit  # noqa: F401
    socketio = SocketIO(
        app,
        cors_allowed_origins=os.environ.get(
            "SOCKETIO_ORIGINS",
            "https://acaiteria-crm.onrender.com"
        ).split(","),
        async_mode="threading",
    )

    @socketio.on("connect")
    def _ws_connect():
        emit("status", {"msg": "Conectado ao CRM"})

    @socketio.on("ping_crm")
    def _ws_ping(data):
        emit("pong_crm", {"echo": data})
except ImportError:
    socketio = None

health_ns = Namespace(
    "health", description="Healthcheck e metadados da API", path="/"
)
api.add_namespace(health_ns)

# Proteção CSRF — ativamos apenas para rotas de formulário (não-API)
# Rotas /api/* são protegidas por session auth + corpo JSON
app.config["WTF_CSRF_CHECK_DEFAULT"] = False
csrf = CSRFProtect(app)


@app.before_request
def _csrf_protect_forms():
    """Aplicar CSRF apenas em formulários HTML (fora de /api/)."""
    if app.config.get("TESTING"):
        return
    if (
        request.method in ("POST", "PUT", "DELETE", "PATCH")
        and not request.path.startswith("/api/")
    ):
        csrf.protect()


# Idle timeout — encerra sessão após 30 min de inatividade
_SESSION_IDLE_MINUTES = 30


@app.before_request
def _check_session_idle():
    """Invalida sessão se inativa por mais de _SESSION_IDLE_MINUTES."""
    if request.path.startswith("/static"):
        return
    now_ts = datetime.now(timezone.utc).timestamp()
    last = session.get("_last_active")
    if last and (now_ts - last) > _SESSION_IDLE_MINUTES * 60:
        usuario_id = session.get("usuario_id")
        session.clear()
        if request.path.startswith("/api/"):
            return  # próximo decorator cuidará do 401
        if usuario_id:
            return redirect("/login")
    session["_last_active"] = now_ts


# Configuração do banco de dados
# Railway fornece postgres:// mas SQLAlchemy 2.x exige postgresql://
# Usar caminho absoluto para SQLite evitar criar bancos em locais diferentes
_db_path = os.path.join(basedir, "instance", "acaiteria.db")
_default_db = f"sqlite:///{_db_path}"
database_url = os.environ.get("DATABASE_URL", _default_db)
if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

# Validar que DATABASE_URL está definida em produção (evitar SQLite efêmero)
if os.environ.get("FLASK_ENV") == "production" and not os.environ.get("DATABASE_URL"):
    raise RuntimeError(
        "DATABASE_URL obrigatória em produção! "
        "Configure no dashboard Render ou render.yaml."
    )

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ECHO"] = os.environ.get("FLASK_ENV") == "development"

# Connection pooling para PostgreSQL em produção
# Render free tier: ~30 conexões máx; com 2 workers, usar pool conservador
if database_url.startswith("postgresql"):
    app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
        "pool_size": int(os.environ.get("DB_POOL_SIZE", "3")),
        "max_overflow": int(os.environ.get("DB_MAX_OVERFLOW", "5")),
        "pool_pre_ping": True,
        "pool_recycle": 3600,
    }

# Inicializar banco de dados
db.init_app(app)

# Criar tabelas automaticamente no primeiro request
# (necessário em produção/cloud)
with app.app_context():
    db.create_all()

    # Auto-migração: adicionar colunas novas em tabelas existentes
    # (SQLite/Postgres)
    _migracoes = [
        ("produto", "estoque_atual", "INTEGER DEFAULT 0"),
        ("produto", "estoque_minimo", "INTEGER DEFAULT 0"),
        ("produto", "volume", "VARCHAR(20)"),
        ("cliente", "pontos_fidelidade", "INTEGER DEFAULT 0"),
        ("cliente", "senha_hash", "VARCHAR(256)"),
        ("venda", "status_pedido", "VARCHAR(30) DEFAULT 'Recebido'"),
        ("venda", "motivo_cancelamento", "TEXT"),
        ("produto", "preco_promocional", "DECIMAL(10,2)"),
        ("venda", "desconto_aplicado", "DECIMAL(10,2) DEFAULT 0"),
    ]
    # Colunas vindas de dict interno – nunca de input externo
    _sql_map = {
        (t, c): tipo for t, c, tipo in _migracoes
    }
    with db.engine.begin() as conn:
        for (tabela, coluna), tipo in _sql_map.items():
            try:
                conn.execute(db.text(
                    "ALTER TABLE {} ADD COLUMN {} {}".format(
                        tabela, coluna, tipo
                    )
                ))
            except Exception:
                pass  # coluna já existe — OK (tanto SQLite como PostgreSQL)

    # Auto-criação de índices para colunas frequentemente filtradas
    _indices = [
        ("idx_cliente_ativo", "cliente", "ativo"),
        ("idx_produto_categoria", "produto", "categoria"),
        ("idx_venda_status_pedido", "venda", "status_pedido"),
        ("idx_venda_data_venda", "venda", "data_venda"),
        ("idx_venda_id_cliente", "venda", "id_cliente"),
        ("idx_fornecedor_email", "fornecedor", "email"),
    ]
    with db.engine.begin() as conn:
        for idx_name, tabela, coluna in _indices:
            try:
                conn.execute(db.text(
                    "CREATE INDEX IF NOT EXISTS {} ON {} ({})".format(
                        idx_name, tabela, coluna
                    )
                ))
            except Exception:
                pass  # índice já existe — OK

# Configurar logging estruturado
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
)
logger = logging.getLogger("acaiteria-crm")


def _erro_interno(e):
    """Loga exceção e retorna resposta segura (sem stack trace em produção)."""
    logger.exception("Erro interno: %s", e)
    if app.config.get("TESTING"):
        return jsonify({"erro": str(e)}), 500
    return jsonify({"erro": "Erro interno do servidor"}), 500


# =============================================================================
# SECURITY HEADERS
# =============================================================================


@app.after_request
def adicionar_headers_seguranca(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = (
        "geolocation=(), camera=(), microphone=()"
    )
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' "
        "https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "worker-src 'self'; "
        "frame-ancestors 'none'"
    )
    if os.environ.get("FLASK_ENV") == "production":
        response.headers["Strict-Transport-Security"] = (
            "max-age=31536000; includeSubDomains"
        )
    # Cache-Control para assets estáticos (1 semana, SW gerencia invalidação)
    if request.path.startswith("/static/"):
        response.headers["Cache-Control"] = (
            "public, max-age=604800, stale-while-revalidate=86400"
        )
    return response


@app.before_request
def gerar_csp_nonce():
    """Gera nonce CSP único por request para scripts inline."""
    g.csp_nonce = secrets.token_hex(16)


@app.context_processor
def inject_user():
    """Injeta dados do usuário logado e nonce CSP em todos os templates."""
    return {
        "usuario_nome": session.get("usuario_nome", ""),
        "papel": session.get("papel", ""),
        "csp_nonce": g.get("csp_nonce", ""),
    }


# -- Helper de validação de e-mail --------------------
_EMAIL_RE = _re.compile(
    r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$"
)


def _email_valido(email: str) -> bool:
    """Retorna True se o e-mail possuir formato válido."""
    return bool(_EMAIL_RE.match(email))


# -- Helper de paginação -------------------------
def get_pagination_params(default_per_page=50):
    """Extrai e valida parâmetros de paginação da query string."""
    pagina = max(1, min(request.args.get("pagina", 1, type=int), 10000))
    por_pagina = max(
        1, min(request.args.get("por_pagina", default_per_page, type=int), 100)
    )
    return pagina, por_pagina


class ClienteCreateSchema(BaseModel):
    nome: str
    telefone: str | None = None
    email: EmailStr | None = None
    observacoes: str | None = None
    consentimento_lgpd: bool = False
    versao_politica: str = "v1.0"

    @field_validator("nome")
    @classmethod
    def nome_obrigatorio(cls, value: str) -> str:
        value = value.strip()
        if len(value) < 2:
            raise ValueError("Nome deve ter ao menos 2 caracteres")
        return value


class ProdutoCreateSchema(BaseModel):
    nome_produto: str
    categoria: str | None = None
    descricao: str | None = None
    preco: float
    volume: str | None = None
    estoque_atual: int | None = 0
    estoque_minimo: int | None = 0

    @field_validator("preco")
    @classmethod
    def preco_positivo(cls, value: float) -> float:
        if value <= 0:
            raise ValueError("Preço deve ser maior que zero")
        return value


class VendaItemSchema(BaseModel):
    id_produto: int
    quantidade: int
    complementos: list[int] = []


class VendaCreateSchema(BaseModel):
    id_cliente: int
    forma_pagamento: str = "Dinheiro"
    observacoes: str | None = None
    desconto_percentual: float = 0.0
    taxa: float = 0.0
    cupom_codigo: str | None = None
    itens: list[VendaItemSchema]

    @field_validator("desconto_percentual")
    @classmethod
    def desconto_valido(cls, v: float) -> float:
        if v < 0 or v > 100:
            raise ValueError("Desconto deve estar entre 0 e 100%")
        return v

    @field_validator("taxa")
    @classmethod
    def taxa_valida(cls, v: float) -> float:
        if v < 0:
            raise ValueError("Taxa não pode ser negativa")
        return v


class ConsentimentoSchema(BaseModel):
    consentimento_lgpd: bool
    versao_politica: str = "v1.0"


class ProdutoUpdateSchema(BaseModel):
    nome_produto: str | None = None
    categoria: str | None = None
    descricao: str | None = None
    preco: float | None = None
    ativo: bool | None = None
    estoque_atual: int | None = None
    estoque_minimo: int | None = None


class LancamentoFinanceiroSchema(BaseModel):
    tipo: str
    categoria: str
    descricao: str | None = None
    valor: float
    data_lancamento: str
    forma_pagamento: str | None = None
    status: str | None = "Pago"
    comprovante: str | None = None
    observacoes: str | None = None

    @field_validator("tipo")
    @classmethod
    def tipo_valido(cls, v: str) -> str:
        if v not in ("receita", "despesa"):
            raise ValueError("Tipo deve ser 'receita' ou 'despesa'")
        return v

    @field_validator("valor")
    @classmethod
    def valor_positivo(cls, v: float) -> float:
        if v <= 0:
            raise ValueError("Valor deve ser maior que zero")
        return v

    @field_validator("status")
    @classmethod
    def status_valido(cls, v: str | None) -> str:
        validos = ("Pago", "Pendente", "Cancelado")
        if v and v not in validos:
            raise ValueError(
                f"Status deve ser um de: {', '.join(validos)}"
            )
        return v or "Pago"


def validar_payload(schema_cls):
    dados = request.get_json(silent=True) or {}
    try:
        return schema_cls.model_validate(dados).model_dump()
    except ValidationError as e:
        raise ValueError(e.errors())


# =============================================================================
# AUTENTICAÇÃO — Login com email + senha e papéis (admin / operador)
# =============================================================================


def login_required(f):
    """Decorator que protege rotas HTML exigindo login."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("usuario_id"):
            return redirect("/login")
        return f(*args, **kwargs)

    return decorated


def api_login_required(f):
    """Decorator que protege rotas API exigindo login."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("usuario_id"):
            return jsonify({"erro": "Autenticação necessária"}), 401
        return f(*args, **kwargs)

    return decorated


def admin_required(f):
    """Decorator que exige papel 'admin' para rotas HTML."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("usuario_id"):
            return redirect("/login")
        if session.get("papel") != "admin":
            return (
                render_template(
                    "index.html",
                    stats=_stats_default(),
                    error="Acesso restrito a administradores",
                ),
                403,
            )
        return f(*args, **kwargs)

    return decorated


def api_admin_required(f):
    """Decorator que exige papel 'admin' para rotas API."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("usuario_id"):
            return jsonify({"erro": "Autenticação necessária"}), 401
        if session.get("papel") != "admin":
            return jsonify({"erro": "Acesso restrito a administradores"}), 403
        return f(*args, **kwargs)

    return decorated


def _stats_default():
    return {
        "total_clientes": 0,
        "total_vendas": 0,
        "faturamento_total": 0,
        "vendas_semana": 0,
        "clientes_consentimento": 0,
        "taxa_consentimento": 0,
    }


def registrar_log(acao, entidade, id_entidade=None, detalhes=None):
    """Registra ação no audit log."""
    try:
        log = LogAcao(
            id_usuario=session.get("usuario_id"),
            acao=acao,
            entidade=entidade,
            id_entidade=id_entidade,
            detalhes=detalhes,
            ip=request.remote_addr,
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        logger.warning("Falha ao registrar log: %s", e)
        try:
            db.session.rollback()
        except Exception:
            pass


@health_ns.route("/health")
class HealthResource(Resource):
    def get(self):
        db_ok = True
        try:
            db.session.execute(db.text("SELECT 1"))
        except Exception:
            db_ok = False
        status = "ok" if db_ok else "degraded"
        code = 200 if db_ok else 503
        return {
            "status": status,
            "service": "acaiteria-crm",
            "database": "connected" if db_ok else "unavailable",
            "timestamp": datetime.now(timezone.utc).isoformat(),
        }, code


# =============================================================================
# ROTAS - AUTENTICAÇÃO
# =============================================================================


@app.route("/offline")
def offline_page():
    """Página offline para PWA"""
    return render_template("offline.html")


@app.route("/login", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def login():
    """Página de login com email e senha"""
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        senha = request.form.get("senha", "")
        usuario = Usuario.query.filter_by(email=email, ativo=True).first()
        if usuario and usuario.verificar_senha(senha):
            # --- 2FA check ---
            tf = TwoFactorSecret.query.filter_by(
                id_usuario=usuario.id_usuario, ativo=True
            ).first()
            if tf:
                codigo_2fa = request.form.get("codigo_2fa", "").strip()
                if not codigo_2fa:
                    # Guardar credenciais na session (server-side)
                    # em vez de expor a senha no HTML
                    session["_2fa_user_id"] = usuario.id_usuario
                    session["_2fa_nonce"] = secrets.token_hex(16)
                    return render_template(
                        "login.html", need_2fa=True,
                        email=email,
                    )
                # Se veio via session (2FA step 2)
                if not session.get("_2fa_user_id"):
                    return render_template(
                        "login.html",
                        erro="Sessão 2FA expirada. Tente novamente.",
                    )
                totp = pyotp.TOTP(tf.secret)
                if not totp.verify(codigo_2fa, valid_window=1):
                    return render_template(
                        "login.html", need_2fa=True,
                        email=email,
                        erro="Código 2FA inválido.",
                    )
                # Limpar tokens temporários
                session.pop("_2fa_user_id", None)
                session.pop("_2fa_nonce", None)
            # --- fim 2FA ---
            session.permanent = True
            session["usuario_id"] = usuario.id_usuario
            session["usuario_nome"] = usuario.nome
            session["papel"] = usuario.papel
            session["autenticado"] = True  # compatibilidade
            registrar_log(
                "login",
                "usuario",
                usuario.id_usuario,
                f"Login de {usuario.nome}",
            )
            return redirect("/")
        logger.warning(
            "Tentativa de login falha para %s de %s",
            email,
            request.remote_addr,
        )
        return render_template("login.html", erro="Email ou senha incorretos.")
    return render_template("login.html")


@app.route("/logout")
def logout():
    """Encerrar sessão"""
    registrar_log("logout", "usuario", session.get("usuario_id"), "Logout")
    session.clear()
    return redirect("/login")


# =============================================================================
# ROTAS - PÁGINA INICIAL
# =============================================================================


@app.route("/")
@login_required
def index():
    """Página inicial - Dashboard"""
    try:
        # Estatísticas gerais
        total_clientes = Cliente.query.filter_by(ativo=True).count()
        total_vendas = Venda.query.count()
        faturamento_total = (
            db.session.query(db.func.sum(Venda.valor_total)).scalar() or 0
        )

        # Últimas vendas (últimos 7 dias)
        vendas_semana = Venda.query.filter(
            Venda.data_venda >= datetime.now(timezone.utc) - timedelta(days=7)
        ).count()

        # Clientes com permissão LGPD
        clientes_consentimento = Cliente.query.filter_by(
            ativo=True, consentimento_lgpd=True
        ).count()

        stats = {
            "total_clientes": total_clientes,
            "total_vendas": total_vendas,
            "faturamento_total": (
                float(faturamento_total) if faturamento_total else 0
            ),
            "vendas_semana": vendas_semana,
            "clientes_consentimento": clientes_consentimento,
            "taxa_consentimento": round(
                (
                    (clientes_consentimento / total_clientes * 100)
                    if total_clientes > 0
                    else 0
                ),
                2,
            ),
        }

        return render_template(
            "index.html",
            stats=stats,
            usuario_nome=session.get("usuario_nome", ""),
            papel=session.get("papel", ""),
        )
    except Exception as e:
        logger.exception("Erro no dashboard: %s", e)
        return render_template(
            "index.html",
            stats=_stats_default(),
            error=str(e),
            usuario_nome=session.get("usuario_nome", ""),
            papel=session.get("papel", ""),
        )


# =============================================================================
# ROTAS - DASHBOARD KPI (API JSON)
# =============================================================================


@app.route("/api/dashboard/kpi", methods=["GET"])
@limiter.limit("30 per minute")
@api_login_required
def dashboard_kpi():
    """Retorna KPIs do dashboard em JSON (para consumo por frontend SPA)."""
    try:
        total_clientes = Cliente.query.filter_by(ativo=True).count()
        total_vendas = Venda.query.count()
        faturamento_total = (
            db.session.query(db.func.sum(Venda.valor_total)).scalar() or 0
        )
        vendas_semana = Venda.query.filter(
            Venda.data_venda >= datetime.now(timezone.utc) - timedelta(days=7)
        ).count()
        clientes_consentimento = Cliente.query.filter_by(
            ativo=True, consentimento_lgpd=True
        ).count()
        ticket_medio = (
            float(faturamento_total / total_vendas)
            if total_vendas > 0 else 0
        )
        produtos_ativos = Produto.query.filter_by(ativo=True).count()
        estoque_baixo = Produto.query.filter(
            Produto.ativo.is_(True),
            Produto.estoque_minimo > 0,
            Produto.estoque_atual <= Produto.estoque_minimo,
        ).count()

        return jsonify({
            "total_clientes": total_clientes,
            "total_vendas": total_vendas,
            "faturamento_total": float(faturamento_total),
            "vendas_semana": vendas_semana,
            "clientes_consentimento": clientes_consentimento,
            "taxa_consentimento": round(
                (clientes_consentimento / total_clientes * 100)
                if total_clientes > 0 else 0, 2
            ),
            "ticket_medio": round(ticket_medio, 2),
            "produtos_ativos": produtos_ativos,
            "estoque_baixo": estoque_baixo,
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - CLIENTES
# =============================================================================


@app.route("/api/clientes", methods=["GET"])
@limiter.limit("120 per minute")
@api_login_required
def listar_clientes():
    """Listar todos os clientes ativos com busca e paginação"""
    try:
        query = Cliente.query.filter_by(ativo=True)

        # Busca por nome, telefone ou email
        busca = request.args.get("busca", "").strip()
        if busca:
            filtro = f"%{busca}%"
            query = query.filter(
                db.or_(
                    Cliente.nome.ilike(filtro),
                    Cliente.telefone.ilike(filtro),
                    Cliente.email.ilike(filtro),
                )
            )

        # Paginação
        pagina, por_pagina = get_pagination_params()

        total = query.count()
        clientes = (
            query.order_by(Cliente.nome)
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )

        return jsonify(
            {
                "clientes": [cliente.to_dict() for cliente in clientes],
                "total": total,
                "pagina": pagina,
                "por_pagina": por_pagina,
                "total_paginas": (total + por_pagina - 1) // por_pagina,
            }
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/clientes", methods=["POST"])
@limiter.limit("30 per minute")
@api_login_required
def criar_cliente():
    """Criar novo cliente"""
    try:
        dados = validar_payload(ClienteCreateSchema)

        # Verificar LGPD
        consentimento = dados.get("consentimento_lgpd", False)
        data_consentimento = None

        if consentimento:
            data_consentimento = datetime.now(timezone.utc)

        # Criar cliente
        cliente = Cliente(
            nome=dados.get("nome"),
            telefone=dados.get("telefone"),
            email=dados.get("email"),
            observacoes=dados.get("observacoes"),
            consentimento_lgpd=consentimento,
            data_consentimento=data_consentimento,
            consentimento_versao=(
                dados.get("versao_politica", "v1.0") if consentimento else None
            ),
            ativo=True,
        )

        db.session.add(cliente)
        db.session.flush()  # gera id_cliente antes do commit

        # Registrar histórico LGPD quando há consentimento
        if consentimento:
            entrada = ConsentimentoHistorico(
                id_cliente=cliente.id_cliente,
                acao="concedeu",
                versao_politica=dados.get("versao_politica", "v1.0"),
                ip_address=request.remote_addr,
                user_agent=request.headers.get("User-Agent", "")[:255],
            )
            db.session.add(entrada)

        db.session.commit()
        registrar_log(
            "criar",
            "cliente",
            cliente.id_cliente,
            f"Cliente criado: {cliente.nome}",
        )

        return jsonify(cliente.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return (
                jsonify({"erro": "Payload invalido", "detalhes": str(e)}),
                400,
            )
        return _erro_interno(e)


@app.route("/api/clientes/<int:id_cliente>", methods=["GET"])
@api_login_required
def obter_cliente(id_cliente):
    """Obter detalhes de um cliente"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        # Detalhes com histórico de vendas (agregação SQL, sem N+1)
        cliente_dict = cliente.to_dict()
        stats = db.session.query(
            db.func.count(Venda.id_venda),
            db.func.coalesce(db.func.sum(Venda.valor_total), 0),
        ).filter(Venda.id_cliente == id_cliente).first()
        cliente_dict["total_vendas"] = stats[0]
        cliente_dict["faturamento_total"] = float(stats[1])

        return jsonify(cliente_dict)
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/clientes/<int:id_cliente>", methods=["PUT"])
@api_login_required
def atualizar_cliente(id_cliente):
    """Atualizar dados de um cliente"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({"erro": "Cliente não encontrado"}), 404
        if not cliente.ativo:
            return (
                jsonify({"erro": "Cliente anonimizado não pode ser editado"}),
                400,
            )

        dados = request.get_json(silent=True) or {}

        # Validar nome se fornecido
        if "nome" in dados:
            nome = (dados["nome"] or "").strip()
            if len(nome) < 2:
                return (
                    jsonify({"erro": "Nome deve ter ao menos 2 caracteres"}),
                    400,
                )
            cliente.nome = nome

        # Validar e-mail se fornecido
        if "email" in dados:
            email = (dados["email"] or "").strip().lower() or None
            if email and not _email_valido(email):
                return jsonify({"erro": "E-mail inválido"}), 400
            # Impedir remoção de e-mail se cliente tem senha
            if not email and cliente.senha_hash:
                return (
                    jsonify(
                        {
                            "erro": (
                                "Não é possível remover o e-mail"
                                " de cliente com senha cadastrada"
                            )
                        }
                    ),
                    400,
                )
            # Verificar duplicidade
            if email:
                existente = Cliente.query.filter(
                    Cliente.email == email,
                    Cliente.ativo == True,  # noqa: E712
                    Cliente.id_cliente != id_cliente,
                ).first()
                if existente:
                    return (
                        jsonify(
                            {"erro": "E-mail já cadastrado por outro cliente"}
                        ),
                        409,
                    )
            cliente.email = email

        # Validar telefone se fornecido
        if "telefone" in dados:
            telefone = (dados["telefone"] or "").strip() or None
            if telefone:
                existente = Cliente.query.filter(
                    Cliente.telefone == telefone,
                    Cliente.ativo == True,  # noqa: E712
                    Cliente.id_cliente != id_cliente,
                ).first()
                if existente:
                    return (
                        jsonify(
                            {
                                "erro": "Telefone já cadastrado"
                                " por outro cliente"
                            }
                        ),
                        409,
                    )
            cliente.telefone = telefone

        if "observacoes" in dados:
            cliente.observacoes = dados.get("observacoes")

        db.session.commit()
        registrar_log(
            "editar", "cliente", id_cliente, f"Cliente editado: {cliente.nome}"
        )

        return jsonify(cliente.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/clientes/<int:id_cliente>", methods=["DELETE"])
@api_login_required
def deletar_cliente(id_cliente):
    """Deletar (anonimizar) um cliente - LGPD"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        # Anonimizar ao invés de deletar (LGPD)
        cliente.nome = f"CLIENTE_ANONIMIZADO_{id_cliente}"
        cliente.telefone = None
        cliente.email = None
        cliente.observacoes = None
        cliente.ativo = False
        cliente.data_exclusao = datetime.now(timezone.utc)

        db.session.commit()
        registrar_log(
            "excluir", "cliente", id_cliente, "Cliente anonimizado (LGPD)"
        )

        return jsonify({"mensagem": "Cliente anonimizado conforme LGPD"}), 200
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - LGPD (consentimento e histórico de auditoria)
# =============================================================================


@app.route("/api/clientes/<int:id_cliente>/consentimento", methods=["PUT"])
@limiter.limit("30 per minute")
@api_login_required
def atualizar_consentimento(id_cliente):
    """Concede/revoga consentimento LGPD."""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        dados = validar_payload(ConsentimentoSchema)
        consentiu = bool(dados.get("consentimento_lgpd"))
        versao = dados.get("versao_politica", "v1.0")

        cliente.consentimento_lgpd = consentiu
        cliente.consentimento_versao = versao if consentiu else None
        cliente.data_consentimento = (
            datetime.now(timezone.utc) if consentiu else None
        )

        entrada = ConsentimentoHistorico(
            id_cliente=id_cliente,
            acao="concedeu" if consentiu else "revogou",
            versao_politica=versao,
            ip_address=request.remote_addr,
            user_agent=request.headers.get("User-Agent", "")[:255],
        )
        db.session.add(entrada)
        db.session.commit()

        return jsonify(
            {
                "mensagem": (
                    "Consentimento "
                    f'{"concedido" if consentiu else "revogado"}'
                    " com sucesso"
                ),
                "cliente": cliente.to_dict(),
            }
        )
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return (
                jsonify({"erro": "Payload invalido", "detalhes": str(e)}),
                400,
            )
        return _erro_interno(e)


@app.route(
    "/api/clientes/<int:id_cliente>/consentimento/historico", methods=["GET"]
)
@limiter.limit("120 per minute")
@api_login_required
def historico_consentimento(id_cliente):
    """Retorna histórico completo de auditoria LGPD do cliente"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        historico = (
            ConsentimentoHistorico.query.filter_by(id_cliente=id_cliente)
            .order_by(ConsentimentoHistorico.data_acao.desc())
            .all()
        )

        return jsonify(
            {
                "id_cliente": id_cliente,
                "nome": cliente.nome,
                "consentimento_atual": cliente.consentimento_lgpd,
                "versao_atual": cliente.consentimento_versao,
                "historico": [h.to_dict() for h in historico],
            }
        )
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - PRODUTOS
# =============================================================================


@app.route("/api/produtos", methods=["GET"])
@limiter.limit("120 per minute")
@api_login_required
def listar_produtos():
    """Listar produtos com filtros opcionais."""
    try:
        incluir_inativos = (
            request.args.get("incluir_inativos", "").lower() == "true"
        )
        query = (
            Produto.query
            if incluir_inativos
            else Produto.query.filter_by(ativo=True)
        )

        busca = request.args.get("busca", "").strip()
        if busca:
            filtro = f"%{busca}%"
            query = query.filter(
                db.or_(
                    Produto.nome_produto.ilike(filtro),
                    Produto.descricao.ilike(filtro),
                )
            )

        categoria = request.args.get("categoria", "").strip()
        if categoria:
            query = query.filter(Produto.categoria.ilike(f"%{categoria}%"))

        pagina, por_pagina = get_pagination_params()
        total = query.count()
        produtos = (
            query.order_by(Produto.nome_produto)
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )
        return jsonify({
            "itens": [produto.to_dict() for produto in produtos],
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "paginas": (total + por_pagina - 1) // por_pagina,
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/produtos/estoque-baixo", methods=["GET"])
@api_login_required
def produtos_estoque_baixo():
    """Lista produtos com estoque abaixo do mínimo (alertas)"""
    try:
        produtos = (
            Produto.query.filter(
                Produto.ativo == True,  # noqa: E712
                Produto.estoque_atual <= Produto.estoque_minimo,
                (Produto.estoque_minimo > 0) | (Produto.estoque_atual > 0),
            )
            .order_by(Produto.estoque_atual)
            .all()
        )
        return jsonify([p.to_dict() for p in produtos])
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/produtos", methods=["POST"])
@limiter.limit("30 per minute")
@api_login_required
def criar_produto():
    """Criar novo produto"""
    try:
        dados = validar_payload(ProdutoCreateSchema)

        produto = Produto(
            nome_produto=dados.get("nome_produto"),
            categoria=dados.get("categoria"),
            descricao=dados.get("descricao"),
            preco=Decimal(str(dados.get("preco"))),
            volume=dados.get("volume"),
            estoque_atual=(
                int(dados.get("estoque_atual", 0))
                if dados.get("estoque_atual") is not None
                else 0
            ),
            estoque_minimo=(
                int(dados.get("estoque_minimo", 5))
                if dados.get("estoque_minimo") is not None
                else 5
            ),
            ativo=True,
        )

        db.session.add(produto)
        db.session.commit()
        _invalidar_cache_vitrine()
        registrar_log(
            "criar",
            "produto",
            produto.id_produto,
            f"Produto criado: {produto.nome_produto}",
        )

        return jsonify(produto.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return (
                jsonify({"erro": "Payload invalido", "detalhes": str(e)}),
                400,
            )
        return _erro_interno(e)


@app.route("/api/produtos/<int:id_produto>", methods=["GET"])
@api_login_required
def obter_produto(id_produto):
    """Obter detalhes de um produto"""
    try:
        produto = db.session.get(Produto, id_produto)
        if not produto:
            return jsonify({"erro": "Produto não encontrado"}), 404
        return jsonify(produto.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/produtos/<int:id_produto>", methods=["PUT"])
@limiter.limit("30 per minute")
@api_login_required
def atualizar_produto(id_produto):
    """Atualizar dados de um produto"""
    try:
        produto = db.session.get(Produto, id_produto)
        if not produto:
            return jsonify({"erro": "Produto não encontrado"}), 404

        dados = request.get_json(silent=True) or {}

        if "nome_produto" in dados and dados["nome_produto"]:
            produto.nome_produto = dados["nome_produto"]
        if "categoria" in dados:
            produto.categoria = dados["categoria"]
        if "descricao" in dados:
            produto.descricao = dados["descricao"]
        if "preco" in dados and dados["preco"] is not None:
            produto.preco = Decimal(str(dados["preco"]))
        if "volume" in dados:
            produto.volume = dados["volume"]
        if "ativo" in dados:
            produto.ativo = bool(dados["ativo"])
        if "estoque_atual" in dados and dados["estoque_atual"] is not None:
            produto.estoque_atual = max(0, int(dados["estoque_atual"]))
        if "estoque_minimo" in dados and dados["estoque_minimo"] is not None:
            produto.estoque_minimo = max(0, int(dados["estoque_minimo"]))

        db.session.commit()
        _invalidar_cache_vitrine()
        registrar_log(
            "editar",
            "produto",
            id_produto,
            f"Produto editado: {produto.nome_produto}",
        )
        return jsonify(produto.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/produtos/<int:id_produto>", methods=["DELETE"])
@limiter.limit("30 per minute")
@api_login_required
def deletar_produto(id_produto):
    """Desativar produto (soft delete)"""
    try:
        produto = db.session.get(Produto, id_produto)
        if not produto:
            return jsonify({"erro": "Produto não encontrado"}), 404

        produto.ativo = False
        db.session.commit()
        _invalidar_cache_vitrine()
        registrar_log(
            "excluir",
            "produto",
            id_produto,
            f"Produto desativado: {produto.nome_produto}",
        )
        return jsonify({"mensagem": "Produto desativado com sucesso"})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/produtos/bulk-update", methods=["PATCH"])
@limiter.limit("10 per minute")
@api_admin_required
def bulk_update_produtos():
    """Atualização em lote de produtos (admin). Max 100 por request.

    Body: {"itens": [{"id_produto": 1, "preco": 15.0, "estoque_atual": 50}, ...]}
    Campos permitidos: preco, estoque_atual, estoque_minimo, ativo.
    """
    try:
        dados = request.get_json(silent=True) or {}
        itens = dados.get("itens")
        if not itens or not isinstance(itens, list):
            return jsonify({"erro": "Campo 'itens' obrigatório"}), 400
        if len(itens) > 100:
            return jsonify({"erro": "Máximo 100 itens por lote"}), 400

        _campos_permitidos = {"preco", "estoque_atual", "estoque_minimo", "ativo"}
        atualizados = []
        erros = []

        for item in itens:
            pid = item.get("id_produto")
            if not pid:
                erros.append({"erro": "id_produto ausente", "item": item})
                continue
            produto = db.session.get(Produto, pid)
            if not produto:
                erros.append({"id_produto": pid, "erro": "não encontrado"})
                continue
            campos_atualizados = []
            for campo in _campos_permitidos:
                if campo in item and item[campo] is not None:
                    if campo == "preco":
                        produto.preco = Decimal(str(item["preco"]))
                    elif campo == "estoque_atual":
                        produto.estoque_atual = max(0, int(item["estoque_atual"]))
                    elif campo == "estoque_minimo":
                        produto.estoque_minimo = max(0, int(item["estoque_minimo"]))
                    elif campo == "ativo":
                        produto.ativo = bool(item["ativo"])
                    campos_atualizados.append(campo)
            if campos_atualizados:
                atualizados.append({
                    "id_produto": pid,
                    "campos": campos_atualizados,
                })

        db.session.commit()
        _invalidar_cache_vitrine()
        registrar_log(
            "editar", "produto", None,
            f"Bulk update: {len(atualizados)} produtos atualizados",
        )
        return jsonify({
            "atualizados": atualizados,
            "erros": erros,
            "total_atualizados": len(atualizados),
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - HISTÓRICO DE AÇÕES (AUDIT LOG)
# =============================================================================


@app.route("/api/logs", methods=["GET"])
@limiter.limit("60 per minute")
@api_admin_required
def listar_logs():
    """Listar histórico de ações (admin)."""
    try:
        pagina, por_pagina = get_pagination_params()
        entidade = request.args.get("entidade")
        acao = request.args.get("acao")

        query = LogAcao.query.order_by(LogAcao.data_hora.desc())

        if entidade:
            query = query.filter(LogAcao.entidade == entidade)
        if acao:
            query = query.filter(LogAcao.acao == acao)

        total = query.count()
        logs = query.offset((pagina - 1) * por_pagina).limit(por_pagina).all()

        return jsonify(
            {
                "logs": [entry.to_dict() for entry in logs],
                "total": total,
                "pagina": pagina,
                "por_pagina": por_pagina,
                "total_paginas": (total + por_pagina - 1) // por_pagina,
            }
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/logs/export-csv", methods=["GET"])
@limiter.limit("10 per minute")
@api_admin_required
def exportar_logs_csv():
    """Exportar audit log como CSV (admin, max 10.000 registros)."""
    try:
        entidade = request.args.get("entidade")
        acao = request.args.get("acao")
        data_inicio = request.args.get("data_inicio")
        data_fim = request.args.get("data_fim")

        query = LogAcao.query.order_by(LogAcao.data_hora.desc())
        if entidade:
            query = query.filter(LogAcao.entidade == entidade)
        if acao:
            query = query.filter(LogAcao.acao == acao)
        if data_inicio:
            try:
                dt = datetime.strptime(data_inicio, "%Y-%m-%d").date()
                query = query.filter(
                    LogAcao.data_hora >= _dia_inicio(dt)
                )
            except ValueError:
                pass
        if data_fim:
            try:
                dt = datetime.strptime(data_fim, "%Y-%m-%d").date()
                query = query.filter(
                    LogAcao.data_hora <= _dia_fim(dt)
                )
            except ValueError:
                pass

        logs = query.limit(10000).all()

        import io
        import csv as _csv
        buf = io.StringIO()
        writer = _csv.writer(buf)
        writer.writerow([
            "id", "data_hora", "acao", "entidade",
            "id_entidade", "id_usuario", "detalhes",
        ])
        for entry in logs:
            writer.writerow([
                entry.id_log,
                entry.data_hora.isoformat() if entry.data_hora else "",
                entry.acao or "",
                entry.entidade or "",
                entry.id_entidade or "",
                entry.id_usuario or "",
                (entry.detalhes or "").replace("\n", " "),
            ])
        output = buf.getvalue()
        buf.close()

        return app.response_class(
            output,
            mimetype="text/csv",
            headers={
                "Content-Disposition":
                    "attachment; filename=audit_log.csv"
            },
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/historico")
@admin_required
def pagina_historico():
    """Página de histórico de ações (admin only)"""
    return render_template("historico.html")


# =============================================================================
# ROTAS - VENDAS
# =============================================================================


@app.route("/api/vendas", methods=["GET"])
@limiter.limit("120 per minute")
@api_login_required
def listar_vendas():
    """Listar vendas com filtros e paginação."""
    try:
        query = Venda.query

        # Filtro por intervalo de datas
        data_inicio = request.args.get("data_inicio", "").strip()
        data_fim = request.args.get("data_fim", "").strip()
        if data_inicio:
            try:
                dt_ini = datetime.strptime(data_inicio, "%Y-%m-%d")
                query = query.filter(Venda.data_venda >= dt_ini)
            except ValueError:
                pass
        if data_fim:
            try:
                dt_fim = datetime.strptime(data_fim, "%Y-%m-%d") + timedelta(
                    days=1
                )
                query = query.filter(Venda.data_venda < dt_fim)
            except ValueError:
                pass

        # Filtro por cliente
        id_cliente = request.args.get("id_cliente", type=int)
        if id_cliente:
            query = query.filter_by(id_cliente=id_cliente)

        # Filtro por forma de pagamento
        forma = request.args.get("forma_pagamento", "").strip()
        if forma:
            query = query.filter(Venda.forma_pagamento.ilike(f"%{forma}%"))

        # Paginação — count ANTES do joinedload (evita JOIN overhead)
        pagina, por_pagina = get_pagination_params()
        total = query.count()

        # Eager-load: evita N+1 ao serializar
        query = query.options(
            joinedload(Venda.cliente),
            joinedload(Venda.itens).joinedload(ItemVenda.produto),
            joinedload(Venda.itens).joinedload(ItemVenda.complementos)
            .joinedload(ItemVendaComplemento.complemento),
        )

        vendas = (
            query.order_by(Venda.data_venda.desc())
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )

        return jsonify(
            {
                "vendas": [venda.to_dict() for venda in vendas],
                "total": total,
                "pagina": pagina,
                "por_pagina": por_pagina,
                "total_paginas": (total + por_pagina - 1) // por_pagina,
            }
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/vendas", methods=["POST"])
@limiter.limit("40 per minute")
@api_login_required
def criar_venda():
    """Criar nova venda"""
    try:
        dados = validar_payload(VendaCreateSchema)

        # Validação complementar de negócio
        if not dados.get("itens") or len(dados["itens"]) == 0:
            return jsonify({"erro": "Venda deve ter pelo menos um item"}), 400

        # Verificar se cliente existe
        cliente = db.session.get(Cliente, dados["id_cliente"])
        if not cliente:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        # LGPD: bloquear venda sem consentimento
        if not cliente.consentimento_lgpd:
            return (
                jsonify(
                    {
                        "erro": (
                            "Venda não permitida:"
                            " cliente sem consentimento LGPD"
                        )
                    }
                ),
                400,
            )

        # Calcular valor total
        valor_total = Decimal("0.00")

        # Criar venda
        venda = Venda(
            id_cliente=dados["id_cliente"],
            forma_pagamento=dados.get("forma_pagamento", "Dinheiro"),
            status_pagamento="Pendente",
            observacoes=dados.get("observacoes"),
        )

        # Adicionar itens
        for item_dados in dados["itens"]:
            # Lock pessimista para evitar race condition no estoque
            produto = (
                db.session.query(Produto)
                .filter_by(id_produto=item_dados["id_produto"])
                .with_for_update()
                .first()
            )
            if not produto:
                return (
                    jsonify(
                        {
                            "erro": (
                                "Produto"
                                f' {item_dados["id_produto"]}'
                                " não encontrado"
                            )
                        }
                    ),
                    404,
                )

            # Validar produto ativo
            if not produto.ativo:
                return (
                    jsonify(
                        {
                            "erro": (
                                f'Produto "{produto.nome_produto}"'
                                " está desativado e não pode"
                                " ser vendido"
                            )
                        }
                    ),
                    400,
                )

            quantidade = int(item_dados["quantidade"])
            if quantidade < 1 or quantidade > 9999:
                return (
                    jsonify({"erro": f"Quantidade inválida: {quantidade}. Deve ser entre 1 e 9999"}),
                    400,
                )

            # Verificar estoque somente se controle ativo
            # (estoque ou mínimo > 0)
            controle_ativo = (produto.estoque_atual or 0) > 0 or (
                produto.estoque_minimo or 0
            ) > 0
            if controle_ativo and produto.estoque_atual < quantidade:
                return (
                    jsonify(
                        {
                            "erro": (
                                f'Estoque insuficiente para'
                                f' "{produto.nome_produto}":'
                                f" disponível {produto.estoque_atual},"
                                f" solicitado {quantidade}"
                            )
                        }
                    ),
                    400,
                )

            preco_unitario = Decimal(str(produto.preco))
            subtotal = preco_unitario * quantidade

            item = ItemVenda(
                id_produto=item_dados["id_produto"],
                quantidade=quantidade,
                preco_unitario=preco_unitario,
                subtotal=subtotal,
            )

            # Processar complementos/toppings do item
            ids_complementos = item_dados.get("complementos") or []
            for id_comp in ids_complementos:
                comp = db.session.get(Complemento, id_comp)
                if comp and comp.ativo:
                    preco_comp = Decimal(
                        str(comp.preco_adicional or 0)
                    )
                    item.complementos.append(
                        ItemVendaComplemento(
                            id_complemento=id_comp,
                            preco_unitario=preco_comp,
                        )
                    )
                    subtotal += preco_comp * quantidade
                    item.subtotal = subtotal

            venda.itens.append(item)
            valor_total += subtotal

            # Descontar estoque atomicamente (row já está locked)
            if controle_ativo:
                produto.estoque_atual = max(
                    0, produto.estoque_atual - quantidade
                )

        # Aplicar desconto e taxa
        desconto_perc = Decimal(str(dados.get("desconto_percentual", 0)))
        taxa = Decimal(str(dados.get("taxa", 0)))
        desconto_valor = valor_total * desconto_perc / Decimal("100")

        # Aplicar cupom de desconto (se informado)
        cupom_codigo = (dados.get("cupom_codigo") or "").strip().upper()
        cupom_obj = None
        if cupom_codigo:
            cupom_obj = CupomDesconto.query.filter_by(
                codigo=cupom_codigo
            ).first()
            if cupom_obj and cupom_obj.valido:
                if cupom_obj.tipo_desconto == "percentual":
                    desconto_valor += (
                        valor_total
                        * cupom_obj.valor_desconto
                        / Decimal("100")
                    )
                else:
                    desconto_valor += min(
                        cupom_obj.valor_desconto, valor_total
                    )
                cupom_obj.usos_realizados = (cupom_obj.usos_realizados or 0) + 1

        valor_total = max(
            valor_total - desconto_valor + taxa, Decimal("0.00")
        ).quantize(Decimal("0.01"))

        venda.valor_total = valor_total
        venda.desconto_aplicado = desconto_valor.quantize(Decimal("0.01"))

        # Criar pagamento
        pagamento = Pagamento(
            valor_pago=valor_total,
            metodo=dados.get("forma_pagamento", "Dinheiro"),
            status="Concluído",
        )
        venda.pagamento = pagamento
        venda.status_pagamento = "Concluído"

        # Acumular pontos de fidelidade (1 ponto por R$1 gasto, cap 999999)
        pontos_ganhos = round(float(valor_total))
        if pontos_ganhos > 0:
            cliente.pontos_fidelidade = min(
                999999, (cliente.pontos_fidelidade or 0) + pontos_ganhos
            )

        db.session.add(venda)
        db.session.commit()

        # Invalidar cache vitrine (estoque atualizado)
        _invalidar_cache_vitrine()

        # Verificar badges de gamificação
        novos_badges = []
        try:
            novos_badges = _verificar_badges(dados["id_cliente"])
        except Exception as e:
            logger.warning("Erro ao verificar badges do cliente %s: %s", dados["id_cliente"], e)

        registrar_log(
            "criar",
            "venda",
            venda.id_venda,
            f"Venda #{venda.id_venda} - R${float(venda.valor_total):.2f}",
        )

        resultado = venda.to_dict()
        resultado["pontos_ganhos"] = pontos_ganhos
        resultado["pontos_total"] = cliente.pontos_fidelidade or 0
        resultado["novos_badges"] = [b.to_dict() for b in novos_badges]
        return jsonify(resultado), 201
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return (
                jsonify({"erro": "Payload invalido", "detalhes": str(e)}),
                400,
            )
        return _erro_interno(e)


@app.route("/api/vendas/<int:id_venda>", methods=["GET"])
@api_login_required
def obter_venda(id_venda):
    """Obter detalhes de uma venda"""
    try:
        venda = db.session.get(Venda, id_venda)
        if not venda:
            return jsonify({"erro": "Venda não encontrada"}), 404

        return jsonify(venda.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/vendas/<int:id_venda>/cancelar", methods=["POST"])
@limiter.limit("10 per minute")
@api_admin_required
def cancelar_venda(id_venda):
    """Cancelar (estornar) uma venda — somente admin.
    Restaura estoque e remove pontos de fidelidade do cliente."""
    try:
        venda = db.session.get(Venda, id_venda)
        if not venda:
            return jsonify({"erro": "Venda não encontrada"}), 404

        if venda.status_pagamento == "Cancelado":
            return (
                jsonify({"erro": "Venda já foi cancelada anteriormente"}),
                400,
            )

        dados = request.get_json(silent=True) or {}
        motivo = (dados.get("motivo") or "").strip()
        if not motivo or len(motivo) < 3:
            return (
                jsonify(
                    {
                        "erro": (
                            "Motivo do cancelamento é"
                            " obrigatório (mínimo 3 caracteres)"
                        )
                    }
                ),
                400,
            )

        # Restaurar estoque dos itens
        for item in venda.itens:
            prod = db.session.get(Produto, item.id_produto)
            if prod:
                prod.estoque_atual = (
                    prod.estoque_atual or 0
                ) + item.quantidade

        # Remover pontos de fidelidade concedidos
        pontos_remover = int(venda.valor_total)
        cliente = db.session.get(Cliente, venda.id_cliente)
        if cliente and pontos_remover > 0:
            cliente.pontos_fidelidade = max(
                0, (cliente.pontos_fidelidade or 0) - pontos_remover
            )

        # Atualizar status da venda
        venda.status_pagamento = "Cancelado"
        venda.status_pedido = "Cancelado"
        venda.motivo_cancelamento = motivo
        venda.observacoes = (
            f'{venda.observacoes or ""}\n[CANCELADO] {motivo}'.strip()
        )

        # Atualizar pagamento
        if venda.pagamento:
            venda.pagamento.status = "Estornado"

        db.session.commit()
        registrar_log(
            "cancelar",
            "venda",
            id_venda,
            f"Venda #{id_venda} cancelada"
            f" — R${float(venda.valor_total):.2f}"
            f" — Motivo: {motivo}",
        )

        return jsonify(
            {
                "mensagem": f"Venda #{id_venda} cancelada com sucesso",
                "estoque_restaurado": True,
                "pontos_removidos": pontos_remover,
                "venda": venda.to_dict(),
            }
        )
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - FIDELIDADE
# =============================================================================


@app.route("/api/clientes/<int:id_cliente>/pontos", methods=["GET"])
@api_login_required
def obter_pontos_fidelidade(id_cliente):
    """Consultar pontos de fidelidade de um cliente"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente or not cliente.ativo:
            return jsonify({"erro": "Cliente não encontrado"}), 404
        return jsonify(
            {
                "id_cliente": cliente.id_cliente,
                "nome": cliente.nome,
                "pontos": cliente.pontos_fidelidade or 0,
            }
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/clientes/<int:id_cliente>/pontos/resgatar", methods=["POST"])
@limiter.limit("30 per minute")
@api_login_required
def resgatar_pontos(id_cliente):
    """Resgatar pontos de fidelidade (cada 100 pontos = R$5 de desconto)"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente or not cliente.ativo:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        dados = request.get_json(silent=True) or {}
        pontos_resgatar = int(dados.get("pontos", 0))

        if pontos_resgatar <= 0:
            return (
                jsonify({"erro": "Quantidade de pontos deve ser positiva"}),
                400,
            )

        pontos_disponiveis = cliente.pontos_fidelidade or 0
        if pontos_resgatar > pontos_disponiveis:
            return (
                jsonify(
                    {
                        "erro": (
                            "Pontos insuficientes."
                            f" Disponível: {pontos_disponiveis}"
                        )
                    }
                ),
                400,
            )

        # Regra: cada 100 pontos = R$5.00 de desconto
        if pontos_resgatar < 100:
            return jsonify({"erro": "Mínimo de 100 pontos para resgate"}), 400

        desconto = Decimal(str((pontos_resgatar // 100) * 5))
        pontos_usados = (pontos_resgatar // 100) * 100  # usa múltiplos de 100

        cliente.pontos_fidelidade = pontos_disponiveis - pontos_usados
        db.session.commit()
        registrar_log(
            "resgatar",
            "fidelidade",
            cliente.id_cliente,
            f"{pontos_usados} pontos resgatados"
            f" → R${float(desconto):.2f} desconto",
        )

        return jsonify(
            {
                "id_cliente": cliente.id_cliente,
                "pontos_resgatados": pontos_usados,
                "desconto_gerado": float(desconto),
                "pontos_restantes": cliente.pontos_fidelidade,
            }
        )
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/fidelidade/ranking", methods=["GET"])
@api_login_required
def ranking_fidelidade():
    """Top 10 clientes com mais pontos"""
    try:
        clientes = (
            Cliente.query.filter(
                Cliente.ativo == True,  # noqa: E712
                Cliente.pontos_fidelidade > 0,
            )
            .order_by(Cliente.pontos_fidelidade.desc())
            .limit(10)
            .all()
        )

        return jsonify(
            [
                {
                    "id_cliente": c.id_cliente,
                    "nome": c.nome,
                    "pontos": c.pontos_fidelidade or 0,
                }
                for c in clientes
            ]
        )
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - BUSCA GLOBAL
# =============================================================================


@app.route("/api/busca", methods=["GET"])
@limiter.limit("60 per minute")
@api_login_required
def busca_global():
    """Busca global: pesquisa clientes e produtos simultaneamente"""
    try:
        q = request.args.get("q", "").strip()
        if len(q) < 2:
            return jsonify({"clientes": [], "produtos": []})

        filtro = f"%{q}%"

        clientes = (
            Cliente.query.filter(
                Cliente.ativo == True,  # noqa: E712
                db.or_(
                    Cliente.nome.ilike(filtro),
                    Cliente.telefone.ilike(filtro),
                    Cliente.email.ilike(filtro),
                ),
            )
            .order_by(Cliente.nome)
            .limit(5)
            .all()
        )

        produtos = (
            Produto.query.filter(
                Produto.ativo == True,  # noqa: E712
                db.or_(
                    Produto.nome_produto.ilike(filtro),
                    Produto.categoria.ilike(filtro),
                ),
            )
            .order_by(Produto.nome_produto)
            .limit(5)
            .all()
        )

        return jsonify(
            {
                "clientes": [
                    {
                        "id": c.id_cliente,
                        "nome": c.nome,
                        "telefone": c.telefone,
                        "email": c.email,
                    }
                    for c in clientes
                ],
                "produtos": [
                    {
                        "id": p.id_produto,
                        "nome": p.nome_produto,
                        "preco": float(p.preco),
                        "categoria": p.categoria,
                    }
                    for p in produtos
                ],
            }
        )
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - DASHBOARD GRÁFICOS
# =============================================================================


@app.route("/api/dashboard/graficos", methods=["GET"])
@api_login_required
def dashboard_graficos():
    """Dados para gráficos do dashboard."""
    try:
        hoje = datetime.now(timezone.utc).date()
        inicio = hoje - timedelta(days=6)

        # Vendas por dia (últimos 7 dias)
        vendas_dia = (
            db.session.query(
                db.func.date(Venda.data_venda).label("dia"),
                db.func.count(Venda.id_venda).label("qtd"),
                db.func.coalesce(db.func.sum(Venda.valor_total), 0).label(
                    "total"
                ),
            )
            .filter(Venda.data_venda >= _dia_inicio(inicio))
            .group_by(db.func.date(Venda.data_venda))
            .order_by(db.func.date(Venda.data_venda))
            .all()
        )

        # Montar dict dia → dados (preencher dias sem vendas com 0)
        mapa = {
            str(r.dia): {"qtd": r.qtd, "total": float(r.total)}
            for r in vendas_dia
        }
        dias = []
        for i in range(7):
            d = inicio + timedelta(days=i)
            ds = str(d)
            dias.append(
                {
                    "data": ds,
                    "label": d.strftime("%d/%m"),
                    "quantidade": mapa.get(ds, {}).get("qtd", 0),
                    "faturamento": mapa.get(ds, {}).get("total", 0),
                }
            )

        # Vendas por forma de pagamento
        pagamentos = (
            db.session.query(
                Pagamento.metodo,
                db.func.count(Pagamento.id_pagamento).label("qtd"),
                db.func.coalesce(db.func.sum(Pagamento.valor_pago), 0).label(
                    "total"
                ),
            )
            .group_by(Pagamento.metodo)
            .all()
        )

        por_pagamento = [
            {"forma": p.metodo, "quantidade": p.qtd, "total": float(p.total)}
            for p in pagamentos
        ]

        # Top 5 produtos mais vendidos
        top_produtos = (
            db.session.query(
                Produto.nome_produto,
                db.func.sum(ItemVenda.quantidade).label("qtd"),
            )
            .join(ItemVenda)
            .filter(Produto.ativo == True)  # noqa: E712
            .group_by(Produto.nome_produto)
            .order_by(db.func.sum(ItemVenda.quantidade).desc())
            .limit(5)
            .all()
        )

        produtos_ranking = [
            {"produto": p.nome_produto, "quantidade": int(p.qtd)}
            for p in top_produtos
        ]

        return jsonify(
            {
                "vendas_por_dia": dias,
                "por_forma_pagamento": por_pagamento,
                "top_produtos": produtos_ranking,
            }
        )
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - RELATÓRIOS
# =============================================================================


@app.route("/api/relatorios/dia-atual", methods=["GET"])
@api_login_required
def relatorio_dia_atual():
    """Relatório de vendas do dia atual"""
    try:
        hoje = datetime.now(timezone.utc).date()

        vendas_hoje = Venda.query.filter(
            Venda.data_venda >= _dia_inicio(hoje),
            Venda.data_venda <= _dia_fim(hoje),
        ).all()

        total_vendas = len(vendas_hoje)
        faturamento = sum(v.valor_total for v in vendas_hoje)

        # Por forma de pagamento
        por_forma = {}
        for venda in vendas_hoje:
            forma = venda.forma_pagamento or "Indefinido"
            if forma not in por_forma:
                por_forma[forma] = Decimal("0.00")
            por_forma[forma] += venda.valor_total

        return jsonify(
            {
                "data": hoje.isoformat(),
                "total_vendas": total_vendas,
                "faturamento_total": float(faturamento),
                "por_forma_pagamento": {
                    k: float(v) for k, v in por_forma.items()
                },
                "ticket_medio": (
                    float(faturamento / total_vendas)
                    if total_vendas > 0
                    else 0
                ),
            }
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/relatorios/por-data", methods=["GET"])
@api_login_required
def relatorio_por_data():
    """Relatório de vendas filtrado por data (YYYY-MM-DD)"""
    try:
        data_str = request.args.get("data", "")
        if not data_str:
            return (
                jsonify(
                    {
                        "erro": (
                            'Parâmetro "data" é obrigatório'
                            " (formato YYYY-MM-DD)"
                        )
                    }
                ),
                400,
            )

        from datetime import date as date_type

        try:
            data_filtro = date_type.fromisoformat(data_str)
        except ValueError:
            return (
                jsonify({"erro": "Formato de data inválido. Use YYYY-MM-DD"}),
                400,
            )

        vendas_dia = Venda.query.options(
            joinedload(Venda.cliente),
            joinedload(Venda.itens).joinedload(ItemVenda.produto),
        ).filter(
            Venda.data_venda >= _dia_inicio(data_filtro),
            Venda.data_venda <= _dia_fim(data_filtro),
        ).all()

        total_vendas = len(vendas_dia)
        faturamento = sum(v.valor_total for v in vendas_dia)

        canceladas = sum(
            1 for v in vendas_dia
            if (v.status_pagamento or "") == "Cancelado"
        )
        total_descontos = sum(
            float(v.desconto_aplicado or 0) for v in vendas_dia
        )
        vendas_ativas = [
            v for v in vendas_dia
            if (v.status_pagamento or "") != "Cancelado"
        ]
        faturamento_liquido = sum(v.valor_total for v in vendas_ativas)

        por_forma = {}
        for venda in vendas_dia:
            forma = venda.forma_pagamento or "Indefinido"
            if forma not in por_forma:
                por_forma[forma] = {"quantidade": 0, "total": Decimal("0.00")}
            por_forma[forma]["quantidade"] += 1
            por_forma[forma]["total"] += venda.valor_total

        return jsonify(
            {
                "data": data_filtro.isoformat(),
                "total_vendas": total_vendas,
                "faturamento_total": float(faturamento),
                "canceladas": canceladas,
                "total_descontos": float(total_descontos),
                "faturamento_liquido": float(faturamento_liquido),
                "por_forma_pagamento": {
                    k: {
                        "quantidade": v["quantidade"],
                        "total": float(v["total"]),
                    }
                    for k, v in por_forma.items()
                },
                "ticket_medio": (
                    float(faturamento / total_vendas)
                    if total_vendas > 0
                    else 0
                ),
                "vendas": [v.to_dict() for v in vendas_dia],
            }
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/relatorios/clientes-frequentes", methods=["GET"])
@api_login_required
def relatorio_clientes_frequentes():
    """Clientes mais frequentes (últimos 30 dias) — com paginação"""
    try:
        dias = request.args.get("dias", 30, type=int)
        dias = min(max(dias, 1), 365)
        limite = request.args.get("limite", 10, type=int)
        limite = min(max(limite, 1), 100)
        pagina = request.args.get("pagina", 1, type=int)
        pagina = max(pagina, 1)

        data_limite = datetime.now(timezone.utc) - timedelta(days=dias)

        base_query = (
            db.session.query(
                Cliente.id_cliente,
                Cliente.nome,
                Cliente.telefone,
                db.func.count(Venda.id_venda).label("total_compras"),
                db.func.sum(Venda.valor_total).label("faturamento"),
                db.func.max(Venda.data_venda).label("ultima_compra"),
            )
            .join(Venda)
            .filter(
                Venda.data_venda >= data_limite,
                Cliente.ativo == True,  # noqa: E712
            )
            .group_by(Cliente.id_cliente, Cliente.nome, Cliente.telefone)
            .order_by(db.func.count(Venda.id_venda).desc())
        )

        total = base_query.count()
        clientes_freq = (
            base_query.offset((pagina - 1) * limite).limit(limite).all()
        )

        resultado = []
        for cliente in clientes_freq:
            resultado.append(
                {
                    "id_cliente": cliente.id_cliente,
                    "nome": cliente.nome,
                    "telefone": cliente.telefone,
                    "total_compras": cliente.total_compras,
                    "faturamento": float(cliente.faturamento),
                    "ultima_compra": (
                        cliente.ultima_compra.isoformat()
                        if cliente.ultima_compra
                        else None
                    ),
                }
            )

        return jsonify(
            {
                "dados": resultado,
                "total": total,
                "pagina": pagina,
                "limite": limite,
                "total_paginas": (
                    (total + limite - 1) // limite if total else 0
                ),
            }
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/relatorios/produtos-ranking", methods=["GET"])
@api_login_required
def relatorio_produtos_ranking():
    """Produtos mais vendidos — com paginação"""
    try:
        limite = request.args.get("limite", 15, type=int)
        limite = min(max(limite, 1), 100)
        pagina = request.args.get("pagina", 1, type=int)
        pagina = max(pagina, 1)

        base_query = (
            db.session.query(
                Produto.id_produto,
                Produto.nome_produto,
                db.func.count(ItemVenda.id_item).label("quantidade"),
                db.func.sum(ItemVenda.subtotal).label("faturamento"),
            )
            .join(ItemVenda)
            .filter(Produto.ativo == True)  # noqa: E712
            .group_by(Produto.id_produto, Produto.nome_produto)
            .order_by(db.func.count(ItemVenda.id_item).desc())
        )

        total = base_query.count()
        produtos_rank = (
            base_query.offset((pagina - 1) * limite).limit(limite).all()
        )

        resultado = []
        for produto in produtos_rank:
            resultado.append(
                {
                    "id_produto": produto.id_produto,
                    "nome_produto": produto.nome_produto,
                    "quantidade_vendida": produto.quantidade,
                    "faturamento": float(produto.faturamento),
                }
            )

        return jsonify(
            {
                "dados": resultado,
                "total": total,
                "pagina": pagina,
                "limite": limite,
                "total_paginas": (
                    (total + limite - 1) // limite if total else 0
                ),
            }
        )
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - EXPORTAÇÃO
# =============================================================================


@app.route("/api/exportar/clientes-csv", methods=["GET"])
@api_login_required
def exportar_clientes_csv():
    """Exportar lista de clientes em CSV"""

    def sanitize_csv(value):
        """Previne CSV formula injection (=, +, -, @, |, %, tab, CR)"""
        if isinstance(value, str) and value:
            stripped = value.strip()
            if stripped and stripped[0] in ("=", "+", "-", "@", "|", "%", "\t", "\r"):
                return "'" + value
        return value

    try:
        clientes = Cliente.query.filter_by(
            ativo=True, consentimento_lgpd=True
        ).all()

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Nome", "Telefone", "Email", "Data de Cadastro"])

        for cliente in clientes:
            writer.writerow(
                [
                    sanitize_csv(cliente.nome),
                    sanitize_csv(cliente.telefone or ""),
                    sanitize_csv(cliente.email or ""),
                    (
                        cliente.data_cadastro.strftime("%Y-%m-%d")
                        if cliente.data_cadastro
                        else ""
                    ),
                ]
            )

        output.seek(0)
        bytes_output = io.BytesIO(output.getvalue().encode("utf-8"))
        bytes_output.seek(0)

        _ts = datetime.now(timezone.utc).strftime(
            "%Y%m%d_%H%M%S"
        )
        return send_file(
            bytes_output,
            mimetype="text/csv",
            as_attachment=True,
            download_name=f"clientes_export_{_ts}.csv",
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/exportar/relatorio-pdf", methods=["GET"])
@api_login_required
def exportar_relatorio_pdf():
    """Gera PDF com relatório de vendas do dia (ou data informada)"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    try:
        data_str = request.args.get(
            "data", datetime.now(timezone.utc).strftime("%Y-%m-%d")
        )
        try:
            data_ref = datetime.strptime(data_str, "%Y-%m-%d").date()
        except ValueError:
            return (
                jsonify({"erro": "Data inválida. Use formato YYYY-MM-DD"}),
                400,
            )

        # Buscar vendas do dia
        vendas = (
            Venda.query.filter(
                Venda.data_venda >= _dia_inicio(data_ref),
                Venda.data_venda <= _dia_fim(data_ref),
            )
            .order_by(Venda.data_venda)
            .all()
        )

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm
        )
        styles = getSampleStyleSheet()
        elements = []

        # Título
        elements.append(
            Paragraph(
                "<b>Combina Açaí — Relatório de Vendas</b>",
                styles["Title"],
            )
        )
        elements.append(
            Paragraph(
                f'Data: {data_ref.strftime("%d/%m/%Y")}', styles["Normal"]
            )
        )
        elements.append(Spacer(1, 0.5 * cm))

        if vendas:
            # Tabela de vendas
            header = ["#", "Cliente", "Itens", "Forma Pgto", "Valor (R$)"]
            rows = [header]
            total_geral = 0
            for v in vendas:
                cliente_nome = v.cliente.nome if v.cliente else "—"
                qtd_itens = (
                    sum(i.quantidade for i in v.itens) if v.itens else 0
                )
                rows.append(
                    [
                        str(v.id_venda),
                        cliente_nome[:25],
                        str(qtd_itens),
                        v.forma_pagamento or "—",
                        f"{float(v.valor_total):.2f}",
                    ]
                )
                total_geral += float(v.valor_total)

            rows.append(["", "", "", "TOTAL", f"{total_geral:.2f}"])

            t = Table(rows, repeatRows=1)
            t.setStyle(
                TableStyle(
                    [
                        (
                            "BACKGROUND",
                            (0, 0),
                            (-1, 0),
                            colors.HexColor("#7B1FA2"),
                        ),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ALIGN", (0, 0), (0, -1), "CENTER"),
                        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
                        (
                            "BACKGROUND",
                            (0, -1),
                            (-1, -1),
                            colors.HexColor("#F3E5F5"),
                        ),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ]
                )
            )
            elements.append(t)
        else:
            elements.append(
                Paragraph(
                    "Nenhuma venda registrada nesta data.", styles["Normal"]
                )
            )

        elements.append(Spacer(1, 1 * cm))
        _now_str = datetime.now(timezone.utc).strftime(
            "%d/%m/%Y %H:%M UTC"
        )
        elements.append(
            Paragraph(
                f"Gerado em {_now_str}"
                " — CRM Açaiteria",
                styles["Italic"],
            )
        )

        doc.build(elements)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f'relatorio_{data_ref.strftime("%Y%m%d")}.pdf',
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/exportar/clientes-xlsx", methods=["GET"])
@api_login_required
def exportar_clientes_xlsx():
    """Exporta clientes com consentimento LGPD em formato Excel (.xlsx)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        clientes = (
            Cliente.query.filter_by(ativo=True, consentimento_lgpd=True)
            .order_by(Cliente.nome)
            .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="7B1FA2", end_color="7B1FA2", fill_type="solid"
        )
        headers = [
            "ID",
            "Nome",
            "Telefone",
            "Email",
            "Pontos",
            "Data Cadastro",
            "Consentimento LGPD",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, c in enumerate(clientes, 2):
            ws.cell(row=i, column=1, value=c.id_cliente)
            ws.cell(row=i, column=2, value=c.nome)
            ws.cell(row=i, column=3, value=c.telefone or "")
            ws.cell(row=i, column=4, value=c.email or "")
            ws.cell(
                row=i, column=5, value=getattr(c, "pontos_fidelidade", 0) or 0
            )
            ws.cell(
                row=i,
                column=6,
                value=(
                    c.data_cadastro.strftime("%d/%m/%Y")
                    if c.data_cadastro
                    else ""
                ),
            )
            ws.cell(
                row=i, column=7, value="Sim" if c.consentimento_lgpd else "Não"
            )

        # Auto-width columns
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col), default=10
            )
            ws.column_dimensions[col[0].column_letter].width = min(
                max_len + 3, 40
            )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        _ts = datetime.now(timezone.utc).strftime(
            "%Y%m%d_%H%M%S"
        )
        return send_file(
            buf,
            mimetype=(
                "application/vnd.openxmlformats"
                "-officedocument.spreadsheetml.sheet"
            ),
            as_attachment=True,
            download_name=f"clientes_{_ts}.xlsx",
        )
    except ImportError:
        return (
            jsonify(
                {
                    "erro": (
                        "openpyxl não instalado."
                        " Execute: pip install openpyxl"
                    )
                }
            ),
            500,
        )
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# PÁGINAS HTML
# =============================================================================


@app.route("/cadastro-cliente")
@login_required
def pagina_cadastro_cliente():
    """Página de cadastro de cliente"""
    return render_template("cadastro_cliente.html")


@app.route("/nova-venda")
@login_required
def pagina_nova_venda():
    """Página de registro de venda"""
    return render_template("venda.html")


@app.route("/vendas")
@login_required
def pagina_vendas():
    """Página de gerenciamento de vendas e pedidos"""
    return render_template("vendas_lista.html")


@app.route("/relatorios")
@login_required
def pagina_relatorios():
    """Página de relatórios"""
    return render_template("relatorios.html")


@app.route("/clientes")
@login_required
def pagina_clientes():
    """Página de gerenciamento de clientes"""
    return render_template("clientes.html")


@app.route("/produtos")
@login_required
def pagina_produtos():
    """Página de gerenciamento de produtos"""
    return render_template("produtos.html")


@app.route("/fechamento")
@login_required
def pagina_fechamento():
    """Página de fechamento diário"""
    return render_template("fechamento.html")


@app.route("/financeiro")
@login_required
def pagina_financeiro():
    """Página de gestão financeira (receitas/despesas)"""
    return render_template("financeiro.html")


@app.route("/sobre")
def sobre_projeto():
    """Página sobre o Projeto Integrador UNIVESP."""
    return render_template("sobre.html")


@app.route("/politica-privacidade")
def politica_privacidade():
    """Página com política de privacidade LGPD"""
    return render_template("politica_privacidade.html")


@app.route("/usuarios")
@admin_required
def pagina_usuarios():
    """Página de gerenciamento de usuários (admin only)"""
    return render_template("usuarios.html")


# =============================================================================
# TOTEM — Cadastro público de clientes (sem login)
# =============================================================================


@app.route("/totem")
def pagina_totem():
    """Totem de auto-cadastro — página pública para tablet no balcão"""
    return render_template("totem_cliente.html")


@app.route("/api/totem/cadastro", methods=["POST"])
@limiter.limit("10 per minute")
def totem_cadastrar_cliente():
    """API pública para auto-cadastro via totem (rate-limited)"""
    try:
        dados = request.get_json(silent=True) or {}

        nome = (dados.get("nome") or "").strip()
        if not nome or len(nome) < 2:
            return (
                jsonify({"erro": "Nome é obrigatório (mínimo 2 caracteres)"}),
                400,
            )

        consentimento = dados.get("consentimento_lgpd", False)
        if not consentimento:
            return jsonify({"erro": "Consentimento LGPD é obrigatório"}), 400

        telefone = (dados.get("telefone") or "").strip() or None
        email = (dados.get("email") or "").strip().lower() or None
        observacoes = (dados.get("observacoes") or "").strip() or None

        # Validar email se fornecido
        if email and not _email_valido(email):
            return jsonify({"erro": "E-mail inválido"}), 400

        # Verificar duplicidade por telefone ou email
        if telefone:
            existente = Cliente.query.filter_by(
                telefone=telefone, ativo=True
            ).first()
            if existente:
                return (
                    jsonify(
                        {
                            "erro": (
                                "Este telefone já está cadastrado."
                                " Fale com o atendente."
                            )
                        }
                    ),
                    409,
                )
        if email:
            existente = Cliente.query.filter_by(
                email=email, ativo=True
            ).first()
            if existente:
                return (
                    jsonify(
                        {
                            "erro": (
                                "Este e-mail já está cadastrado."
                                " Fale com o atendente."
                            )
                        }
                    ),
                    409,
                )

        cliente = Cliente(
            nome=nome,
            telefone=telefone,
            email=email,
            observacoes=observacoes,
            consentimento_lgpd=True,
            data_consentimento=datetime.now(timezone.utc),
            consentimento_versao=dados.get("versao_politica", "v1.0"),
            ativo=True,
        )

        # Senha opcional no totem (cliente pode criar depois)
        senha = (dados.get("senha") or "").strip()
        if senha and len(senha) >= 6:
            cliente.set_senha(senha)

        db.session.add(cliente)
        db.session.flush()

        # Registrar histórico LGPD
        entrada = ConsentimentoHistorico(
            id_cliente=cliente.id_cliente,
            acao="concedeu",
            versao_politica=dados.get("versao_politica", "v1.0"),
            ip_address=request.remote_addr,
            user_agent=request.headers.get("User-Agent", "")[:255],
        )
        db.session.add(entrada)

        # Bônus de boas-vindas: +10 pontos
        cliente.pontos_fidelidade = 10

        db.session.commit()
        registrar_log(
            "criar",
            "cliente",
            cliente.id_cliente,
            f"Auto-cadastro via totem: {cliente.nome}",
        )

        return (
            jsonify(
                {
                    "id_cliente": cliente.id_cliente,
                    "nome": cliente.nome,
                    "pontos_fidelidade": cliente.pontos_fidelidade,
                    "mensagem": "Cadastro realizado com sucesso!",
                }
            ),
            201,
        )

    except Exception as e:
        db.session.rollback()
        logger.exception("Erro no auto-cadastro totem: %s", e)
        return jsonify({"erro": "Erro interno. Tente novamente."}), 500


# =============================================================================
# VITRINE PÚBLICA — Visão do cliente (sem login obrigatório)
# =============================================================================


@app.route("/vitrine")
def vitrine():
    """Página pública — vitrine da açaiteria para clientes."""
    return render_template("vitrine.html")


@app.route("/api/vitrine/produtos", methods=["GET"])
@limiter.limit("60 per minute")
@cache.cached(timeout=120, query_string=True, key_prefix="vitrine_produtos")
def vitrine_produtos():
    """API pública — lista produtos ativos para a vitrine."""
    try:
        categoria = request.args.get("categoria", "").strip()
        query = Produto.query.filter_by(ativo=True)
        if categoria:
            query = query.filter(Produto.categoria.ilike(f"%{categoria}%"))
        produtos = query.order_by(Produto.nome_produto).limit(500).all()
        return jsonify([
            {
                "id_produto": p.id_produto,
                "nome_produto": p.nome_produto,
                "categoria": p.categoria,
                "descricao": p.descricao,
                "preco": float(p.preco),
                "preco_promocional": (
                    float(p.preco_promocional)
                    if p.preco_promocional else None
                ),
                "volume": p.volume,
                "foto_url": p.foto_url,
                "estoque_disponivel": (p.estoque_atual or 0) > 0 if (
                    (p.estoque_atual or 0) > 0
                    or (p.estoque_minimo or 0) > 0
                ) else True,
            }
            for p in produtos
        ])
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/vitrine/categorias", methods=["GET"])
@limiter.limit("60 per minute")
def vitrine_categorias():
    """API pública — lista categorias disponíveis."""
    try:
        cats = (
            db.session.query(Produto.categoria)
            .filter(Produto.ativo.is_(True), Produto.categoria.isnot(None))
            .distinct()
            .order_by(Produto.categoria)
            .all()
        )
        return jsonify([c[0] for c in cats if c[0]])
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS — COMPLEMENTOS / TOPPINGS (Admin + Vitrine)
# =============================================================================


@app.route("/api/complementos", methods=["GET"])
@limiter.limit("120 per minute")
@api_login_required
def listar_complementos():
    """Listar complementos (admin)."""
    try:
        incluir_inativos = (
            request.args.get("incluir_inativos", "").lower() == "true"
        )
        query = (
            Complemento.query
            if incluir_inativos
            else Complemento.query.filter_by(ativo=True)
        )
        pagina, por_pagina = get_pagination_params(default_per_page=100)
        total = query.count()
        comps = (
            query.order_by(Complemento.categoria, Complemento.nome)
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )
        return jsonify({
            "complementos": [c.to_dict() for c in comps],
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": max(1, math.ceil(total / por_pagina)),
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/complementos", methods=["POST"])
@limiter.limit("30 per minute")
@api_login_required
def criar_complemento():
    """Criar novo complemento."""
    try:
        dados = request.get_json(silent=True) or {}
        nome = (dados.get("nome") or "").strip()
        if not nome:
            return jsonify({"erro": "Nome é obrigatório"}), 400

        comp = Complemento(
            nome=nome,
            categoria=dados.get("categoria"),
            unidade_medida=dados.get("unidade_medida"),
            preco_adicional=Decimal(str(dados.get("preco_adicional", 0))),
            ativo=True,
        )
        db.session.add(comp)
        db.session.commit()
        registrar_log(
            "criar", "complemento", comp.id_complemento,
            f"Complemento criado: {comp.nome}",
        )
        return jsonify(comp.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/complementos/<int:cid>", methods=["GET"])
@api_login_required
def obter_complemento(cid):
    """Obter detalhes de um complemento."""
    try:
        comp = db.session.get(Complemento, cid)
        if not comp:
            return jsonify({"erro": "Complemento não encontrado"}), 404
        return jsonify(comp.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/complementos/<int:cid>", methods=["PUT"])
@limiter.limit("30 per minute")
@api_login_required
def atualizar_complemento(cid):
    """Atualizar complemento existente."""
    try:
        comp = db.session.get(Complemento, cid)
        if not comp:
            return jsonify({"erro": "Complemento não encontrado"}), 404
        dados = request.get_json(silent=True) or {}
        if "nome" in dados and dados["nome"]:
            comp.nome = dados["nome"].strip()
        if "categoria" in dados:
            comp.categoria = dados["categoria"]
        if "unidade_medida" in dados:
            comp.unidade_medida = dados["unidade_medida"]
        if "preco_adicional" in dados:
            comp.preco_adicional = Decimal(
                str(dados["preco_adicional"])
            )
        if "ativo" in dados:
            comp.ativo = bool(dados["ativo"])
        db.session.commit()
        registrar_log(
            "editar", "complemento", cid,
            f"Complemento editado: {comp.nome}",
        )
        return jsonify(comp.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/complementos/<int:cid>", methods=["DELETE"])
@limiter.limit("30 per minute")
@api_login_required
def deletar_complemento(cid):
    """Desativar complemento (soft delete)."""
    try:
        comp = db.session.get(Complemento, cid)
        if not comp:
            return jsonify({"erro": "Complemento não encontrado"}), 404
        comp.ativo = False
        db.session.commit()
        registrar_log(
            "excluir", "complemento", cid,
            f"Complemento desativado: {comp.nome}",
        )
        return jsonify({"mensagem": "Complemento desativado"})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/vitrine/complementos", methods=["GET"])
@limiter.limit("60 per minute")
def vitrine_complementos():
    """API pública — lista complementos ativos para a vitrine."""
    try:
        comps = (
            Complemento.query
            .filter_by(ativo=True)
            .order_by(Complemento.categoria, Complemento.nome)
            .all()
        )
        return jsonify([c.to_dict() for c in comps])
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# AUTENTICAÇÃO DE CLIENTES — Login / Cadastro com senha
# =============================================================================


@app.route("/cliente/login", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def cliente_login():
    """Login do cliente (email ou telefone + senha)."""
    if request.method == "POST":
        identificador = (
            request.form.get("identificador", "").strip().lower()
        )
        senha = request.form.get("senha", "")

        if not identificador or not senha:
            return render_template(
                "cliente_login.html", erro="Preencha todos os campos."
            )

        # Buscar por email ou telefone
        telefone_limpo = _re.sub(r"\D", "", identificador)
        cliente = Cliente.query.filter(
            Cliente.ativo.is_(True),
            db.or_(
                db.func.lower(Cliente.email) == identificador,
                db.func.replace(
                    db.func.replace(
                        db.func.replace(Cliente.telefone, "(", ""),
                        ")", "",
                    ),
                    "-", "",
                ).like(f"%{telefone_limpo}%")
                if telefone_limpo and len(telefone_limpo) >= 8
                else db.func.lower(Cliente.email) == identificador,
            ),
        ).first()

        if cliente and cliente.verificar_senha(senha):
            session.clear()
            session.permanent = True
            session["cliente_id"] = cliente.id_cliente
            session["cliente_nome"] = cliente.nome
            session["tipo_usuario"] = "cliente"
            return redirect("/cliente/painel")

        logger.warning(
            "Login cliente falhou: %s de %s",
            identificador,
            request.remote_addr,
        )
        return render_template(
            "cliente_login.html", erro="Credenciais incorretas."
        )

    return render_template("cliente_login.html")


@app.route("/cliente/cadastro", methods=["GET", "POST"])
@limiter.limit("5 per minute", methods=["POST"])
def cliente_cadastro():
    """Cadastro do cliente com senha."""
    if request.method == "POST":
        dados = request.form
        nome = (dados.get("nome") or "").strip()
        telefone = (dados.get("telefone") or "").strip() or None
        email = (dados.get("email") or "").strip().lower() or None
        senha = dados.get("senha", "")
        observacoes = (dados.get("observacoes") or "").strip() or None
        consentimento = dados.get("consentimento_lgpd") == "on"

        erros = []
        if not nome or len(nome) < 2:
            erros.append("Nome deve ter ao menos 2 caracteres.")
        if not senha or len(senha) < 8:
            erros.append("Senha deve ter ao menos 8 caracteres.")
        elif not _re.search(r"[A-Z]", senha):
            erros.append(
                "Senha deve conter pelo menos 1 letra maiúscula."
            )
        elif not _re.search(r"[0-9]", senha):
            erros.append(
                "Senha deve conter pelo menos 1 número."
            )
        if not consentimento:
            erros.append("Consentimento LGPD é obrigatório.")
        if not email and not telefone:
            erros.append(
                "Informe ao menos email ou telefone para login."
            )
        if email and not _email_valido(email):
            erros.append("E-mail inválido.")

        # Verificar duplicatas (normaliza telefone para só dígitos)
        if not erros and telefone:
            tel_digits = _re.sub(r"\D", "", telefone)
            # Busca SQL com LIKE nos últimos 8 dígitos (evita carregar todos)
            sufixo = tel_digits[-8:] if len(tel_digits) >= 8 else tel_digits
            candidatos = Cliente.query.filter_by(
                ativo=True
            ).filter(
                Cliente.telefone.ilike(f"%{sufixo}%")
            ).all()
            for c in candidatos:
                if _re.sub(r"\D", "", c.telefone or "") == tel_digits:
                    erros.append(
                        "Este telefone já está cadastrado. "
                        "Use a opção 'Faça login'."
                    )
                    break
        if not erros and email:
            existente = Cliente.query.filter_by(
                email=email, ativo=True
            ).first()
            if existente:
                erros.append(
                    "Este e-mail já está cadastrado. "
                    "Use a opção 'Faça login'."
                )

        if erros:
            return render_template(
                "cliente_cadastro.html", erros=erros, dados=dados
            )

        try:
            cliente = Cliente(
                nome=nome,
                telefone=telefone,
                email=email,
                observacoes=observacoes,
                consentimento_lgpd=True,
                data_consentimento=datetime.now(timezone.utc),
                consentimento_versao="v1.0",
                ativo=True,
            )
            cliente.set_senha(senha)
            db.session.add(cliente)
            db.session.flush()

            entrada = ConsentimentoHistorico(
                id_cliente=cliente.id_cliente,
                acao="concedeu",
                versao_politica="v1.0",
                ip_address=request.remote_addr,
                user_agent=request.headers.get(
                    "User-Agent", ""
                )[:255],
            )
            db.session.add(entrada)

            cliente.pontos_fidelidade = 10
            db.session.commit()
            registrar_log(
                "criar",
                "cliente",
                cliente.id_cliente,
                f"Auto-cadastro com senha: {cliente.nome}",
            )

            # Login automático após cadastro
            session.permanent = True
            session["cliente_id"] = cliente.id_cliente
            session["cliente_nome"] = cliente.nome
            session["tipo_usuario"] = "cliente"
            return redirect("/cliente/painel")

        except Exception as e:
            db.session.rollback()
            logger.exception("Erro no cadastro cliente: %s", e)
            return render_template(
                "cliente_cadastro.html",
                erros=["Erro interno. Tente novamente."],
                dados=dados,
            )

    return render_template("cliente_cadastro.html", erros=[], dados={})


def cliente_login_required(f):
    """Decorator que exige login do cliente."""

    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("cliente_id"):
            return redirect("/cliente/login")
        return f(*args, **kwargs)

    return decorated


@app.route("/cliente/painel")
@cliente_login_required
def cliente_painel():
    """Painel do cliente — pontos, histórico de compras."""
    cliente = db.session.get(Cliente, session["cliente_id"])
    if not cliente or not cliente.ativo:
        session.clear()
        return redirect("/cliente/login")

    # Histórico de compras (com itens e produtos eager-loaded)
    vendas = (
        Venda.query.filter_by(id_cliente=cliente.id_cliente)
        .options(
            joinedload(Venda.itens).joinedload(ItemVenda.produto)
        )
        .order_by(Venda.data_venda.desc())
        .limit(20)
        .all()
    )

    return render_template(
        "cliente_painel.html", cliente=cliente, vendas=vendas
    )


@app.route("/cliente/logout")
def cliente_logout():
    """Logout do cliente — limpa sessão completamente."""
    session.clear()
    return redirect("/vitrine")


@app.route("/api/cliente/perfil", methods=["GET"])
@limiter.limit("30 per minute")
@cliente_login_required
def cliente_perfil_api():
    """API — perfil do cliente logado (pontos, badges, resumo)."""
    try:
        cliente = db.session.get(Cliente, session["cliente_id"])
        if not cliente or not cliente.ativo:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        total_compras = Venda.query.filter_by(
            id_cliente=cliente.id_cliente
        ).count()
        total_gasto = float(
            db.session.query(
                db.func.coalesce(db.func.sum(Venda.valor_total), 0)
            ).filter(Venda.id_cliente == cliente.id_cliente).scalar()
        )
        badges = BadgeCliente.query.filter_by(
            id_cliente=cliente.id_cliente
        ).order_by(BadgeCliente.data_conquista).all()

        return jsonify({
            "id_cliente": cliente.id_cliente,
            "nome": cliente.nome,
            "email": cliente.email,
            "telefone": cliente.telefone,
            "pontos_fidelidade": cliente.pontos_fidelidade or 0,
            "total_compras": total_compras,
            "total_gasto": round(total_gasto, 2),
            "badges": [b.to_dict() for b in badges],
            "membro_desde": (
                cliente.data_cadastro.isoformat()
                if cliente.data_cadastro else None
            ),
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/cliente/carrinho/checkout", methods=["POST"])
@limiter.limit("10 per minute")
@cliente_login_required
def cliente_checkout():
    """Finaliza o pedido do cliente (carrinho → Venda)."""
    try:
        dados = request.get_json(silent=True)
        if not dados or not dados.get("itens"):
            return jsonify({"erro": "Carrinho vazio."}), 400

        forma_pagamento = (
            dados.get("forma_pagamento", "").strip() or "Não informado"
        )
        formas_validas = [
            "Pix", "Cartão de Crédito", "Cartão de Débito",
            "Dinheiro", "Não informado",
        ]
        if forma_pagamento not in formas_validas:
            return jsonify({"erro": "Forma de pagamento inválida."}), 400

        cliente = db.session.get(Cliente, session["cliente_id"])
        if not cliente or not cliente.ativo:
            return jsonify({"erro": "Cliente não encontrado."}), 404

        if not cliente.consentimento_lgpd:
            return jsonify(
                {"erro": "Consentimento LGPD necessário para compras."}
            ), 403

        itens_req = dados["itens"]
        if len(itens_req) > 50:
            return jsonify({"erro": "Máximo 50 itens por pedido."}), 400

        itens_venda = []
        valor_total = 0

        for item in itens_req:
            id_produto = item.get("id_produto")
            quantidade = item.get("quantidade", 1)
            if (
                not id_produto
                or not isinstance(quantidade, int)
                or quantidade < 1
                or quantidade > 99
            ):
                return jsonify(
                    {"erro": "Item inválido no carrinho."}
                ), 400

            # Lock pessimista para evitar race condition no estoque
            produto = (
                db.session.query(Produto)
                .filter_by(id_produto=id_produto, ativo=True)
                .with_for_update()
                .first()
            )
            if not produto:
                return jsonify(
                    {"erro": f"Produto ID {id_produto} indisponível."}
                ), 400

            # Verificar estoque (consistente com admin criar_venda)
            controle_ativo = (
                (produto.estoque_atual or 0) > 0
                or (produto.estoque_minimo or 0) > 0
            )
            if controle_ativo and (
                produto.estoque_atual or 0
            ) < quantidade:
                return jsonify(
                    {
                        "erro": (
                            f'Estoque insuficiente para'
                            f' "{produto.nome_produto}":'
                            f" disponível"
                            f" {produto.estoque_atual},"
                            f" solicitado {quantidade}"
                        )
                    }
                ), 400

            preco = (
                float(produto.preco_promocional)
                if produto.preco_promocional
                else float(produto.preco)
            )
            subtotal = preco * quantidade
            itens_venda.append(
                ItemVenda(
                    id_produto=produto.id_produto,
                    quantidade=quantidade,
                    preco_unitario=preco,
                    subtotal=subtotal,
                )
            )
            valor_total += subtotal

            # Descontar estoque atomicamente (row já locked)
            if controle_ativo:
                produto.estoque_atual = max(
                    0, (produto.estoque_atual or 0) - quantidade
                )

        observacoes = (
            dados.get("observacoes", "").strip()[:500] or None
        )

        venda = Venda(
            id_cliente=cliente.id_cliente,
            valor_total=valor_total,
            forma_pagamento=forma_pagamento,
            status_pagamento="Concluído",
            observacoes=observacoes,
        )
        db.session.add(venda)
        db.session.flush()

        for iv in itens_venda:
            iv.id_venda = venda.id_venda
            db.session.add(iv)

        # Criar registro de pagamento (consistente com admin criar_venda)
        pagamento = Pagamento(
            id_venda=venda.id_venda,
            valor_pago=valor_total,
            metodo=forma_pagamento,
            status="Concluído",
        )
        db.session.add(pagamento)

        # 1 ponto por real gasto (cap 999999)
        pontos_ganhos = round(float(valor_total))
        cliente.pontos_fidelidade = min(
            999999, (cliente.pontos_fidelidade or 0) + pontos_ganhos
        )

        db.session.commit()

        # Invalidar cache vitrine (estoque atualizado)
        _invalidar_cache_vitrine()

        registrar_log(
            "criar", "venda", venda.id_venda,
            f"Pedido online cliente {cliente.nome}: "
            f"R${valor_total:.2f}, {len(itens_venda)} itens",
        )

        return jsonify({
            "sucesso": True,
            "id_venda": venda.id_venda,
            "valor_total": valor_total,
            "pontos_ganhos": pontos_ganhos,
            "pontos_total": cliente.pontos_fidelidade,
            "mensagem": (
                f"Pedido #{venda.id_venda} realizado! "
                f"+{pontos_ganhos} pontos"
            ),
        }), 201

    except Exception as e:
        db.session.rollback()
        logger.exception("Erro no checkout: %s", e)
        return jsonify({"erro": "Erro ao processar pedido."}), 500


# =============================================================================
# ROTAS - GESTÃO DE USUÁRIOS (admin only)
# =============================================================================


class UsuarioCreateSchema(BaseModel):
    nome: str
    email: EmailStr
    senha: str
    papel: str = "operador"

    @field_validator("nome")
    @classmethod
    def nome_valido(cls, v: str) -> str:
        v = v.strip()
        if len(v) < 2:
            raise ValueError("Nome deve ter ao menos 2 caracteres")
        return v

    @field_validator("senha")
    @classmethod
    def senha_forte(cls, v: str) -> str:
        if len(v) < 8:
            raise ValueError("Senha deve ter ao menos 8 caracteres")
        if not _re.search(r"[A-Z]", v):
            raise ValueError(
                "Senha deve conter pelo menos 1 letra maiúscula"
            )
        if not _re.search(r"[0-9]", v):
            raise ValueError(
                "Senha deve conter pelo menos 1 número"
            )
        return v

    @field_validator("papel")
    @classmethod
    def papel_valido(cls, v: str) -> str:
        if v not in ("admin", "operador"):
            raise ValueError('Papel deve ser "admin" ou "operador"')
        return v


@app.route("/api/usuarios", methods=["GET"])
@api_admin_required
def listar_usuarios():
    """Listar todos os usuários (admin only) — com paginação."""
    try:
        pagina, por_pagina = get_pagination_params(default_per_page=50)
        query = Usuario.query.order_by(Usuario.nome)
        total = query.count()
        usuarios = (
            query.offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )
        return jsonify({
            "usuarios": [u.to_dict() for u in usuarios],
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": max(1, math.ceil(total / por_pagina)),
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/usuarios", methods=["POST"])
@limiter.limit("10 per minute")
@api_admin_required
def criar_usuario():
    """Criar novo usuário (admin only)"""
    try:
        dados = validar_payload(UsuarioCreateSchema)

        if Usuario.query.filter_by(email=dados["email"].lower()).first():
            return jsonify({"erro": "Email já cadastrado"}), 409

        usuario = Usuario(
            nome=dados["nome"],
            email=dados["email"].lower(),
            papel=dados.get("papel", "operador"),
        )
        usuario.set_senha(dados["senha"])

        db.session.add(usuario)
        db.session.commit()
        registrar_log(
            "criar", "usuario", usuario.id_usuario,
            f"Novo usuario: {usuario.nome} ({usuario.papel})",
        )
        return jsonify(usuario.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return (
                jsonify({"erro": "Payload inválido", "detalhes": str(e)}),
                400,
            )
        return _erro_interno(e)


@app.route("/api/usuarios/<int:id_usuario>", methods=["PUT"])
@limiter.limit("10 per minute")
@api_admin_required
def atualizar_usuario(id_usuario):
    """Atualizar usuário (admin only)"""
    try:
        usuario = db.session.get(Usuario, id_usuario)
        if not usuario:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        dados = request.get_json(silent=True) or {}

        if "nome" in dados and dados["nome"]:
            usuario.nome = dados["nome"].strip()
        if "email" in dados and dados["email"]:
            novo_email = dados["email"].strip().lower()
            existente = Usuario.query.filter_by(email=novo_email).first()
            if existente and existente.id_usuario != id_usuario:
                return (
                    jsonify({"erro": "Email já cadastrado por outro usuário"}),
                    409,
                )
            usuario.email = novo_email
        if "papel" in dados and dados["papel"] in ("admin", "operador"):
            usuario.papel = dados["papel"]
        if "senha" in dados and dados["senha"]:
            if len(dados["senha"]) < 8:
                return (
                    jsonify({"erro": "Senha deve ter ao menos 8 caracteres"}),
                    400,
                )
            usuario.set_senha(dados["senha"])
        if "ativo" in dados:
            usuario.ativo = bool(dados["ativo"])

        db.session.commit()
        registrar_log(
            "atualizar", "usuario", id_usuario,
            f"Usuario atualizado: {usuario.nome}",
        )
        return jsonify(usuario.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/usuarios/<int:id_usuario>", methods=["DELETE"])
@api_admin_required
def deletar_usuario(id_usuario):
    """Desativar usuário (admin only, não pode desativar a si mesmo)"""
    try:
        if session.get("usuario_id") == id_usuario:
            return (
                jsonify(
                    {"erro": "Não é possível desativar seu próprio usuário"}
                ),
                400,
            )

        usuario = db.session.get(Usuario, id_usuario)
        if not usuario:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        usuario.ativo = False
        db.session.commit()
        registrar_log(
            "desativar", "usuario", id_usuario,
            f"Usuario desativado: {usuario.nome}",
        )
        return jsonify({"mensagem": f"Usuário {usuario.nome} desativado"})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/me", methods=["GET"])
@api_login_required
def usuario_atual():
    """Retorna dados do usuário logado"""
    try:
        usuario = db.session.get(Usuario, session.get("usuario_id"))
        if not usuario:
            return jsonify({"erro": "Usuário não encontrado"}), 404
        return jsonify(usuario.to_dict())
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - SUPORTE (Tickets + Chat)
# =============================================================================


@app.route("/suporte")
@login_required
def pagina_suporte():
    return render_template("suporte.html")


class TicketCreateSchema(BaseModel):
    assunto: str
    categoria: str = "duvida"
    prioridade: str = "normal"
    mensagem: str

    @field_validator("assunto")
    @classmethod
    def assunto_valido(cls, v: str) -> str:
        v = v.strip()
        if len(v) < 3:
            raise ValueError("Assunto deve ter ao menos 3 caracteres")
        if len(v) > 200:
            raise ValueError("Assunto deve ter no máximo 200 caracteres")
        return v

    @field_validator("categoria")
    @classmethod
    def categoria_valida(cls, v: str) -> str:
        if v not in ("duvida", "problema", "sugestao", "outro"):
            raise ValueError("Categoria inválida")
        return v

    @field_validator("prioridade")
    @classmethod
    def prioridade_valida(cls, v: str) -> str:
        if v not in ("baixa", "normal", "alta", "urgente"):
            raise ValueError("Prioridade inválida")
        return v

    @field_validator("mensagem")
    @classmethod
    def mensagem_valida(cls, v: str) -> str:
        v = v.strip()
        if len(v) < 5:
            raise ValueError("Mensagem deve ter ao menos 5 caracteres")
        return v


@app.route("/api/suporte/tickets", methods=["GET"])
@api_login_required
def listar_tickets():
    """Lista tickets: admin vê todos, operador vê os próprios"""
    try:
        pagina, por_pagina = get_pagination_params(default_per_page=20)
        status_filtro = request.args.get("status")

        query = TicketSuporte.query
        if session.get("papel") != "admin":
            query = query.filter_by(id_usuario=session["usuario_id"])
        if status_filtro:
            query = query.filter_by(status=status_filtro)

        query = query.order_by(TicketSuporte.data_atualizacao.desc())
        total = query.count()
        tickets = query.offset((pagina - 1) * por_pagina).limit(por_pagina).all()

        return jsonify(
            {
                "dados": [t.to_dict() for t in tickets],
                "total": total,
                "pagina": pagina,
                "limite": por_pagina,
                "total_paginas": (total + por_pagina - 1) // por_pagina,
            }
        )
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/suporte/tickets", methods=["POST"])
@limiter.limit("10 per minute")
@api_login_required
def criar_ticket():
    """Abre um novo ticket de suporte"""
    try:
        dados = validar_payload(TicketCreateSchema)

        ticket = TicketSuporte(
            id_usuario=session["usuario_id"],
            assunto=dados["assunto"],
            categoria=dados["categoria"],
            prioridade=dados["prioridade"],
        )
        db.session.add(ticket)
        db.session.flush()

        msg = MensagemTicket(
            id_ticket=ticket.id_ticket,
            id_usuario=session["usuario_id"],
            conteudo=dados["mensagem"],
        )
        db.session.add(msg)
        db.session.commit()

        registrar_log(
            "criar", "ticket_suporte", ticket.id_ticket, dados["assunto"]
        )
        return jsonify(ticket.to_dict()), 201
    except ValueError as e:
        return jsonify({"erro": "Dados inválidos", "detalhes": str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/suporte/tickets/<int:id_ticket>", methods=["GET"])
@api_login_required
def obter_ticket(id_ticket):
    """Retorna ticket com mensagens (chat)"""
    try:
        ticket = db.session.get(TicketSuporte, id_ticket)
        if not ticket:
            return jsonify({"erro": "Ticket não encontrado"}), 404

        if session.get(
            "papel"
        ) != "admin" and ticket.id_usuario != session.get("usuario_id"):
            return jsonify({"erro": "Acesso negado"}), 403

        return jsonify(ticket.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/suporte/tickets/<int:id_ticket>/mensagens", methods=["POST"])
@limiter.limit("30 per minute")
@api_login_required
def enviar_mensagem_ticket(id_ticket):
    """Envia mensagem em um ticket (chat)"""
    try:
        ticket = db.session.get(TicketSuporte, id_ticket)
        if not ticket:
            return jsonify({"erro": "Ticket não encontrado"}), 404

        if session.get(
            "papel"
        ) != "admin" and ticket.id_usuario != session.get("usuario_id"):
            return jsonify({"erro": "Acesso negado"}), 403

        if ticket.status == "fechado":
            return jsonify({"erro": "Este ticket está fechado"}), 400

        dados = request.get_json(silent=True) or {}
        conteudo = (dados.get("conteudo") or "").strip()
        if len(conteudo) < 1:
            return jsonify({"erro": "Mensagem não pode ser vazia"}), 400

        msg = MensagemTicket(
            id_ticket=id_ticket,
            id_usuario=session["usuario_id"],
            conteudo=conteudo,
        )
        db.session.add(msg)

        if ticket.status == "aberto" and session.get("papel") == "admin":
            ticket.status = "em_andamento"

        db.session.commit()
        registrar_log(
            "mensagem", "ticket", id_ticket,
            f"Mensagem enviada no ticket #{id_ticket}",
        )
        return jsonify(msg.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/suporte/tickets/<int:id_ticket>/status", methods=["PUT"])
@api_login_required
def atualizar_status_ticket(id_ticket):
    """Atualiza status do ticket."""
    try:
        ticket = db.session.get(TicketSuporte, id_ticket)
        if not ticket:
            return jsonify({"erro": "Ticket não encontrado"}), 404

        dados = request.get_json(silent=True) or {}
        novo_status = dados.get("status")
        if novo_status not in (
            "aberto",
            "em_andamento",
            "resolvido",
            "fechado",
        ):
            return jsonify({"erro": "Status inválido"}), 400

        if session.get("papel") != "admin":
            if ticket.id_usuario != session.get("usuario_id"):
                return jsonify({"erro": "Acesso negado"}), 403
            if novo_status not in ("fechado",):
                return (
                    jsonify(
                        {"erro": "Operador só pode fechar o próprio ticket"}
                    ),
                    403,
                )

        ticket.status = novo_status
        db.session.commit()

        registrar_log(
            "atualizar", "ticket_suporte", id_ticket, f"status → {novo_status}"
        )
        return jsonify(ticket.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - FORNECEDORES
# =============================================================================


@app.route("/api/fornecedores", methods=["GET"])
@api_login_required
def listar_fornecedores():
    """Listar fornecedores com busca opcional."""
    try:
        busca = (request.args.get("busca") or "").strip()
        query = Fornecedor.query
        if busca:
            query = query.filter(
                db.or_(
                    Fornecedor.nome.ilike(f"%{busca}%"),
                    Fornecedor.cnpj.ilike(f"%{busca}%"),
                )
            )
        incluir_inativos = request.args.get("incluir_inativos") == "true"
        if not incluir_inativos:
            query = query.filter_by(ativo=True)
        pagina, por_pagina = get_pagination_params(default_per_page=50)
        total = query.count()
        fornecedores = (
            query.order_by(Fornecedor.nome)
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )
        return jsonify({
            "fornecedores": [f.to_dict() for f in fornecedores],
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": max(1, math.ceil(total / por_pagina)),
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/fornecedores", methods=["POST"])
@limiter.limit("20 per minute")
@api_admin_required
def criar_fornecedor():
    """Criar novo fornecedor."""
    try:
        dados = request.get_json(silent=True) or {}
        nome = (dados.get("nome") or "").strip()
        if not nome or len(nome) < 2:
            return (
                jsonify({"erro": "Nome obrigatório (mínimo 2 caracteres)"}),
                400,
            )
        cnpj = (dados.get("cnpj") or "").strip() or None
        if cnpj:
            existente = Fornecedor.query.filter_by(cnpj=cnpj).first()
            if existente:
                return jsonify({"erro": "CNPJ já cadastrado"}), 400

        fornecedor = Fornecedor(
            nome=nome,
            cnpj=cnpj,
            telefone=(dados.get("telefone") or "").strip() or None,
            email=(dados.get("email") or "").strip() or None,
            endereco=(dados.get("endereco") or "").strip() or None,
            observacoes=(dados.get("observacoes") or "").strip() or None,
        )
        db.session.add(fornecedor)
        db.session.commit()
        registrar_log(
            "criar", "fornecedor", fornecedor.id_fornecedor, nome
        )
        return jsonify(fornecedor.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/fornecedores/<int:fid>", methods=["GET"])
@api_login_required
def obter_fornecedor(fid):
    """Obter detalhes de um fornecedor."""
    try:
        fornecedor = db.session.get(Fornecedor, fid)
        if not fornecedor:
            return jsonify({"erro": "Fornecedor não encontrado"}), 404
        return jsonify(fornecedor.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/fornecedores/<int:fid>", methods=["PUT"])
@api_admin_required
def atualizar_fornecedor(fid):
    """Atualizar dados de um fornecedor."""
    try:
        fornecedor = db.session.get(Fornecedor, fid)
        if not fornecedor:
            return jsonify({"erro": "Fornecedor não encontrado"}), 404

        dados = request.get_json(silent=True) or {}
        if "nome" in dados:
            nome = (dados["nome"] or "").strip()
            if len(nome) < 2:
                return (
                    jsonify({"erro": "Nome mínimo 2 caracteres"}), 400
                )
            fornecedor.nome = nome
        if "cnpj" in dados:
            cnpj = (dados["cnpj"] or "").strip() or None
            if cnpj:
                dup = Fornecedor.query.filter(
                    Fornecedor.cnpj == cnpj,
                    Fornecedor.id_fornecedor != fid,
                ).first()
                if dup:
                    return jsonify({"erro": "CNPJ já cadastrado"}), 400
            fornecedor.cnpj = cnpj
        for campo in ("telefone", "email", "endereco", "observacoes"):
            if campo in dados:
                setattr(
                    fornecedor, campo,
                    (dados[campo] or "").strip() or None,
                )
        if "ativo" in dados:
            fornecedor.ativo = bool(dados["ativo"])

        db.session.commit()
        registrar_log(
            "editar", "fornecedor", fid, fornecedor.nome
        )
        return jsonify(fornecedor.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/fornecedores/<int:fid>", methods=["DELETE"])
@api_admin_required
def desativar_fornecedor(fid):
    """Desativar fornecedor (soft delete)."""
    try:
        fornecedor = db.session.get(Fornecedor, fid)
        if not fornecedor:
            return jsonify({"erro": "Fornecedor não encontrado"}), 404
        fornecedor.ativo = False
        db.session.commit()
        registrar_log("desativar", "fornecedor", fid, fornecedor.nome)
        return jsonify({"mensagem": "Fornecedor desativado"})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - COMPRAS DE ESTOQUE
# =============================================================================


@app.route("/compras")
@login_required
def pagina_compras():
    """Página de gerenciamento de compras/estoque"""
    return render_template("compras.html")


@app.route("/api/compras", methods=["GET"])
@api_login_required
def listar_compras():
    """Listar compras de estoque com filtros."""
    try:
        pagina, por_pagina = get_pagination_params()
        query = CompraEstoque.query

        status = request.args.get("status")
        if status:
            query = query.filter_by(status=status)

        id_fornecedor = request.args.get("id_fornecedor", type=int)
        if id_fornecedor:
            query = query.filter_by(id_fornecedor=id_fornecedor)

        total = query.count()
        total_paginas = max(1, math.ceil(total / por_pagina))
        pagina = min(pagina, total_paginas)

        compras = (
            query.order_by(CompraEstoque.data_compra.desc())
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )
        return jsonify({
            "compras": [c.to_dict() for c in compras],
            "total": total,
            "pagina": pagina,
            "total_paginas": total_paginas,
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/compras", methods=["POST"])
@limiter.limit("20 per minute")
@api_admin_required
def criar_compra():
    """Registrar nova compra de estoque."""
    try:
        dados = request.get_json(silent=True) or {}

        id_fornecedor = dados.get("id_fornecedor")
        if not id_fornecedor:
            return jsonify({"erro": "Fornecedor obrigatório"}), 400

        fornecedor = db.session.get(Fornecedor, id_fornecedor)
        if not fornecedor:
            return jsonify({"erro": "Fornecedor não encontrado"}), 404

        itens = dados.get("itens")
        if not itens or not isinstance(itens, list) or len(itens) == 0:
            return (
                jsonify({"erro": "Compra deve ter pelo menos um item"}), 400
            )

        compra = CompraEstoque(
            id_fornecedor=id_fornecedor,
            nota_fiscal=(dados.get("nota_fiscal") or "").strip() or None,
            status=dados.get("status", "Pendente"),
            observacoes=(dados.get("observacoes") or "").strip() or None,
        )

        valor_total = Decimal("0.00")
        for item_dados in itens:
            id_produto = item_dados.get("id_produto")
            produto = db.session.get(Produto, id_produto)
            if not produto:
                return (
                    jsonify(
                        {"erro": f"Produto {id_produto} não encontrado"}
                    ),
                    404,
                )
            quantidade = int(item_dados.get("quantidade", 0))
            if quantidade <= 0:
                return (
                    jsonify({"erro": "Quantidade deve ser positiva"}), 400
                )
            preco_unit = Decimal(
                str(item_dados.get("preco_unitario", 0))
            )
            if preco_unit <= 0:
                return (
                    jsonify({"erro": "Preço unitário deve ser positivo"}),
                    400,
                )
            subtotal = preco_unit * quantidade
            item = ItemCompra(
                id_produto=id_produto,
                quantidade=quantidade,
                preco_unitario=preco_unit,
                subtotal=subtotal,
            )
            compra.itens.append(item)
            valor_total += subtotal

            # Se status é Recebido, atualizar estoque imediatamente
            if compra.status == "Recebido":
                produto.estoque_atual = (
                    produto.estoque_atual or 0
                ) + quantidade

        compra.valor_total = valor_total
        db.session.add(compra)
        db.session.commit()
        registrar_log(
            "criar", "compra", compra.id_compra,
            f"Compra #{compra.id_compra} — R${float(valor_total):.2f}"
            f" — Fornecedor: {fornecedor.nome}",
        )
        return jsonify(compra.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/compras/<int:cid>", methods=["GET"])
@api_login_required
def obter_compra(cid):
    """Obter detalhes de uma compra."""
    try:
        compra = db.session.get(CompraEstoque, cid)
        if not compra:
            return jsonify({"erro": "Compra não encontrada"}), 404
        return jsonify(compra.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/compras/<int:cid>", methods=["PUT"])
@api_admin_required
def atualizar_compra(cid):
    """Atualizar compra (nota fiscal, observações)."""
    try:
        compra = db.session.get(CompraEstoque, cid)
        if not compra:
            return jsonify({"erro": "Compra não encontrada"}), 404
        if compra.status == "Cancelado":
            return jsonify({"erro": "Compra cancelada não pode ser editada"}), 400

        dados = request.get_json(silent=True) or {}
        if "nota_fiscal" in dados:
            compra.nota_fiscal = (dados["nota_fiscal"] or "").strip() or None
        if "observacoes" in dados:
            compra.observacoes = (dados["observacoes"] or "").strip() or None

        db.session.commit()
        registrar_log("editar", "compra", cid, f"Compra #{cid} atualizada")
        return jsonify(compra.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/compras/<int:cid>/receber", methods=["POST"])
@api_admin_required
def receber_compra(cid):
    """Marcar compra como recebida — atualiza estoque."""
    try:
        compra = db.session.get(CompraEstoque, cid)
        if not compra:
            return jsonify({"erro": "Compra não encontrada"}), 404
        if compra.status == "Recebido":
            return jsonify({"erro": "Compra já recebida"}), 400
        if compra.status == "Cancelado":
            return jsonify({"erro": "Compra cancelada"}), 400

        # Atualizar estoque de cada produto
        for item in compra.itens:
            produto = db.session.get(Produto, item.id_produto)
            if produto:
                produto.estoque_atual = (
                    produto.estoque_atual or 0
                ) + item.quantidade

        compra.status = "Recebido"
        db.session.commit()
        registrar_log(
            "receber", "compra", cid,
            f"Compra #{cid} recebida — estoque atualizado",
        )
        return jsonify({
            "mensagem": f"Compra #{cid} recebida — estoque atualizado",
            "compra": compra.to_dict(),
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/compras/<int:cid>/cancelar", methods=["POST"])
@api_admin_required
def cancelar_compra(cid):
    """Cancelar compra — reverte estoque se já recebida."""
    try:
        compra = db.session.get(CompraEstoque, cid)
        if not compra:
            return jsonify({"erro": "Compra não encontrada"}), 404
        if compra.status == "Cancelado":
            return jsonify({"erro": "Compra já cancelada"}), 400

        # Se já recebida, reverter estoque
        if compra.status == "Recebido":
            for item in compra.itens:
                produto = db.session.get(Produto, item.id_produto)
                if produto:
                    produto.estoque_atual = max(
                        0, (produto.estoque_atual or 0) - item.quantidade
                    )

        compra.status = "Cancelado"
        db.session.commit()
        registrar_log(
            "cancelar", "compra", cid, f"Compra #{cid} cancelada"
        )
        return jsonify({
            "mensagem": f"Compra #{cid} cancelada",
            "compra": compra.to_dict(),
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - EDITAR VENDA + STATUS PEDIDO
# =============================================================================


@app.route("/api/vendas/<int:id_venda>", methods=["PUT"])
@api_admin_required
def editar_venda(id_venda):
    """Editar venda existente — forma pagamento, observações, desconto."""
    try:
        venda = db.session.get(Venda, id_venda)
        if not venda:
            return jsonify({"erro": "Venda não encontrada"}), 404
        if venda.status_pagamento == "Cancelado":
            return (
                jsonify({"erro": "Venda cancelada não pode ser editada"}),
                400,
            )

        dados = request.get_json(silent=True) or {}
        campos_editados = []

        if "forma_pagamento" in dados:
            venda.forma_pagamento = dados["forma_pagamento"]
            if venda.pagamento:
                venda.pagamento.metodo = dados["forma_pagamento"]
            campos_editados.append("forma_pagamento")

        if "observacoes" in dados:
            venda.observacoes = (dados["observacoes"] or "").strip() or None
            campos_editados.append("observacoes")

        if "status_pagamento" in dados:
            status_validos = ("Pendente", "Concluído", "Cancelado")
            if dados["status_pagamento"] not in status_validos:
                return (
                    jsonify({"erro": f"Status inválido. Use: {status_validos}"}),
                    400,
                )
            venda.status_pagamento = dados["status_pagamento"]
            if venda.pagamento:
                venda.pagamento.status = dados["status_pagamento"]
            campos_editados.append("status_pagamento")

        db.session.commit()
        registrar_log(
            "editar", "venda", id_venda,
            f"Venda #{id_venda} editada: {', '.join(campos_editados)}",
        )
        return jsonify(venda.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route(
    "/api/vendas/<int:id_venda>/status-pedido", methods=["PUT"]
)
@api_login_required
def atualizar_status_pedido(id_venda):
    """Atualizar status do pedido (workflow iFood-like)."""
    try:
        venda = db.session.get(Venda, id_venda)
        if not venda:
            return jsonify({"erro": "Venda não encontrada"}), 404

        dados = request.get_json(silent=True) or {}
        novo_status = (dados.get("status_pedido") or "").strip()
        _status_validos = (
            "Recebido", "Preparando", "Pronto", "Entregue", "Cancelado"
        )
        if novo_status not in _status_validos:
            return (
                jsonify(
                    {"erro": f"Status inválido. Use: {_status_validos}"}
                ),
                400,
            )

        # Workflow: não pode retroceder (exceto cancelar)
        _ordem = {
            "Recebido": 0, "Preparando": 1, "Pronto": 2,
            "Entregue": 3, "Cancelado": 99,
        }
        status_atual = venda.status_pedido or "Recebido"
        if (
            novo_status != "Cancelado"
            and _ordem.get(novo_status, 0) <= _ordem.get(status_atual, 0)
        ):
            return (
                jsonify(
                    {
                        "erro": (
                            f"Não pode retroceder de"
                            f" '{status_atual}' para '{novo_status}'"
                        )
                    }
                ),
                400,
            )

        venda.status_pedido = novo_status
        db.session.commit()
        registrar_log(
            "status_pedido", "venda", id_venda,
            f"Pedido #{id_venda}: {status_atual} → {novo_status}",
        )
        return jsonify({
            "mensagem": f"Pedido #{id_venda} → {novo_status}",
            "venda": venda.to_dict(),
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - CUPONS DE DESCONTO
# =============================================================================


@app.route("/api/cupons", methods=["GET"])
@api_login_required
def listar_cupons():
    """Listar cupons de desconto."""
    try:
        query = CupomDesconto.query
        ativo_only = request.args.get("ativo_only") != "false"
        if ativo_only:
            query = query.filter_by(ativo=True)
        pagina, por_pagina = get_pagination_params(default_per_page=50)
        total = query.count()
        cupons = (
            query.order_by(CupomDesconto.data_criacao.desc())
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )
        return jsonify({
            "cupons": [c.to_dict() for c in cupons],
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": max(1, math.ceil(total / por_pagina)),
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/cupons", methods=["POST"])
@limiter.limit("20 per minute")
@api_admin_required
def criar_cupom():
    """Criar novo cupom de desconto."""
    try:
        dados = request.get_json(silent=True) or {}
        codigo = (dados.get("codigo") or "").strip().upper()
        if not codigo or len(codigo) < 3:
            return (
                jsonify({"erro": "Código obrigatório (mínimo 3 caracteres)"}),
                400,
            )

        # Validar código único
        existente = CupomDesconto.query.filter_by(codigo=codigo).first()
        if existente:
            return jsonify({"erro": f"Código '{codigo}' já existe"}), 400

        tipo = dados.get("tipo_desconto", "percentual")
        if tipo not in ("percentual", "fixo"):
            return (
                jsonify({"erro": "tipo_desconto: 'percentual' ou 'fixo'"}),
                400,
            )

        valor = Decimal(str(dados.get("valor_desconto", 0)))
        if valor <= 0:
            return (
                jsonify({"erro": "Valor do desconto deve ser positivo"}), 400
            )
        if tipo == "percentual" and valor > 100:
            return (
                jsonify({"erro": "Percentual máximo é 100%"}), 400
            )

        cupom = CupomDesconto(
            codigo=codigo,
            descricao=(dados.get("descricao") or "").strip() or None,
            tipo_desconto=tipo,
            valor_desconto=valor,
            valor_minimo_pedido=Decimal(
                str(dados.get("valor_minimo_pedido", 0))
            ),
            usos_maximos=int(dados.get("usos_maximos", 0)),
        )

        data_fim = dados.get("data_fim")
        if data_fim:
            cupom.data_fim = datetime.fromisoformat(data_fim)

        db.session.add(cupom)
        db.session.commit()
        registrar_log(
            "criar", "cupom", cupom.id_cupom,
            f"Cupom {codigo} — {tipo} {float(valor)}",
        )
        return jsonify(cupom.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/cupons/<int:cid>", methods=["PUT"])
@api_admin_required
def atualizar_cupom(cid):
    """Atualizar cupom de desconto."""
    try:
        cupom = db.session.get(CupomDesconto, cid)
        if not cupom:
            return jsonify({"erro": "Cupom não encontrado"}), 404

        dados = request.get_json(silent=True) or {}
        if "descricao" in dados:
            cupom.descricao = (dados["descricao"] or "").strip() or None
        if "ativo" in dados:
            cupom.ativo = bool(dados["ativo"])
        if "usos_maximos" in dados:
            cupom.usos_maximos = int(dados["usos_maximos"])
        if "data_fim" in dados:
            cupom.data_fim = (
                datetime.fromisoformat(dados["data_fim"])
                if dados["data_fim"]
                else None
            )

        db.session.commit()
        registrar_log("editar", "cupom", cid, cupom.codigo)
        return jsonify(cupom.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/cupons/<int:cid>", methods=["DELETE"])
@api_admin_required
def desativar_cupom(cid):
    """Desativar cupom."""
    try:
        cupom = db.session.get(CupomDesconto, cid)
        if not cupom:
            return jsonify({"erro": "Cupom não encontrado"}), 404
        cupom.ativo = False
        db.session.commit()
        registrar_log("desativar", "cupom", cid, cupom.codigo)
        return jsonify({"mensagem": f"Cupom {cupom.codigo} desativado"})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/cupons/validar", methods=["POST"])
@limiter.limit("20 per minute")
@api_login_required
def validar_cupom():
    """Validar cupom no checkout."""
    try:
        dados = request.get_json(silent=True) or {}
        codigo = (dados.get("codigo") or "").strip().upper()
        if not codigo:
            return jsonify({"erro": "Código obrigatório"}), 400

        cupom = CupomDesconto.query.filter_by(codigo=codigo).first()
        if not cupom:
            return jsonify({"erro": "Cupom não encontrado"}), 404

        if not cupom.valido:
            return jsonify({"erro": "Cupom expirado ou esgotado"}), 400

        valor_pedido = Decimal(str(dados.get("valor_pedido", 0)))
        if (
            cupom.valor_minimo_pedido
            and valor_pedido < cupom.valor_minimo_pedido
        ):
            return (
                jsonify(
                    {
                        "erro": (
                            f"Pedido mínimo de"
                            f" R${float(cupom.valor_minimo_pedido):.2f}"
                        )
                    }
                ),
                400,
            )

        # Calcular desconto
        if cupom.tipo_desconto == "percentual":
            desconto = valor_pedido * cupom.valor_desconto / Decimal("100")
        else:
            desconto = min(cupom.valor_desconto, valor_pedido)

        return jsonify({
            "valido": True,
            "cupom": cupom.to_dict(),
            "desconto_calculado": float(desconto),
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - PROMOÇÕES (preço promocional nos produtos)
# =============================================================================


@app.route("/api/produtos/<int:id_produto>/promocao", methods=["PUT"])
@api_admin_required
def definir_promocao(id_produto):
    """Definir ou remover preço promocional de um produto."""
    try:
        produto = db.session.get(Produto, id_produto)
        if not produto:
            return jsonify({"erro": "Produto não encontrado"}), 404

        dados = request.get_json(silent=True) or {}
        preco_promo = dados.get("preco_promocional")

        if preco_promo is None or preco_promo == "" or preco_promo is False:
            produto.preco_promocional = None
            db.session.commit()
            registrar_log(
                "remover_promocao", "produto", id_produto,
                f"Promoção removida: {produto.nome_produto}",
            )
            return jsonify({
                "mensagem": "Promoção removida",
                "produto": produto.to_dict(),
            })

        preco_promo = Decimal(str(preco_promo))
        if preco_promo <= 0:
            return (
                jsonify({"erro": "Preço promocional deve ser positivo"}),
                400,
            )
        if preco_promo >= produto.preco:
            return (
                jsonify(
                    {"erro": "Preço promocional deve ser menor que o preço regular"}
                ),
                400,
            )

        produto.preco_promocional = preco_promo
        db.session.commit()
        registrar_log(
            "promocao", "produto", id_produto,
            f"{produto.nome_produto}: R${float(produto.preco):.2f}"
            f" → R${float(preco_promo):.2f}",
        )
        return jsonify({
            "mensagem": "Preço promocional definido",
            "produto": produto.to_dict(),
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# AGENTE IA — Machine Learning com TF-IDF + Cosine Similarity
# =============================================================================
# Implementação de NLP/ML sem dependências externas (scikit-learn-free).
# Usa TF-IDF (Term Frequency–Inverse Document Frequency) para vetorizar
# textos e Cosine Similarity para medir semelhança semântica.
# =============================================================================

# ---- Pré-processamento de texto (NLP pipeline) ----

_STOPWORDS_PT = frozenset(
    [
        "a",
        "ao",
        "aos",
        "aquela",
        "aquele",
        "as",
        "até",
        "com",
        "como",
        "da",
        "das",
        "de",
        "dela",
        "dele",
        "do",
        "dos",
        "e",
        "ela",
        "ele",
        "em",
        "entre",
        "era",
        "essa",
        "esse",
        "esta",
        "este",
        "eu",
        "foi",
        "for",
        "há",
        "isso",
        "isto",
        "já",
        "lhe",
        "lhes",
        "mas",
        "me",
        "meu",
        "minha",
        "muito",
        "na",
        "nas",
        "nem",
        "no",
        "nos",
        "nossa",
        "nosso",
        "nós",
        "num",
        "numa",
        "o",
        "os",
        "ou",
        "para",
        "pela",
        "pelo",
        "por",
        "qual",
        "quando",
        "que",
        "quem",
        "se",
        "sem",
        "ser",
        "seu",
        "sua",
        "são",
        "só",
        "também",
        "te",
        "tem",
        "tinha",
        "toda",
        "todo",
        "tu",
        "tua",
        "tudo",
        "um",
        "uma",
        "uns",
        "vai",
        "você",
        "vos",
        "à",
        "é",
    ]
)

# Stemmer RSLP simplificado para português
_SUFIXOS_PT = [
    "amentos",
    "imento",
    "amente",
    "idades",
    "ização",
    "ações",
    "mente",
    "idade",
    "ação",
    "ância",
    "ência",
    "ável",
    "ível",
    "ante",
    "ente",
    "ando",
    "endo",
    "indo",
    "ções",
    "ores",
    "ados",
    "idos",
    "ando",
    "endo",
    "indo",
    "aram",
    "eram",
    "iram",
    "avam",
    "ário",
    "ária",
    "eiro",
    "eira",
    "ado",
    "ido",
    "ção",
    "oso",
    "osa",
    "dor",
    "tor",
    "nte",
    "ais",
    "eis",
    "ões",
    "ais",
    "mos",
    "ria",
    "ndo",
    "ar",
    "er",
    "ir",
    "ou",
    "am",
    "em",
    "ia",
    "as",
    "es",
    "os",
]


def _normalizar_texto(texto):
    """Remove acentos e normaliza para lowercase ASCII."""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower()


def _stem_pt(palavra):
    """Stemmer simplificado para português — remove sufixos comuns."""
    if len(palavra) <= 3:
        return palavra
    for sufixo in _SUFIXOS_PT:
        if len(palavra) - len(sufixo) >= 3 and palavra.endswith(sufixo):
            return palavra[: -len(sufixo)]
    return palavra


def _tokenizar(texto):
    """Pipeline NLP: normaliza → tokeniza → remove stopwords → stem."""
    texto_norm = _normalizar_texto(texto)
    tokens = _re.findall(r"[a-z0-9]+", texto_norm)
    return [
        _stem_pt(t) for t in tokens if t not in _STOPWORDS_PT and len(t) > 1
    ]


# ---- TF-IDF Engine ----


class _TFIDFEngine:
    """Motor TF-IDF + Cosine Similarity para classificação de texto."""

    def __init__(self):
        self.documentos = []  # Lista de dicts originais
        self.doc_tokens = []  # Tokens processados por documento
        self.idf = {}  # IDF por termo
        self.tfidf_matrix = []  # Vetores TF-IDF dos documentos
        self.vocabulario = {}  # termo → index

    def treinar(self, base_conhecimento):
        """Treina o modelo com a base de conhecimento (fit)."""
        self.documentos = base_conhecimento
        self.doc_tokens = []

        # Gerar corpus: combina palavras-chave
        # + texto da resposta para cada doc
        for item in base_conhecimento:
            corpus_text = (
                " ".join(item["palavras"]) + " " + item.get("treino", "")
            )
            tokens = _tokenizar(corpus_text)
            self.doc_tokens.append(tokens)

        n_docs = len(self.doc_tokens)
        if n_docs == 0:
            return

        # Construir vocabulário
        all_terms = set()
        for tokens in self.doc_tokens:
            all_terms.update(tokens)
        self.vocabulario = {
            term: i for i, term in enumerate(sorted(all_terms))
        }

        # Calcular IDF: log(N / df_t) + 1  (suavizado)
        df = {}
        for tokens in self.doc_tokens:
            for term in set(tokens):
                df[term] = df.get(term, 0) + 1
        self.idf = {
            term: math.log(n_docs / df.get(term, 1)) + 1.0
            for term in self.vocabulario
        }

        # Calcular TF-IDF matrix
        self.tfidf_matrix = []
        for tokens in self.doc_tokens:
            vec = self._calcular_tfidf(tokens)
            self.tfidf_matrix.append(vec)

    def _calcular_tfidf(self, tokens):
        """Calcula vetor TF-IDF para uma lista de tokens."""
        tf = {}
        for t in tokens:
            tf[t] = tf.get(t, 0) + 1
        # Normalizar TF: tf / max_tf
        max_tf = max(tf.values()) if tf else 1
        vec = {}
        for term, count in tf.items():
            if term in self.idf:
                vec[term] = (count / max_tf) * self.idf[term]
        return vec

    def _cosine_similarity(self, vec_a, vec_b):
        """Calcula similaridade do cosseno entre dois vetores esparsos."""
        # Dot product
        termos_comuns = set(vec_a.keys()) & set(vec_b.keys())
        if not termos_comuns:
            return 0.0
        dot = sum(vec_a[t] * vec_b[t] for t in termos_comuns)
        # Normas
        norm_a = math.sqrt(sum(v * v for v in vec_a.values()))
        norm_b = math.sqrt(sum(v * v for v in vec_b.values()))
        if norm_a == 0 or norm_b == 0:
            return 0.0
        return dot / (norm_a * norm_b)

    def classificar(self, texto):
        """Classifica texto via cosine similarity."""
        tokens = _tokenizar(texto)
        if not tokens:
            return None, 0.0

        query_vec = self._calcular_tfidf(tokens)
        if not query_vec:
            return None, 0.0

        melhor_idx = -1
        melhor_sim = 0.0

        for i, doc_vec in enumerate(self.tfidf_matrix):
            sim = self._cosine_similarity(query_vec, doc_vec)
            if sim > melhor_sim:
                melhor_sim = sim
                melhor_idx = i

        if melhor_idx >= 0 and melhor_sim > 0.05:
            return self.documentos[melhor_idx], melhor_sim
        return None, 0.0


# ---- Base de Conhecimento (Training Data) ----

_IA_KNOWLEDGE_BASE = [
    {
        "palavras": [
            "senha",
            "password",
            "login",
            "acessar",
            "entrar",
            "logar",
            "autenticar",
        ],
        "treino": (
            "esqueci minha senha como faco login"
            " nao consigo entrar acessar sistema"
            " autenticacao credenciais email usuario"
        ),
        "resposta": "🔑 **Problemas com login?**\n\n"
        "1. Verifique se o email está correto (tudo minúsculo)\n"
        "2. Confira se o Caps Lock está desativado\n"
        "3. A senha deve ter no mínimo **8 caracteres**\n"
        "4. Tente limpar os cookies do navegador\n"
        "5. Se persistir, peça ao administrador para redefinir sua senha\n\n"
        "💡 Dica: use um gerenciador de senhas para não esquecer!",
        "categoria": "duvida",
    },
    {
        "palavras": [
            "venda",
            "registrar",
            "vender",
            "compra",
            "pedido",
            "carrinho",
        ],
        "treino": (
            "como registrar nova venda fazer vender"
            " produto pedido cliente compra carrinho"
            " finalizar forma pagamento pix"
            " dinheiro cartao"
        ),
        "resposta": "🛒 **Como registrar uma venda:**\n\n"
        "1. Vá em **Vendas → Registrar Venda** (ou atalho **V**)\n"
        "2. Selecione o cliente (precisa ter consentimento LGPD)\n"
        "3. Adicione os produtos e quantidade\n"
        "4. Aplique desconto % ou taxa se necessário\n"
        "5. Escolha a forma de pagamento (Pix, Dinheiro, Cartão)\n"
        "6. Clique em **Finalizar Venda**\n\n"
        "⚠️ O cliente precisa ter consentimento LGPD ativo!",
        "categoria": "duvida",
    },
    {
        "palavras": ["cliente", "cadastrar", "cadastro", "novo", "registrar"],
        "treino": (
            "como cadastrar novo cliente registro"
            " formulario nome telefone email totem"
            " autocadastro dados pessoais"
        ),
        "resposta": "👤 **Cadastro de clientes:**\n\n"
        "1. Vá em **Clientes → Cadastrar Cliente**\n"
        "2. Preencha nome, telefone e email\n"
        "3. O cliente deve aceitar o consentimento LGPD\n"
        "4. Clique em **Cadastrar**\n"
        "5. Ou use o **Totem de Auto-Cadastro**"
        " para o próprio cliente se cadastrar!\n\n"
        "📋 Sem consentimento LGPD, não é possível registrar vendas.",
        "categoria": "duvida",
    },
    {
        "palavras": [
            "lgpd",
            "privacidade",
            "dados",
            "consentimento",
            "exclusao",
            "esquecimento",
            "anonimizar",
        ],
        "treino": (
            "lei geral protecao dados pessoais"
            " consentimento revogar direito"
            " esquecimento anonimizacao"
            " privacidade politica termos"
            " lgpd compliance"
        ),
        "resposta": "🔒 **LGPD — Lei Geral de Proteção de Dados:**\n\n"
        "• Todo cliente tem direito a saber quais dados coletamos\n"
        "• O consentimento pode ser revogado a qualquer momento\n"
        '• O "direito ao esquecimento" anonimiza os dados do cliente\n'
        "• Revogação de consentimento bloqueia novas vendas\n"
        "• Histórico de consentimentos fica registrado por auditoria\n\n"
        "📋 Acesse: Menu → LGPD para ver a política completa.",
        "categoria": "duvida",
    },
    {
        "palavras": [
            "estoque",
            "produto",
            "acabou",
            "falta",
            "minimo",
            "repor",
        ],
        "treino": (
            "estoque baixo produto acabou faltando"
            " reposicao controle estoque minimo"
            " quantidade disponivel acabando"
            " alerta inventario"
        ),
        "resposta": "📦 **Gestão de estoque:**\n\n"
        "• Vá em **Produtos** no menu (ou atalho **P**)\n"
        "• Cada produto tem estoque atual e mínimo\n"
        "• Quando estoque fica abaixo do mínimo, aparece alerta no Dashboard\n"
        "• O estoque é decrementado automaticamente a cada venda\n"
        "• Edite o produto para ajustar quantidades\n\n"
        "💡 Configure o estoque mínimo para receber alertas!",
        "categoria": "duvida",
    },
    {
        "palavras": [
            "relatorio",
            "financeiro",
            "faturamento",
            "fechamento",
            "caixa",
            "grafico",
        ],
        "treino": (
            "relatorio financeiro faturamento vendas"
            " dia fechamento caixa exportar csv pdf"
            " xlsx grafico dashboard estatistica"
            " receita lucro"
        ),
        "resposta": "📊 **Relatórios e Financeiro:**\n\n"
        "• **Relatórios** → Vendas do dia,"
        " clientes frequentes, ranking de produtos\n"
        "• **Fechamento de Caixa** → Consolidação"
        " por data e forma de pagamento\n"
        "• Exporte em **CSV**, **PDF** ou **XLSX**\n"
        "• Gráficos interativos no Dashboard\n"
        "• Use atalho **R** para acessar rápido\n\n"
        "💰 O Dashboard mostra estatísticas em tempo real!",
        "categoria": "duvida",
    },
    {
        "palavras": [
            "erro",
            "bug",
            "funciona",
            "travou",
            "problema",
            "falha",
            "quebrado",
            "crash",
        ],
        "treino": (
            "erro bug nao funciona travou problema"
            " falha quebrado tela branca pagina"
            " carregando crash lento devagar"
            " congelou sistema fora"
        ),
        "resposta": "🐛 **Problemas técnicos:**\n\n"
        "1. Tente recarregar a página (**F5** ou **Ctrl+R**)\n"
        "2. Limpe o cache do navegador\n"
        "3. Verifique se está usando um navegador atualizado\n"
        "4. Tente outro navegador (Chrome, Firefox, Edge)\n"
        "5. Verifique sua conexão com a internet\n"
        "6. Se persistir, descreva o erro em detalhe neste ticket\n\n"
        "📸 Se possível, envie uma captura de tela do erro.",
        "categoria": "problema",
    },
    {
        "palavras": [
            "whatsapp",
            "compartilhar",
            "comprovante",
            "recibo",
            "enviar",
            "imprimir",
        ],
        "treino": (
            "whatsapp compartilhar comprovante recibo"
            " enviar impressao imprimir nota fiscal"
            " cupom mensagem zap zapzap"
        ),
        "resposta": "📱 **Compartilhamento e impressão:**\n\n"
        "• Após registrar uma venda, clique no"
        " botão **WhatsApp** para compartilhar"
        " o comprovante\n"
        "• O sistema gera automaticamente um recibo formatado\n"
        "• Você também pode exportar relatórios em PDF/CSV\n"
        "• Para imprimir, use **Ctrl+P** na página de relatórios\n\n"
        "💡 Use o atalho de teclado **V** para ir direto para Nova Venda!",
        "categoria": "duvida",
    },
    {
        "palavras": ["atalho", "teclado", "shortcut", "tecla", "rapido"],
        "treino": (
            "atalho teclado shortcut tecla rapido"
            " acesso direto hotkey key press"
            " combinacao comando"
        ),
        "resposta": "⌨️ **Atalhos de teclado disponíveis:**\n\n"
        "• **H** → Dashboard (Início)\n"
        "• **V** → Nova Venda\n"
        "• **C** → Clientes\n"
        "• **P** → Produtos\n"
        "• **R** → Relatórios\n"
        "• **S** → Suporte\n"
        "• **/** → Busca global\n"
        "• **?** → Lista de atalhos\n\n"
        "⚡ Funciona em qualquer página (exceto campos de texto)!",
        "categoria": "duvida",
    },
    {
        "palavras": [
            "fidelidade",
            "pontos",
            "desconto",
            "recompensa",
            "beneficio",
            "programa",
        ],
        "treino": (
            "programa fidelidade pontos desconto"
            " recompensa beneficio bonus cashback"
            " rede ranking cliente fiel vantagem"
        ),
        "resposta": "🌟 **Programa de Fidelidade:**\n\n"
        "• A cada R$1 gasto, o cliente ganha 1 ponto\n"
        "• 100 pontos = R$5,00 de desconto\n"
        "• Os pontos são creditados automaticamente a cada venda\n"
        "• Veja o ranking no Dashboard (Top Fidelidade)\n\n"
        "💜 Clientes fiéis são o coração do negócio!",
        "categoria": "duvida",
    },
    {
        "palavras": [
            "acai",
            "sabor",
            "cardapio",
            "menu",
            "tamanho",
            "copo",
            "tigela",
        ],
        "treino": (
            "acai sabor cardapio menu tamanho copo"
            " tigela 300ml 500ml 700ml fruta"
            " topping cobertura complemento"
            " granola banana morango"
        ),
        "resposta": "🍇 **Sobre o Cardápio Açaí:**\n\n"
        "• Os produtos e tamanhos são cadastrados em **Produtos**\n"
        "• Cada item tem nome, preço, categoria e estoque\n"
        "• Categorias ajudam a organizar: Açaí, Complementos, Bebidas\n"
        "• Edite preços e disponibilidade a qualquer momento\n\n"
        "🛍️ Gerencie tudo em: Menu → Produtos",
        "categoria": "duvida",
    },
    {
        "palavras": [
            "suporte",
            "ticket",
            "ajuda",
            "contato",
            "atendimento",
            "chat",
        ],
        "treino": (
            "suporte ticket ajuda contato falar"
            " alguem atendimento chat conversar"
            " mensagem resposta demora urgente"
        ),
        "resposta": "🎧 **Central de Suporte:**\n\n"
        "• Crie um **ticket** descrevendo sua dúvida ou problema\n"
        "• Escolha a categoria e prioridade adequadas\n"
        "• A equipe responderá o mais rápido possível\n"
        "• Use a **IA** para respostas rápidas automáticas\n"
        "• Acompanhe o status do ticket nesta página\n\n"
        "⚡ Tickets urgentes têm prioridade na fila!",
        "categoria": "duvida",
    },
]

# ---- Instância global do motor ML (treinado no startup) ----
_ia_engine = _TFIDFEngine()
_ia_engine.treinar(_IA_KNOWLEDGE_BASE)
logger.info(
    "IA ML Engine treinada: %d documentos, %d termos no vocabulário",
    len(_IA_KNOWLEDGE_BASE),
    len(_ia_engine.vocabulario),
)


def _ia_classificar_mensagem(texto):
    """Classifica a mensagem usando TF-IDF + Cosine Similarity (ML)."""
    doc, similaridade = _ia_engine.classificar(texto)

    if doc and similaridade > 0.08:
        return {
            "resposta": doc["resposta"],
            "confianca": min(similaridade * 1.5, 1.0),  # Escala: 0-1.0
            "categoria_sugerida": doc["categoria"],
            "metodo": "tfidf_cosine_similarity",
            "similaridade_bruta": round(similaridade, 4),
        }

    return {
        "resposta": "🤖 Não encontrei uma resposta"
        " automática para sua pergunta.\n\n"
        "Sugestões:\n"
        '• Tente reformular usando palavras-chave'
        ' (ex: "senha", "venda", "estoque")\n'
        "• Abra um ticket para a equipe responder pessoalmente\n"
        "• Use **?** para ver atalhos disponíveis\n\n"
        "📝 A IA aprende com as perguntas mais frequentes!",
        "confianca": 0.0,
        "categoria_sugerida": "duvida",
        "metodo": "nenhum_match",
        "similaridade_bruta": 0.0,
    }


@app.route("/api/suporte/ia-resposta", methods=["POST"])
@limiter.limit("20 per minute")
@api_login_required
def ia_resposta():
    """Agente IA — classifica perguntas via TF-IDF + Cosine Similarity (ML)"""
    try:
        dados = request.get_json(silent=True) or {}
        mensagem = (dados.get("mensagem") or "").strip()
        if len(mensagem) < 2:
            return jsonify({"erro": "Mensagem muito curta"}), 400

        resultado = _ia_classificar_mensagem(mensagem)
        return jsonify(
            {
                "resposta": resultado["resposta"],
                "confianca": round(resultado["confianca"], 3),
                "categoria_sugerida": resultado["categoria_sugerida"],
                "metodo": resultado.get("metodo", "tfidf"),
                "similaridade": resultado.get("similaridade_bruta", 0.0),
                "ia": True,
            }
        )
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# PIX — Geração de QR Code (BRCode EMV padrão BACEN)
# =============================================================================

# Chave PIX da loja (obrigatória via env)
_PIX_CHAVE = os.environ.get("PIX_CHAVE", "")
_PIX_NOME = os.environ.get("PIX_NOME", "Combina Acai")
_PIX_CIDADE = os.environ.get("PIX_CIDADE", "Lorena")


def _pix_campo(tag, valor):
    """Monta campo EMV: ID (2 chars) + Length (2 chars) + Value."""
    return "{}{:02d}{}".format(tag, len(valor), valor)


def _pix_crc16(payload):
    """CRC-16/CCITT-FALSE conforme especificação BRCode."""
    crc = 0xFFFF
    for byte in payload.encode("utf-8"):
        crc ^= byte << 8
        for _ in range(8):
            if crc & 0x8000:
                crc = (crc << 1) ^ 0x1021
            else:
                crc <<= 1
            crc &= 0xFFFF
    return "{:04X}".format(crc)


def _gerar_pix_payload(valor, txid="***"):
    """Gera payload PIX (BRCode) no padrão EMV estático."""
    # 00 - Payload Format Indicator
    pfi = _pix_campo("00", "01")
    # 01 - Point of Initiation Method (12 = estático)
    pim = _pix_campo("01", "12")
    # 26 - Merchant Account Information (PIX)
    gui = _pix_campo("00", "br.gov.bcb.pix")
    chave = _pix_campo("01", _PIX_CHAVE)
    mai = _pix_campo("26", gui + chave)
    # 52 - Merchant Category Code
    mcc = _pix_campo("52", "0000")
    # 53 - Transaction Currency (986 = BRL)
    trc = _pix_campo("53", "986")
    # 54 - Transaction Amount (se informado)
    tra = ""
    if valor and float(valor) > 0:
        tra = _pix_campo("54", "{:.2f}".format(float(valor)))
    # 58 - Country Code
    cc = _pix_campo("58", "BR")
    # 59 - Merchant Name (max 25 chars, ASCII)
    nome = _PIX_NOME[:25]
    mn = _pix_campo("59", nome)
    # 60 - Merchant City (max 15 chars)
    cidade = _PIX_CIDADE[:15]
    mc = _pix_campo("60", cidade)
    # 62 - Additional Data Field Template
    ref = _pix_campo("05", txid[:25])
    adft = _pix_campo("62", ref)
    # 63 - CRC16 (placeholder com 4 chars)
    payload_sem_crc = pfi + pim + mai + mcc + trc + tra + cc + mn + mc + adft
    payload_sem_crc += "6304"
    crc = _pix_crc16(payload_sem_crc)
    return payload_sem_crc + crc


@app.route("/api/pix/qrcode")
@limiter.limit("30 per minute")
@api_login_required
def pix_qrcode():
    """Gera payload PIX BRCode para QR Code."""
    try:
        if not _PIX_CHAVE:
            return jsonify({"erro": "Chave PIX não configurada"}), 503
        valor = request.args.get("valor", 0, type=float)
        txid = request.args.get("txid", "AcaiCRM")
        # Sanitizar txid (apenas alfanuméricos)
        txid = _re.sub(r"[^a-zA-Z0-9]", "", txid)[:25] or "AcaiCRM"
        if valor < 0 or valor > 99999.99:
            return jsonify({"erro": "Valor inválido"}), 400
        payload = _gerar_pix_payload(valor, txid)
        return jsonify({
            "payload": payload,
            "chave": _PIX_CHAVE,
            "valor": round(valor, 2),
            "nome": _PIX_NOME,
            "cidade": _PIX_CIDADE,
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# IA / ML — Recomendação de Produtos (Collaborative Filtering)
# =============================================================================


def _construir_matriz_compras():
    """Constrói matriz cliente×produto a partir do histórico de vendas."""
    vendas = (
        Venda.query.options(joinedload(Venda.itens))
        .order_by(Venda.data_venda.desc())
        .limit(2000)
        .all()
    )
    # {id_cliente: {id_produto: quantidade_total}}
    matriz = {}
    for v in vendas:
        cid = v.id_cliente
        if cid not in matriz:
            matriz[cid] = {}
        for item in v.itens:
            pid = item.id_produto
            matriz[cid][pid] = matriz[cid].get(pid, 0) + item.quantidade
    return matriz


def _similaridade_clientes(vetor_a, vetor_b):
    """Cosine similarity entre dois vetores de compra (dicts esparsos)."""
    comuns = set(vetor_a.keys()) & set(vetor_b.keys())
    if not comuns:
        return 0.0
    dot = sum(vetor_a[k] * vetor_b[k] for k in comuns)
    norm_a = math.sqrt(sum(v * v for v in vetor_a.values()))
    norm_b = math.sqrt(sum(v * v for v in vetor_b.values()))
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)


@app.route("/api/ia/recomendacoes/<int:id_cliente>")
@limiter.limit("30 per minute")
@api_login_required
def ia_recomendacoes(id_cliente):
    """ML — Recomendação de produtos via Collaborative Filtering."""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        matriz = _construir_matriz_compras()
        compras_alvo = matriz.get(id_cliente, {})

        if not compras_alvo:
            # Cold start: recomendar mais vendidos
            top = (
                db.session.query(
                    ItemVenda.id_produto,
                    db.func.sum(ItemVenda.quantidade).label("total"),
                )
                .group_by(ItemVenda.id_produto)
                .order_by(db.text("total DESC"))
                .limit(5)
                .all()
            )
            produtos_rec = []
            pids_top = [pid for pid, _ in top]
            prods_map = {
                p.id_produto: p
                for p in Produto.query.filter(
                    Produto.id_produto.in_(pids_top), Produto.ativo.is_(True)
                ).all()
            } if pids_top else {}
            for pid, total in top:
                p = prods_map.get(pid)
                if p:
                    produtos_rec.append({
                        **p.to_dict(),
                        "score": 1.0,
                        "motivo": "popular",
                    })
            return jsonify({
                "recomendacoes": produtos_rec,
                "metodo": "popularidade_cold_start",
                "cliente": cliente.nome,
            })

        # KNN: encontrar clientes mais similares
        similaridades = []
        for cid, compras in matriz.items():
            if cid == id_cliente:
                continue
            sim = _similaridade_clientes(compras_alvo, compras)
            if sim > 0.1:
                similaridades.append((cid, sim, compras))

        similaridades.sort(key=lambda x: x[1], reverse=True)
        vizinhos = similaridades[:10]  # Top-10 vizinhos

        # Calcular scores para produtos não comprados pelo alvo
        scores = {}
        for _cid, sim, compras in vizinhos:
            for pid, qtd in compras.items():
                if pid not in compras_alvo:
                    scores[pid] = scores.get(pid, 0) + sim * qtd

        # Ordenar e pegar top-5 — bulk load (evita N+1)
        ranking = sorted(scores.items(), key=lambda x: x[1], reverse=True)
        max_score = ranking[0][1] if ranking else 1
        top5_pids = [pid for pid, _ in ranking[:5]]
        prods_map = {
            p.id_produto: p
            for p in Produto.query.filter(
                Produto.id_produto.in_(top5_pids), Produto.ativo.is_(True)
            ).all()
        } if top5_pids else {}

        produtos_rec = []
        for pid, score in ranking[:5]:
            p = prods_map.get(pid)
            if p:
                produtos_rec.append({
                    **p.to_dict(),
                    "score": round(score / max_score, 3),
                    "motivo": "collaborative_filtering",
                })

        return jsonify({
            "recomendacoes": produtos_rec,
            "metodo": "collaborative_filtering_knn",
            "vizinhos_encontrados": len(vizinhos),
            "cliente": cliente.nome,
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# IA / ML — Segmentação de Clientes (Análise RFM)
# =============================================================================


def _calcular_rfm():
    """Calcula scores RFM (Recency, Frequency, Monetary) por cliente."""
    agora = datetime.now(timezone.utc)

    # Consulta agregada: evita N+1 (1 query em vez de N)
    stats = (
        db.session.query(
            Venda.id_cliente,
            db.func.max(Venda.data_venda).label("ultima"),
            db.func.count(Venda.id_venda).label("frequency"),
            db.func.sum(Venda.valor_total).label("monetary"),
        )
        .group_by(Venda.id_cliente)
        .all()
    )
    stats_map = {
        s.id_cliente: s for s in stats
    }

    clientes = Cliente.query.filter(
        Cliente.ativo.is_(True),
        Cliente.id_cliente.in_(stats_map.keys()),
    ).all()

    rfm_data = []
    for c in clientes:
        s = stats_map[c.id_cliente]
        ultima = s.ultima
        if ultima.tzinfo is None:
            ultima = ultima.replace(tzinfo=timezone.utc)
        recency = (agora - ultima).days

        rfm_data.append({
            "id_cliente": c.id_cliente,
            "nome": c.nome,
            "recency": recency,
            "frequency": s.frequency,
            "monetary": round(float(s.monetary), 2),
            "pontos": c.pontos_fidelidade or 0,
        })

    return rfm_data


def _segmentar_rfm(rfm_data):
    """Segmenta clientes em grupos com base nos scores RFM."""
    if not rfm_data:
        return []

    # Calcular percentis usando ordenação simples
    def _percentil_rank(dados, campo, inverso=False):
        """Atribui rank percentil (1-5) para cada item."""
        ordenados = sorted(
            dados, key=lambda x: x[campo], reverse=inverso
        )
        n = len(ordenados)
        for i, item in enumerate(ordenados):
            item[f"{campo}_score"] = min(int((i / max(n - 1, 1)) * 4) + 1, 5)

    _percentil_rank(rfm_data, "recency", inverso=False)   # Menor = melhor
    _percentil_rank(rfm_data, "frequency", inverso=True)   # Maior = melhor
    _percentil_rank(rfm_data, "monetary", inverso=True)     # Maior = melhor

    for item in rfm_data:
        r = item["recency_score"]
        f = item["frequency_score"]
        m = item["monetary_score"]
        media = (r + f + m) / 3.0

        if media >= 4.0:
            segmento = "Campeão"
        elif f >= 4 and m >= 3:
            segmento = "Cliente Fiel"
        elif r >= 4 and f <= 2:
            segmento = "Novo Promissor"
        elif r <= 2 and f >= 3:
            segmento = "Em Risco"
        elif r <= 2 and f <= 2:
            segmento = "Hibernando"
        else:
            segmento = "Regular"

        item["segmento"] = segmento
        item["rfm_score"] = round(media, 2)

    return rfm_data


@app.route("/api/ia/segmentacao")
@limiter.limit("10 per minute")
@api_login_required
def ia_segmentacao():
    """ML — Segmentação RFM de clientes."""
    try:
        rfm_data = _calcular_rfm()
        segmentados = _segmentar_rfm(rfm_data)

        # Resumo por segmento
        resumo = {}
        for item in segmentados:
            seg = item["segmento"]
            if seg not in resumo:
                resumo[seg] = {"total": 0, "receita": 0}
            resumo[seg]["total"] += 1
            resumo[seg]["receita"] += item["monetary"]

        for seg in resumo:
            resumo[seg]["receita"] = round(resumo[seg]["receita"], 2)

        return jsonify({
            "segmentacao": segmentados,
            "resumo": resumo,
            "total_clientes": len(segmentados),
            "metodo": "rfm_analysis",
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# IA / ML — Análise de Tendências de Vendas
# =============================================================================


@app.route("/api/ia/tendencias")
@limiter.limit("10 per minute")
@api_login_required
def ia_tendencias():
    """ML — Tendências de vendas com regressão linear simples."""
    try:
        dias = request.args.get("dias", 30, type=int)
        dias = min(max(dias, 7), 365)
        inicio = datetime.now(timezone.utc) - timedelta(days=dias)

        vendas = (
            Venda.query
            .filter(Venda.data_venda >= inicio)
            .order_by(Venda.data_venda)
            .all()
        )

        # Agrupar vendas por dia
        vendas_dia = {}
        for v in vendas:
            dia = v.data_venda.strftime("%Y-%m-%d")
            if dia not in vendas_dia:
                vendas_dia[dia] = {"qtd": 0, "receita": 0.0}
            vendas_dia[dia]["qtd"] += 1
            vendas_dia[dia]["receita"] += float(v.valor_total)

        # Regressão linear simples para projeção de receita
        if len(vendas_dia) >= 2:
            dias_list = sorted(vendas_dia.keys())
            y_vals = [vendas_dia[d]["receita"] for d in dias_list]
            n = len(y_vals)
            x_vals = list(range(n))

            # y = a + bx (mínimos quadrados)
            x_mean = sum(x_vals) / n
            y_mean = sum(y_vals) / n
            numerador = sum(
                (x - x_mean) * (y - y_mean)
                for x, y in zip(x_vals, y_vals)
            )
            denominador = sum((x - x_mean) ** 2 for x in x_vals)
            b = numerador / denominador if denominador != 0 else 0
            a = y_mean - b * x_mean

            # R² (coeficiente de determinação)
            ss_res = sum((y - (a + b * x)) ** 2
                         for x, y in zip(x_vals, y_vals))
            ss_tot = sum((y - y_mean) ** 2 for y in y_vals)
            r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0

            # Projeção para próximos 7 dias
            projecao = []
            ultimo_dia = datetime.strptime(dias_list[-1], "%Y-%m-%d").date()
            for i in range(1, 8):
                dia_futuro = ultimo_dia + timedelta(days=i)
                valor_proj = max(a + b * (n + i - 1), 0)
                projecao.append({
                    "data": dia_futuro.isoformat(),
                    "receita_projetada": round(valor_proj, 2),
                })

            tendencia = "alta" if b > 0.5 else "baixa" if b < -0.5 else "estável"
        else:
            a, b, r_squared = 0, 0, 0
            projecao = []
            tendencia = "dados_insuficientes"

        # Produto mais vendido no período
        produto_top = (
            db.session.query(
                Produto.nome_produto,
                db.func.sum(ItemVenda.quantidade).label("total"),
            )
            .join(ItemVenda, Produto.id_produto == ItemVenda.id_produto)
            .join(Venda, ItemVenda.id_venda == Venda.id_venda)
            .filter(Venda.data_venda >= inicio)
            .group_by(Produto.nome_produto)
            .order_by(db.text("total DESC"))
            .first()
        )

        return jsonify({
            "periodo_dias": dias,
            "total_vendas": len(vendas),
            "receita_total": round(
                sum(d["receita"] for d in vendas_dia.values()), 2
            ),
            "media_diaria": round(
                sum(d["receita"] for d in vendas_dia.values())
                / max(len(vendas_dia), 1), 2
            ),
            "tendencia": tendencia,
            "regressao": {
                "intercepto": round(a, 4),
                "coeficiente": round(b, 4),
                "r_squared": round(r_squared, 4),
            },
            "projecao_7dias": projecao,
            "produto_mais_vendido": (
                {"nome": produto_top[0], "quantidade": produto_top[1]}
                if produto_top else None
            ),
            "vendas_por_dia": {
                k: round(v["receita"], 2)
                for k, v in sorted(vendas_dia.items())
            },
            "metodo": "regressao_linear_simples",
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# IA / ML — Feedback Loop (Melhoria Contínua do Chatbot)
# =============================================================================

_ia_feedback_stats = {"positivo": 0, "negativo": 0}


def _ia_feedback_from_db():
    """Conta feedbacks da IA a partir do audit log (persistente)."""
    try:
        pos = LogAcao.query.filter_by(
            entidade="ia_feedback", acao="positivo"
        ).count()
        neg = LogAcao.query.filter_by(
            entidade="ia_feedback", acao="negativo"
        ).count()
        return {"positivo": pos, "negativo": neg}
    except Exception:
        return _ia_feedback_stats


@app.route("/api/suporte/ia-feedback", methods=["POST"])
@limiter.limit("30 per minute")
@api_login_required
def ia_feedback():
    """Feedback loop — registra se resposta da IA foi útil."""
    try:
        dados = request.get_json(silent=True) or {}
        util = dados.get("util")  # True/False
        pergunta = (dados.get("pergunta") or "").strip()[:500]

        if util is None or not isinstance(util, bool):
            return jsonify({"erro": "Campo 'util' (bool) obrigatório"}), 400

        # Persistir no audit log (sobrevive a restarts)
        acao = "positivo" if util else "negativo"
        log = LogAcao(
            id_usuario=session.get("usuario_id"),
            acao=acao,
            entidade="ia_feedback",
            detalhes=pergunta[:500] if pergunta else None,
            ip=request.remote_addr,
        )
        db.session.add(log)
        db.session.commit()

        stats = _ia_feedback_from_db()
        total = stats["positivo"] + stats["negativo"]
        taxa_sucesso = stats["positivo"] / total if total > 0 else 0

        logger.info(
            "IA Feedback: %s | pergunta='%s' | taxa=%.1f%%",
            "positivo" if util else "negativo",
            pergunta[:80],
            taxa_sucesso * 100,
        )

        return jsonify({
            "registrado": True,
            "stats": {
                "positivo": stats["positivo"],
                "negativo": stats["negativo"],
                "taxa_sucesso": round(taxa_sucesso, 3),
            },
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/ia/stats")
@limiter.limit("30 per minute")
@api_login_required
def ia_stats():
    """Estatísticas do motor de IA/ML."""
    try:
        fb_stats = _ia_feedback_from_db()
        total_fb = fb_stats["positivo"] + fb_stats["negativo"]
        return jsonify({
            "engine": {
                "tipo": "TF-IDF + Cosine Similarity",
                "documentos_treinados": len(_ia_engine.documentos),
                "vocabulario_tamanho": len(_ia_engine.vocabulario),
                "threshold_similaridade": 0.08,
            },
            "modulos": [
                "chatbot_tfidf",
                "recomendacao_collaborative_filtering",
                "segmentacao_rfm",
                "tendencias_regressao_linear",
                "feedback_loop",
            ],
            "feedback": {
                **fb_stats,
                "total": total_fb,
                "taxa_sucesso": round(
                    fb_stats["positivo"] / total_fb
                    if total_fb > 0 else 0, 3
                ),
            },
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - RELATÓRIOS COM FILTROS AVANÇADOS
# =============================================================================


@app.route("/api/relatorios/vendas-filtradas", methods=["GET"])
@api_login_required
def relatorio_vendas_filtradas():
    """Relatório de vendas com filtros avançados."""
    try:
        query = Venda.query

        # Filtro por período
        data_inicio = request.args.get("data_inicio")
        data_fim = request.args.get("data_fim")
        if data_inicio:
            query = query.filter(
                Venda.data_venda >= _dia_inicio(
                    datetime.strptime(data_inicio, "%Y-%m-%d").date()
                ) if isinstance(data_inicio, str) else
                Venda.data_venda >= _dia_inicio(data_inicio)
            )
        if data_fim:
            query = query.filter(
                Venda.data_venda <= _dia_fim(
                    datetime.strptime(data_fim, "%Y-%m-%d").date()
                ) if isinstance(data_fim, str) else
                Venda.data_venda <= _dia_fim(data_fim)
            )

        # Filtro por forma de pagamento
        forma = request.args.get("forma_pagamento")
        if forma:
            query = query.filter(Venda.forma_pagamento == forma)

        # Filtro por status do pedido
        status = request.args.get("status_pedido")
        if status:
            query = query.filter(Venda.status_pedido == status)

        # Filtro por status de pagamento
        status_pag = request.args.get("status_pagamento")
        if status_pag:
            query = query.filter(Venda.status_pagamento == status_pag)

        # Filtro por cliente
        id_cliente = request.args.get("id_cliente", type=int)
        if id_cliente:
            query = query.filter(Venda.id_cliente == id_cliente)

        # Filtro por produto
        id_produto = request.args.get("id_produto", type=int)
        if id_produto:
            query = query.join(ItemVenda).filter(
                ItemVenda.id_produto == id_produto
            )

        # Filtro por faixa de valor
        valor_min = request.args.get("valor_min", type=float)
        valor_max = request.args.get("valor_max", type=float)
        if valor_min is not None:
            query = query.filter(Venda.valor_total >= valor_min)
        if valor_max is not None:
            query = query.filter(Venda.valor_total <= valor_max)

        # Paginação
        pagina, por_pagina = get_pagination_params()
        total = query.count()
        vendas = (
            query.options(
                joinedload(Venda.cliente),
                joinedload(Venda.itens).joinedload(ItemVenda.produto),
            )
            .order_by(Venda.data_venda.desc())
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )

        # Totalizadores
        total_faturamento = float(
            query.with_entities(
                db.func.coalesce(db.func.sum(Venda.valor_total), 0)
            ).scalar()
        )
        ticket_medio = (
            total_faturamento / total if total > 0 else 0
        )

        # Resumo por forma de pagamento (dentro do filtro)
        resumo_pagamento = (
            query.with_entities(
                Venda.forma_pagamento,
                db.func.count(Venda.id_venda),
                db.func.sum(Venda.valor_total),
            )
            .group_by(Venda.forma_pagamento)
            .all()
        )

        return jsonify({
            "vendas": [v.to_dict() for v in vendas],
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": (total + por_pagina - 1) // por_pagina,
            "totalizadores": {
                "faturamento": round(total_faturamento, 2),
                "ticket_medio": round(ticket_medio, 2),
                "quantidade": total,
            },
            "por_forma_pagamento": [
                {
                    "forma": r[0] or "Indefinido",
                    "quantidade": r[1],
                    "total": float(r[2] or 0),
                }
                for r in resumo_pagamento
            ],
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - GESTÃO FINANCEIRA (receitas / despesas manuais)
# =============================================================================


@app.route("/api/financeiro", methods=["GET"])
@api_login_required
def listar_lancamentos():
    """Lista lançamentos financeiros com filtros e resumo."""
    try:
        query = LancamentoFinanceiro.query

        tipo = request.args.get("tipo")
        if tipo:
            query = query.filter(LancamentoFinanceiro.tipo == tipo)

        categoria = request.args.get("categoria")
        if categoria:
            query = query.filter(
                LancamentoFinanceiro.categoria == categoria
            )

        status = request.args.get("status")
        if status:
            query = query.filter(
                LancamentoFinanceiro.status == status
            )

        data_inicio = request.args.get("data_inicio")
        data_fim = request.args.get("data_fim")
        if data_inicio:
            query = query.filter(
                LancamentoFinanceiro.data_lancamento >= data_inicio
            )
        if data_fim:
            query = query.filter(
                LancamentoFinanceiro.data_lancamento <= data_fim
            )

        busca = request.args.get("busca", "").strip()
        if busca:
            term = f"%{busca}%"
            query = query.filter(
                db.or_(
                    LancamentoFinanceiro.descricao.ilike(term),
                    LancamentoFinanceiro.categoria.ilike(term),
                    LancamentoFinanceiro.comprovante.ilike(term),
                )
            )

        pagina, por_pagina = get_pagination_params()
        total = query.count()
        lancamentos = (
            query.order_by(
                LancamentoFinanceiro.data_lancamento.desc(),
                LancamentoFinanceiro.id_lancamento.desc(),
            )
            .offset((pagina - 1) * por_pagina)
            .limit(por_pagina)
            .all()
        )

        # Resumo geral (sem paginação, respeitando filtros)
        total_receitas = float(
            query.filter(
                LancamentoFinanceiro.tipo == "receita",
                LancamentoFinanceiro.status != "Cancelado",
            ).with_entities(
                db.func.coalesce(
                    db.func.sum(LancamentoFinanceiro.valor), 0
                )
            ).scalar()
        )
        total_despesas = float(
            query.filter(
                LancamentoFinanceiro.tipo == "despesa",
                LancamentoFinanceiro.status != "Cancelado",
            ).with_entities(
                db.func.coalesce(
                    db.func.sum(LancamentoFinanceiro.valor), 0
                )
            ).scalar()
        )

        return jsonify({
            "lancamentos": [lf.to_dict() for lf in lancamentos],
            "total": total,
            "pagina": pagina,
            "por_pagina": por_pagina,
            "total_paginas": math.ceil(total / por_pagina) if total else 0,
            "resumo": {
                "total_receitas": round(total_receitas, 2),
                "total_despesas": round(total_despesas, 2),
                "saldo": round(total_receitas - total_despesas, 2),
            },
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/financeiro", methods=["POST"])
@api_login_required
def criar_lancamento():
    """Cria um novo lançamento financeiro."""
    try:
        dados = validar_payload(LancamentoFinanceiroSchema)
    except ValueError as e:
        return jsonify({"erro": str(e)}), 400

    try:
        from datetime import date as date_type

        try:
            dt = date_type.fromisoformat(dados["data_lancamento"])
        except ValueError:
            return jsonify(
                {"erro": "Formato de data inválido. Use YYYY-MM-DD"}
            ), 400

        lanc = LancamentoFinanceiro(
            tipo=dados["tipo"],
            categoria=dados["categoria"],
            descricao=dados.get("descricao"),
            valor=Decimal(str(dados["valor"])),
            data_lancamento=dt,
            forma_pagamento=dados.get("forma_pagamento"),
            status=dados.get("status", "Pago"),
            comprovante=dados.get("comprovante"),
            observacoes=dados.get("observacoes"),
            id_usuario=session.get("usuario_id"),
        )
        db.session.add(lanc)
        db.session.commit()

        registrar_log(
            "criar", "lancamento_financeiro",
            lanc.id_lancamento,
            f"{lanc.tipo}: {lanc.categoria} R${lanc.valor}",
        )

        return jsonify(lanc.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/financeiro/<int:lid>", methods=["GET"])
@api_admin_required
def obter_lancamento(lid):
    """Obter detalhes de um lançamento financeiro."""
    lanc = db.session.get(LancamentoFinanceiro, lid)
    if not lanc:
        return jsonify({"erro": "Lançamento não encontrado"}), 404
    return jsonify(lanc.to_dict())


@app.route("/api/financeiro/<int:lid>", methods=["PUT"])
@api_admin_required
def atualizar_lancamento(lid):
    """Atualiza um lançamento financeiro existente."""
    lanc = db.session.get(LancamentoFinanceiro, lid)
    if not lanc:
        return jsonify({"erro": "Lançamento não encontrado"}), 404

    dados = request.get_json(silent=True) or {}

    try:
        if "tipo" in dados:
            if dados["tipo"] not in ("receita", "despesa"):
                return jsonify(
                    {"erro": "Tipo deve ser 'receita' ou 'despesa'"}
                ), 400
            lanc.tipo = dados["tipo"]

        if "categoria" in dados:
            lanc.categoria = dados["categoria"]
        if "descricao" in dados:
            lanc.descricao = dados["descricao"]
        if "valor" in dados:
            v = float(dados["valor"])
            if v <= 0:
                return jsonify(
                    {"erro": "Valor deve ser maior que zero"}
                ), 400
            lanc.valor = Decimal(str(v))
        if "data_lancamento" in dados:
            from datetime import date as date_type

            try:
                lanc.data_lancamento = date_type.fromisoformat(
                    dados["data_lancamento"]
                )
            except ValueError:
                return jsonify(
                    {"erro": "Data inválida. Use YYYY-MM-DD"}
                ), 400
        if "forma_pagamento" in dados:
            lanc.forma_pagamento = dados["forma_pagamento"]
        if "status" in dados:
            if dados["status"] not in ("Pago", "Pendente", "Cancelado"):
                return jsonify({"erro": "Status inválido"}), 400
            lanc.status = dados["status"]
        if "comprovante" in dados:
            lanc.comprovante = dados["comprovante"]
        if "observacoes" in dados:
            lanc.observacoes = dados["observacoes"]

        db.session.commit()

        registrar_log(
            "editar", "lancamento_financeiro",
            lanc.id_lancamento,
            f"{lanc.tipo}: {lanc.categoria} R${lanc.valor}",
        )

        return jsonify(lanc.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/financeiro/<int:lid>", methods=["DELETE"])
@api_admin_required
def deletar_lancamento(lid):
    """Cancela (soft delete) um lançamento financeiro."""
    lanc = db.session.get(LancamentoFinanceiro, lid)
    if not lanc:
        return jsonify({"erro": "Lançamento não encontrado"}), 404

    try:
        desc = f"{lanc.tipo}: {lanc.categoria} R${lanc.valor}"
        lanc.status = "Cancelado"
        db.session.commit()

        registrar_log(
            "excluir", "lancamento_financeiro",
            lid, desc,
        )

        return jsonify({"mensagem": "Lançamento cancelado com sucesso"})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/financeiro/resumo", methods=["GET"])
@api_login_required
def resumo_financeiro():
    """Resumo financeiro consolidado (lançamentos + vendas + compras)."""
    try:
        data_inicio_str = request.args.get("data_inicio")
        data_fim_str = request.args.get("data_fim")

        # Converter strings para datetime (portável SQLite + PostgreSQL)
        data_inicio = None
        data_fim = None
        if data_inicio_str:
            try:
                data_inicio = _dia_inicio(
                    datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
                )
            except ValueError:
                return jsonify(
                    {"erro": "data_inicio inválida. Use YYYY-MM-DD"}
                ), 400
        if data_fim_str:
            try:
                data_fim = _dia_fim(
                    datetime.strptime(data_fim_str, "%Y-%m-%d").date()
                )
            except ValueError:
                return jsonify(
                    {"erro": "data_fim inválida. Use YYYY-MM-DD"}
                ), 400

        # --- Lançamentos manuais ---
        q_lanc = LancamentoFinanceiro.query.filter(
            LancamentoFinanceiro.status != "Cancelado"
        )
        if data_inicio:
            q_lanc = q_lanc.filter(
                LancamentoFinanceiro.data_lancamento >= data_inicio
            )
        if data_fim:
            q_lanc = q_lanc.filter(
                LancamentoFinanceiro.data_lancamento <= data_fim
            )

        receitas_manual = float(
            q_lanc.filter(
                LancamentoFinanceiro.tipo == "receita"
            ).with_entities(
                db.func.coalesce(
                    db.func.sum(LancamentoFinanceiro.valor), 0
                )
            ).scalar()
        )
        despesas_manual = float(
            q_lanc.filter(
                LancamentoFinanceiro.tipo == "despesa"
            ).with_entities(
                db.func.coalesce(
                    db.func.sum(LancamentoFinanceiro.valor), 0
                )
            ).scalar()
        )

        # --- Vendas (receitas automáticas) ---
        q_vendas = Venda.query.filter(
            Venda.status_pagamento != "Cancelado"
        )
        if data_inicio:
            q_vendas = q_vendas.filter(
                Venda.data_venda >= data_inicio
            )
        if data_fim:
            q_vendas = q_vendas.filter(
                Venda.data_venda <= data_fim
            )
        receitas_vendas = float(
            q_vendas.with_entities(
                db.func.coalesce(db.func.sum(Venda.valor_total), 0)
            ).scalar()
        )

        # --- Compras de Estoque (despesas automáticas) ---
        q_compras = CompraEstoque.query.filter(
            CompraEstoque.status != "Cancelado"
        )
        if data_inicio:
            q_compras = q_compras.filter(
                CompraEstoque.data_compra >= data_inicio
            )
        if data_fim:
            q_compras = q_compras.filter(
                CompraEstoque.data_compra <= data_fim
            )
        despesas_compras = float(
            q_compras.with_entities(
                db.func.coalesce(
                    db.func.sum(CompraEstoque.valor_total), 0
                )
            ).scalar()
        )

        total_receitas = receitas_vendas + receitas_manual
        total_despesas = despesas_compras + despesas_manual
        saldo = total_receitas - total_despesas

        return jsonify({
            "total_receitas": round(total_receitas, 2),
            "total_despesas": round(total_despesas, 2),
            "saldo": round(saldo, 2),
            "detalhamento": {
                "receitas_vendas": round(receitas_vendas, 2),
                "receitas_manual": round(receitas_manual, 2),
                "despesas_compras": round(despesas_compras, 2),
                "despesas_manual": round(despesas_manual, 2),
            },
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/financeiro/categorias", methods=["GET"])
@api_login_required
def categorias_financeiro():
    """Retorna categorias distintas usadas nos lançamentos."""
    try:
        cats = (
            db.session.query(LancamentoFinanceiro.categoria)
            .distinct()
            .order_by(LancamentoFinanceiro.categoria)
            .all()
        )
        return jsonify([c[0] for c in cats])
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - GAMIFICAÇÃO (BADGES)
# =============================================================================


# Definição de badges disponíveis
# check(total_vendas, total_gasto, pontos) — SEM queries inline
BADGES_DEFINICAO = {
    "primeira_compra": {
        "nome": "Primeira Compra",
        "descricao": "Fez o primeiro pedido!",
        "icone": "🎉",
        "check": lambda v, g, p: v >= 1,
    },
    "fiel_10": {
        "nome": "Cliente Fiel",
        "descricao": "Completou 10 compras",
        "icone": "🏆",
        "check": lambda v, g, p: v >= 10,
    },
    "fiel_50": {
        "nome": "Super Fiel",
        "descricao": "Completou 50 compras!",
        "icone": "💎",
        "check": lambda v, g, p: v >= 50,
    },
    "gastador_100": {
        "nome": "Gastador",
        "descricao": "Gastou mais de R$100 no total",
        "icone": "💰",
        "check": lambda v, g, p: g >= 100,
    },
    "gastador_500": {
        "nome": "VIP",
        "descricao": "Gastou mais de R$500 — você é VIP!",
        "icone": "👑",
        "check": lambda v, g, p: g >= 500,
    },
    "pontos_100": {
        "nome": "Colecionador",
        "descricao": "Acumulou 100 pontos de fidelidade",
        "icone": "🌟",
        "check": lambda v, g, p: p >= 100,
    },
}


def _verificar_badges(id_cliente):
    """Verifica e concede badges ao cliente (3 queries em vez de N)."""
    cliente = db.session.get(Cliente, id_cliente)
    if not cliente:
        return []

    stats = db.session.query(
        db.func.count(Venda.id_venda),
        db.func.coalesce(db.func.sum(Venda.valor_total), 0),
    ).filter(Venda.id_cliente == id_cliente).first()
    total_vendas = stats[0]
    total_gasto = float(stats[1])
    pontos = cliente.pontos_fidelidade or 0

    badges_existentes = {
        row[0]
        for row in db.session.query(BadgeCliente.codigo).filter_by(
            id_cliente=id_cliente
        ).all()
    }

    novos = []
    for codigo, defn in BADGES_DEFINICAO.items():
        if codigo not in badges_existentes:
            if defn["check"](total_vendas, total_gasto, pontos):
                badge = BadgeCliente(
                    id_cliente=id_cliente,
                    codigo=codigo,
                    nome=defn["nome"],
                    descricao=defn["descricao"],
                    icone=defn["icone"],
                )
                db.session.add(badge)
                novos.append(badge)

    if novos:
        db.session.commit()
    return novos


@app.route("/api/clientes/<int:id_cliente>/badges", methods=["GET"])
@api_login_required
def listar_badges(id_cliente):
    """Listar badges conquistados por um cliente."""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente or not cliente.ativo:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        # Verificar novos badges
        novos = _verificar_badges(id_cliente)

        badges = BadgeCliente.query.filter_by(
            id_cliente=id_cliente
        ).order_by(BadgeCliente.data_conquista).all()

        # Listar badges não conquistados
        conquistados = {b.codigo for b in badges}
        disponiveis = [
            {
                "codigo": cod,
                "nome": d["nome"],
                "descricao": d["descricao"],
                "icone": d["icone"],
                "conquistado": False,
            }
            for cod, d in BADGES_DEFINICAO.items()
            if cod not in conquistados
        ]

        return jsonify({
            "badges": [b.to_dict() for b in badges],
            "novos": [b.to_dict() for b in novos],
            "disponiveis": disponiveis,
            "total_conquistados": len(badges),
            "total_disponiveis": len(BADGES_DEFINICAO),
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - EXTRATO DO CLIENTE (HISTÓRICO DE PAGAMENTOS)
# =============================================================================


@app.route(
    "/api/clientes/<int:id_cliente>/extrato", methods=["GET"]
)
@api_login_required
def extrato_cliente(id_cliente):
    """Extrato completo do cliente: compras, pontos, cupons."""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente or not cliente.ativo:
            return jsonify({"erro": "Cliente não encontrado"}), 404

        # Histórico de vendas
        vendas = (
            Venda.query.filter_by(id_cliente=id_cliente)
            .order_by(Venda.data_venda.desc())
            .limit(50)
            .all()
        )

        timeline = []
        total_gasto = Decimal("0.00")
        for v in vendas:
            total_gasto += v.valor_total
            timeline.append({
                "tipo": "compra",
                "data": v.data_venda.isoformat(),
                "descricao": (
                    f"Pedido #{v.id_venda} — "
                    f"{len(v.itens)} iten(s)"
                ),
                "valor": float(v.valor_total),
                "forma_pagamento": v.forma_pagamento,
                "status": v.status_pagamento,
                "status_pedido": v.status_pedido or "Recebido",
                "itens": [
                    {
                        "produto": i.produto.nome_produto,
                        "qtd": i.quantidade,
                        "subtotal": float(i.subtotal),
                    }
                    for i in v.itens
                ],
            })

        # Resumo
        total_compras = len(vendas)
        pontos = cliente.pontos_fidelidade or 0
        desconto_disponivel = (pontos // 100) * 5

        return jsonify({
            "cliente": {
                "id": cliente.id_cliente,
                "nome": cliente.nome,
                "pontos": pontos,
                "desconto_disponivel": desconto_disponivel,
                "desde": (
                    cliente.data_cadastro.isoformat()
                    if cliente.data_cadastro
                    else None
                ),
            },
            "resumo": {
                "total_compras": total_compras,
                "total_gasto": float(total_gasto),
                "ticket_medio": (
                    float(total_gasto / total_compras)
                    if total_compras > 0
                    else 0
                ),
            },
            "timeline": timeline,
        })
    except Exception as e:
        return _erro_interno(e)


@app.route("/cliente/extrato")
@cliente_login_required
def cliente_extrato_page():
    """Página de extrato do cliente logado."""
    cliente = db.session.get(Cliente, session["cliente_id"])
    if not cliente or not cliente.ativo:
        session.clear()
        return redirect("/cliente/login")
    return render_template(
        "cliente_extrato.html", cliente=cliente
    )


# =============================================================================
# ROTAS - PREVISÃO DE ESTOQUE (ML)
# =============================================================================


@app.route("/api/estoque/previsao", methods=["GET"])
@api_login_required
def previsao_estoque():
    """Previsão de consumo de estoque baseada em vendas recentes."""
    try:
        dias_analise = request.args.get("dias", 30, type=int)
        dias_analise = min(max(dias_analise, 7), 365)
        data_inicio = (
            datetime.now(timezone.utc) - timedelta(days=dias_analise)
        )

        # Consumo médio por produto nos últimos N dias
        # Subconsulta: total vendido por produto no período
        sub = (
            db.session.query(
                ItemVenda.id_produto,
                db.func.sum(ItemVenda.quantidade).label("total_vendido"),
            )
            .join(Venda, Venda.id_venda == ItemVenda.id_venda)
            .filter(Venda.data_venda >= data_inicio)
            .group_by(ItemVenda.id_produto)
            .subquery()
        )

        consumo = (
            db.session.query(
                Produto.id_produto,
                Produto.nome_produto,
                Produto.categoria,
                Produto.estoque_atual,
                Produto.estoque_minimo,
                db.func.coalesce(
                    sub.c.total_vendido, 0
                ).label("total_vendido"),
            )
            .outerjoin(
                sub, sub.c.id_produto == Produto.id_produto
            )
            .filter(Produto.ativo == True)  # noqa: E712
            .group_by(
                Produto.id_produto, Produto.nome_produto,
                Produto.categoria, Produto.estoque_atual,
                Produto.estoque_minimo,
                sub.c.total_vendido,
            )
            .all()
        )

        previsoes = []
        for p in consumo:
            vendido = int(p.total_vendido)
            media_diaria = vendido / dias_analise if vendido > 0 else 0
            estoque = p.estoque_atual or 0
            dias_restantes = (
                int(estoque / media_diaria)
                if media_diaria > 0
                else None
            )
            sugestao_compra = 0
            if media_diaria > 0:
                # Sugerir estoque para 14 dias
                necessario_14d = int(media_diaria * 14)
                if estoque < necessario_14d:
                    sugestao_compra = necessario_14d - estoque

            nivel = "ok"
            if dias_restantes is not None:
                if dias_restantes <= 3:
                    nivel = "critico"
                elif dias_restantes <= 7:
                    nivel = "alerta"

            previsoes.append({
                "id_produto": p.id_produto,
                "nome": p.nome_produto,
                "categoria": p.categoria,
                "estoque_atual": estoque,
                "estoque_minimo": p.estoque_minimo or 0,
                "vendido_periodo": vendido,
                "media_diaria": round(media_diaria, 2),
                "dias_restantes": dias_restantes,
                "sugestao_compra": sugestao_compra,
                "nivel": nivel,
            })

        # Ordenar: críticos primeiro
        ordem_nivel = {"critico": 0, "alerta": 1, "ok": 2}
        previsoes.sort(
            key=lambda x: (
                ordem_nivel.get(x["nivel"], 3),
                x["dias_restantes"] or 9999,
            )
        )

        return jsonify({
            "dias_analise": dias_analise,
            "total_produtos": len(previsoes),
            "criticos": sum(
                1 for p in previsoes if p["nivel"] == "critico"
            ),
            "alertas": sum(
                1 for p in previsoes if p["nivel"] == "alerta"
            ),
            "previsoes": previsoes,
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# FEATURE #24 — API VERSIONING (v1 proxy transparente)
# =============================================================================


@app.route("/api/version", methods=["GET"])
def api_version():
    """Retorna versão da API e prefixos disponíveis."""
    return jsonify({
        "version": "1.0",
        "prefix": "/api",
        "versioned_prefix": "/api/v1",
        "changelog": "v1.0 — versão inicial estável",
    })


@app.route(
    "/api/v1/<path:subpath>",
    methods=["GET", "POST", "PUT", "DELETE", "PATCH"],
)
def api_v1_proxy(subpath):
    """API v1 — proxy transparente para endpoints atuais."""
    target = f"/api/{subpath}"
    qs = request.query_string.decode()
    if qs:
        target += f"?{qs}"
    return redirect(target, code=307)


# =============================================================================
# FEATURE #25 — OPENAPI / SWAGGER EXPORT
# =============================================================================


@app.route("/api/openapi.json", methods=["GET"])
def openapi_export():
    """Exporta especificação OpenAPI 2.0 (Swagger) como JSON."""
    try:
        spec = api.__schema__
        return jsonify(spec)
    except Exception:
        return jsonify({
            "swagger": "2.0",
            "info": {"title": "Acaiteria CRM API", "version": "1.0"},
            "paths": {},
        })


# =============================================================================
# FEATURE #6 — CACHE: invalidação manual ao criar/editar dados
# =============================================================================


def _invalidar_cache_vitrine():
    """Invalida cache de endpoints da vitrine após alteração de produtos."""
    try:
        cache.delete("vitrine_produtos")
        # Limpar todas as variações com query_string
        # SimpleCache suporta clear() mas não pattern delete;
        # forçar limpeza completa é seguro com SimpleCache
        from flask_caching import Cache as _CacheRef  # noqa: F401
        backend = cache.cache
        # SimpleCache armazena em dict _cache; remover chaves com prefixo
        if hasattr(backend, '_cache'):
            keys_to_del = [
                k for k in list(backend._cache.keys())
                if 'vitrine_produtos' in str(k)
                or 'dashboard_kpi' in str(k)
            ]
            for k in keys_to_del:
                backend._cache.pop(k, None)
        else:
            cache.clear()
    except Exception:
        pass  # Cache miss is not critical


# =============================================================================
# FEATURE #8 — CURSOR PAGINATION (alternativa a OFFSET/LIMIT)
# =============================================================================


@app.route("/api/vendas/cursor", methods=["GET"])
@api_login_required
def vendas_cursor_pagination():
    """Lista vendas com cursor-pagination (mais eficiente para grandes bases).

    Params: after_id (int), limit (int, max 100), status (str)
    """
    try:
        after_id = request.args.get("after_id", type=int)
        limit = min(request.args.get("limit", 20, type=int), 100)
        status = request.args.get("status", "").strip()

        query = Venda.query
        if after_id:
            query = query.filter(Venda.id_venda < after_id)
        if status:
            query = query.filter(Venda.status_pagamento == status)

        vendas = (
            query.order_by(Venda.id_venda.desc())
            .limit(limit + 1)
            .all()
        )

        has_next = len(vendas) > limit
        if has_next:
            vendas = vendas[:limit]

        next_cursor = vendas[-1].id_venda if vendas and has_next else None

        return jsonify({
            "vendas": [v.to_dict() for v in vendas],
            "next_cursor": next_cursor,
            "has_next": has_next,
            "limit": limit,
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# FEATURE #2 — 2FA (Two-Factor Authentication) ENDPOINTS
# =============================================================================


@app.route("/api/2fa/setup", methods=["POST"])
@api_login_required
def setup_2fa():
    """Gera secret TOTP e URI para configurar no app autenticador."""
    try:
        uid = session.get("usuario_id")
        usuario = db.session.get(Usuario, uid)
        if not usuario:
            return jsonify({"erro": "Usuário não encontrado"}), 404

        existing = TwoFactorSecret.query.filter_by(
            id_usuario=uid
        ).first()
        if existing and existing.ativo:
            return jsonify({"erro": "2FA já está ativo"}), 400

        secret = pyotp.random_base32()

        if existing:
            existing.secret = secret
            existing.ativo = False
            existing.data_ativacao = None
        else:
            existing = TwoFactorSecret(
                id_usuario=uid, secret=secret, ativo=False,
            )
            db.session.add(existing)
        db.session.commit()

        totp = pyotp.TOTP(secret)
        uri = totp.provisioning_uri(
            name=usuario.email,
            issuer_name="Açaiteria CRM",
        )

        return jsonify({
            "secret": secret,
            "uri": uri,
            "mensagem": "Use o app autenticador para escanear o QR code.",
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/2fa/verify", methods=["POST"])
@api_login_required
def verify_2fa():
    """Verifica código TOTP e ativa 2FA."""
    try:
        uid = session.get("usuario_id")
        dados = request.get_json(silent=True) or {}
        codigo = str(dados.get("codigo", "")).strip()

        if not codigo:
            return jsonify({"erro": "Código obrigatório"}), 400

        tf = TwoFactorSecret.query.filter_by(id_usuario=uid).first()
        if not tf:
            return jsonify({"erro": "Execute setup primeiro"}), 400

        totp = pyotp.TOTP(tf.secret)
        if not totp.verify(codigo, valid_window=1):
            return jsonify({"erro": "Código inválido"}), 400

        tf.ativo = True
        tf.data_ativacao = datetime.now(timezone.utc)
        db.session.commit()

        registrar_log("ativar", "2fa", uid, "2FA ativado com sucesso")

        return jsonify({"mensagem": "2FA ativado com sucesso!"})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/2fa/disable", methods=["POST"])
@api_login_required
def disable_2fa():
    """Desativa 2FA do usuário logado."""
    try:
        uid = session.get("usuario_id")
        tf = TwoFactorSecret.query.filter_by(
            id_usuario=uid, ativo=True
        ).first()
        if not tf:
            return jsonify({"erro": "2FA não está ativo"}), 400

        tf.ativo = False
        db.session.commit()

        registrar_log("desativar", "2fa", uid, "2FA desativado")
        return jsonify({"mensagem": "2FA desativado com sucesso."})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/2fa/status", methods=["GET"])
@api_login_required
def status_2fa():
    """Retorna status do 2FA do usuário logado."""
    uid = session.get("usuario_id")
    tf = TwoFactorSecret.query.filter_by(id_usuario=uid).first()
    return jsonify({
        "ativo": bool(tf and tf.ativo),
        "data_ativacao": (
            tf.data_ativacao.isoformat()
            if tf and tf.data_ativacao else None
        ),
    })


# =============================================================================
# FEATURE #14 — UPLOAD DE FOTO DO PRODUTO (Base64)
# =============================================================================


@app.route("/api/produtos/<int:pid>/foto", methods=["POST"])
@api_admin_required
def upload_foto_produto(pid):
    """Upload de foto do produto em Base64 (max 500KB)."""
    try:
        produto = db.session.get(Produto, pid)
        if not produto:
            return jsonify({"erro": "Produto não encontrado"}), 404

        dados = request.get_json(silent=True) or {}
        foto_url = dados.get("foto_url", "").strip()

        if not foto_url:
            return jsonify({"erro": "foto_url obrigatório"}), 400

        # Aceitar data:image/... base64 ou URL externa
        if foto_url.startswith("data:image/"):
            import base64
            try:
                header, data = foto_url.split(",", 1)
                decoded = base64.b64decode(data)
                if len(decoded) > 512_000:  # 500KB max
                    return jsonify(
                        {"erro": "Imagem excede 500KB"}
                    ), 400
            except Exception:
                return jsonify(
                    {"erro": "Base64 inválido"}
                ), 400
        elif not foto_url.startswith("https://"):
            return jsonify(
                {"erro": "URL deve usar HTTPS ou ser data:image/"}
            ), 400

        produto.foto_url = foto_url
        db.session.commit()
        _invalidar_cache_vitrine()

        registrar_log(
            "editar", "produto", pid, "Foto atualizada"
        )
        return jsonify({
            "mensagem": "Foto atualizada",
            "produto": produto.to_dict(),
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# FEATURE #15 — WEBHOOKS — notificar sistemas externos
# =============================================================================


def _disparar_webhooks(evento, payload):
    """Dispara webhooks para o evento especificado (fire-and-forget)."""
    from concurrent.futures import ThreadPoolExecutor
    import urllib.request
    import json as _json

    hooks = WebhookConfig.query.filter_by(
        evento=evento, ativo=True
    ).all()
    if not hooks:
        return

    def _send(url, data, secret):
        try:
            body = _json.dumps(data).encode("utf-8")
            req = urllib.request.Request(
                url, data=body,
                headers={"Content-Type": "application/json"},
            )
            if secret:
                import hashlib
                import hmac
                sig = hmac.new(
                    secret.encode(), body, hashlib.sha256
                ).hexdigest()
                req.add_header("X-Webhook-Signature", sig)
            urllib.request.urlopen(req, timeout=10)
        except Exception as exc:
            logger.warning("Webhook falhou: %s — %s", url, exc)

    with ThreadPoolExecutor(max_workers=5) as pool:
        for hook in hooks:
            pool.submit(_send, hook.url, payload, hook.secret)


@app.route("/api/webhooks", methods=["GET"])
@api_admin_required
def listar_webhooks():
    """Lista webhooks configurados."""
    hooks = WebhookConfig.query.all()
    return jsonify([h.to_dict() for h in hooks])


@app.route("/api/webhooks", methods=["POST"])
@api_admin_required
def criar_webhook():
    """Cria configuração de webhook."""
    try:
        dados = request.get_json(silent=True) or {}
        evento = dados.get("evento", "").strip()
        url = dados.get("url", "").strip()

        eventos_validos = [
            "venda_criada", "pedido_pronto", "cliente_novo",
            "estoque_baixo",
        ]
        if evento not in eventos_validos:
            return jsonify({
                "erro": f"Evento inválido. Válidos: {eventos_validos}"
            }), 400
        if not url or not url.startswith("https://"):
            return jsonify(
                {"erro": "URL obrigatória (HTTPS)"}
            ), 400

        hook = WebhookConfig(
            evento=evento, url=url,
            secret=secrets.token_hex(32),
        )
        db.session.add(hook)
        db.session.commit()

        return jsonify(hook.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/webhooks/<int:wid>", methods=["DELETE"])
@api_admin_required
def deletar_webhook(wid):
    """Remove webhook."""
    hook = db.session.get(WebhookConfig, wid)
    if not hook:
        return jsonify({"erro": "Webhook não encontrado"}), 404
    db.session.delete(hook)
    db.session.commit()
    return jsonify({"mensagem": "Webhook removido"})


# =============================================================================
# FEATURE #16 — COMBOS / KITS
# =============================================================================


@app.route("/api/combos", methods=["GET"])
@api_login_required
def listar_combos():
    """Lista combos/kits disponíveis."""
    combos = ComboKit.query.filter_by(ativo=True).all()
    return jsonify([c.to_dict() for c in combos])


@app.route("/api/combos/<int:cid>", methods=["GET"])
@api_login_required
def obter_combo(cid):
    """Obter detalhes de um combo/kit."""
    combo = db.session.get(ComboKit, cid)
    if not combo:
        return jsonify({"erro": "Combo não encontrado"}), 404
    return jsonify(combo.to_dict())


@app.route("/api/combos", methods=["POST"])
@api_admin_required
def criar_combo():
    """Cria combo/kit com produtos e preço especial."""
    try:
        dados = request.get_json(silent=True) or {}
        nome = dados.get("nome", "").strip()
        preco = dados.get("preco_combo")
        itens = dados.get("itens", [])

        if not nome or preco is None:
            return jsonify(
                {"erro": "nome e preco_combo obrigatórios"}
            ), 400
        if not itens or not isinstance(itens, list):
            return jsonify(
                {"erro": "itens deve ser uma lista com ao menos 1 item"}
            ), 400

        combo = ComboKit(
            nome=nome,
            descricao=dados.get("descricao", ""),
            preco_combo=Decimal(str(preco)),
        )
        for item in itens:
            pid = item.get("id_produto")
            qtd = item.get("quantidade", 1)
            prod = db.session.get(Produto, pid)
            if not prod or not prod.ativo:
                return jsonify(
                    {"erro": f"Produto {pid} não encontrado ou inativo"}
                ), 400
            combo.itens.append(ComboKitItem(
                id_produto=pid, quantidade=qtd,
            ))

        db.session.add(combo)
        db.session.commit()

        registrar_log(
            "criar", "combo", combo.id_combo, f"Combo: {nome}"
        )
        return jsonify(combo.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/combos/<int:cid>", methods=["PUT"])
@api_admin_required
def atualizar_combo(cid):
    """Atualiza combo/kit."""
    try:
        combo = db.session.get(ComboKit, cid)
        if not combo:
            return jsonify({"erro": "Combo não encontrado"}), 404

        dados = request.get_json(silent=True) or {}
        if "nome" in dados:
            combo.nome = dados["nome"].strip()
        if "descricao" in dados:
            combo.descricao = dados["descricao"]
        if "preco_combo" in dados:
            combo.preco_combo = Decimal(str(dados["preco_combo"]))
        if "ativo" in dados:
            combo.ativo = bool(dados["ativo"])

        db.session.commit()
        return jsonify(combo.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/combos/<int:cid>", methods=["DELETE"])
@api_admin_required
def deletar_combo(cid):
    """Desativa combo/kit."""
    combo = db.session.get(ComboKit, cid)
    if not combo:
        return jsonify({"erro": "Combo não encontrado"}), 404
    combo.ativo = False
    db.session.commit()
    return jsonify({"mensagem": "Combo desativado"})


# =============================================================================
# FEATURE #4 — PROGRAMA DE INDICAÇÃO (REFERRAL)
# =============================================================================


@app.route("/api/indicacoes/codigo/<int:id_cliente>", methods=["GET"])
@api_login_required
def gerar_codigo_indicacao(id_cliente):
    """Gera ou retorna código de indicação único do cliente."""
    cliente = db.session.get(Cliente, id_cliente)
    if not cliente or not cliente.ativo:
        return jsonify({"erro": "Cliente não encontrado"}), 404

    # Verificar se já tem código de indicação gerado
    existente = Indicacao.query.filter_by(
        id_cliente_indicador=id_cliente,
        id_cliente_indicado=None,
    ).first()
    if existente:
        codigo = existente.codigo_indicacao
    else:
        codigo = f"IND{id_cliente:04d}{secrets.token_hex(3).upper()}"
        placeholder = Indicacao(
            id_cliente_indicador=id_cliente,
            codigo_indicacao=codigo,
        )
        db.session.add(placeholder)
        db.session.commit()

    return jsonify({
        "codigo": codigo,
        "cliente": cliente.nome,
        "link": f"/totem?ref={codigo}",
    })


@app.route("/api/indicacoes/validar", methods=["POST"])
@api_login_required
def validar_indicacao():
    """Valida código de indicação e concede bônus aos dois."""
    BONUS_PONTOS = 50
    try:
        dados = request.get_json(silent=True) or {}
        codigo = dados.get("codigo", "").strip().upper()
        id_indicado = dados.get("id_cliente_indicado")

        if not codigo or not id_indicado:
            return jsonify(
                {"erro": "codigo e id_cliente_indicado obrigatórios"}
            ), 400

        # Encontrar indicação existente com esse código
        ref = Indicacao.query.filter_by(
            codigo_indicacao=codigo
        ).first()
        if not ref:
            return jsonify(
                {"erro": "Código de indicação inválido"}
            ), 404

        indicador = db.session.get(Cliente, ref.id_cliente_indicador)
        indicado = db.session.get(Cliente, id_indicado)
        if not indicador or not indicado:
            return jsonify(
                {"erro": "Cliente não encontrado"}
            ), 404
        if indicador.id_cliente == indicado.id_cliente:
            return jsonify(
                {"erro": "Não pode indicar a si mesmo"}
            ), 400

        # Verificar se já existe indicação entre eles
        ja_existe = Indicacao.query.filter_by(
            id_cliente_indicador=indicador.id_cliente,
            id_cliente_indicado=indicado.id_cliente,
        ).first()
        if ja_existe:
            return jsonify(
                {"erro": "Indicação já registrada"}
            ), 400

        # Atualizar registro pendente e conceder bônus
        ref.id_cliente_indicado = indicado.id_cliente
        ref.bonus_concedido = True

        indicador.pontos_fidelidade = (
            (indicador.pontos_fidelidade or 0) + BONUS_PONTOS
        )
        indicado.pontos_fidelidade = (
            (indicado.pontos_fidelidade or 0) + BONUS_PONTOS
        )
        db.session.commit()

        return jsonify({
            "mensagem": (
                f"Indicação validada! +{BONUS_PONTOS} pontos "
                f"para {indicador.nome} e {indicado.nome}"
            ),
            "indicador_pontos": indicador.pontos_fidelidade,
            "indicado_pontos": indicado.pontos_fidelidade,
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/indicacoes", methods=["GET"])
@api_login_required
def listar_indicacoes():
    """Lista todas as indicações registradas."""
    indicacoes = Indicacao.query.order_by(
        Indicacao.data_indicacao.desc()
    ).limit(100).all()
    return jsonify([i.to_dict() for i in indicacoes])


# =============================================================================
# FEATURE #17 — AGENDAMENTO DE PEDIDOS
# =============================================================================


@app.route("/api/vendas/<int:vid>/agendar", methods=["PUT"])
@api_login_required
def agendar_pedido(vid):
    """Define data/hora de agendamento para retirada do pedido."""
    try:
        venda = db.session.get(Venda, vid)
        if not venda:
            return jsonify({"erro": "Venda não encontrada"}), 404

        dados = request.get_json(silent=True) or {}
        data_str = dados.get("data_agendamento", "").strip()
        if not data_str:
            return jsonify(
                {"erro": "data_agendamento obrigatório (ISO 8601)"}
            ), 400

        try:
            dt = datetime.fromisoformat(data_str)
        except ValueError:
            return jsonify(
                {"erro": "Formato inválido. Use ISO 8601"}
            ), 400

        agora = datetime.now(timezone.utc)
        if dt.tzinfo:
            if dt < agora:
                return jsonify(
                    {"erro": "Data deve ser no futuro"}
                ), 400
        else:
            if dt < agora.replace(tzinfo=None):
                return jsonify(
                    {"erro": "Data deve ser no futuro"}
                ), 400

        venda.data_agendamento = dt
        db.session.commit()

        return jsonify({
            "mensagem": "Pedido agendado",
            "data_agendamento": dt.isoformat(),
            "venda": venda.to_dict(),
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/agendamentos", methods=["GET"])
@api_login_required
def listar_agendamentos():
    """Lista vendas agendadas (futuras)."""
    try:
        vendas = (
            Venda.query
            .filter(Venda.data_agendamento.isnot(None))
            .filter(Venda.status_pedido != "Cancelado")
            .order_by(Venda.data_agendamento.asc())
            .limit(50)
            .all()
        )
        return jsonify([
            {
                **v.to_dict(),
                "data_agendamento": (
                    v.data_agendamento.isoformat()
                    if v.data_agendamento else None
                ),
            }
            for v in vendas
        ])
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# FEATURE #18 — ASSINATURAS / PLANOS MENSAIS
# =============================================================================


@app.route("/api/assinaturas/planos", methods=["GET"])
@api_login_required
def listar_planos():
    """Lista planos de assinatura disponíveis."""
    planos = Assinatura.query.filter_by(ativo=True).all()
    return jsonify([p.to_dict() for p in planos])


@app.route("/api/assinaturas/planos", methods=["POST"])
@api_admin_required
def criar_plano():
    """Cria novo plano de assinatura."""
    try:
        dados = request.get_json(silent=True) or {}
        nome = dados.get("nome_plano", "").strip()
        preco = dados.get("preco_mensal")
        limite = dados.get("limite_usos", 10)

        if not nome or preco is None:
            return jsonify(
                {"erro": "nome_plano e preco_mensal obrigatórios"}
            ), 400

        plano = Assinatura(
            nome_plano=nome,
            descricao=dados.get("descricao", ""),
            preco_mensal=Decimal(str(preco)),
            limite_usos=int(limite),
        )
        db.session.add(plano)
        db.session.commit()

        registrar_log(
            "criar", "assinatura", plano.id_assinatura,
            f"Plano: {nome}",
        )
        return jsonify(plano.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/assinaturas/planos/<int:pid>", methods=["GET"])
@api_login_required
def obter_plano(pid):
    """Obter detalhes de um plano de assinatura."""
    plano = db.session.get(Assinatura, pid)
    if not plano:
        return jsonify({"erro": "Plano não encontrado"}), 404
    return jsonify(plano.to_dict())


@app.route("/api/assinaturas/planos/<int:pid>", methods=["PUT"])
@api_admin_required
def atualizar_plano(pid):
    """Atualiza plano de assinatura."""
    try:
        plano = db.session.get(Assinatura, pid)
        if not plano:
            return jsonify({"erro": "Plano não encontrado"}), 404

        dados = request.get_json(silent=True) or {}
        if "nome_plano" in dados:
            plano.nome_plano = dados["nome_plano"].strip()
        if "descricao" in dados:
            plano.descricao = dados["descricao"]
        if "preco_mensal" in dados:
            plano.preco_mensal = Decimal(str(dados["preco_mensal"]))
        if "limite_usos" in dados:
            plano.limite_usos = int(dados["limite_usos"])
        if "ativo" in dados:
            plano.ativo = bool(dados["ativo"])

        db.session.commit()
        registrar_log(
            "editar", "assinatura", pid,
            f"Plano editado: {plano.nome_plano}",
        )
        return jsonify(plano.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/assinaturas/planos/<int:pid>", methods=["DELETE"])
@api_admin_required
def desativar_plano(pid):
    """Desativa plano de assinatura (soft delete)."""
    plano = db.session.get(Assinatura, pid)
    if not plano:
        return jsonify({"erro": "Plano não encontrado"}), 404
    plano.ativo = False
    db.session.commit()
    registrar_log(
        "excluir", "assinatura", pid,
        f"Plano desativado: {plano.nome_plano}",
    )
    return jsonify({"mensagem": "Plano desativado"})


@app.route("/api/assinaturas/assinar", methods=["POST"])
@api_login_required
def assinar_plano():
    """Vincula cliente a um plano de assinatura."""
    try:
        dados = request.get_json(silent=True) or {}
        id_plano = dados.get("id_assinatura")
        id_cli = dados.get("id_cliente")

        if not id_plano or not id_cli:
            return jsonify(
                {"erro": "id_assinatura e id_cliente obrigatórios"}
            ), 400

        plano = db.session.get(Assinatura, id_plano)
        if not plano or not plano.ativo:
            return jsonify(
                {"erro": "Plano não encontrado ou inativo"}
            ), 404

        cliente = db.session.get(Cliente, id_cli)
        if not cliente or not cliente.ativo:
            return jsonify(
                {"erro": "Cliente não encontrado"}
            ), 404

        from datetime import date as date_type
        hoje = date_type.today()
        fim = hoje + timedelta(days=30)

        assinatura = AssinaturaCliente(
            id_assinatura=id_plano,
            id_cliente=id_cli,
            data_inicio=hoje,
            data_fim=fim,
        )
        db.session.add(assinatura)
        db.session.commit()

        return jsonify(assinatura.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/assinaturas/<int:acid>/usar", methods=["POST"])
@api_login_required
def usar_assinatura(acid):
    """Registra uso de uma assinatura ativa."""
    try:
        assi = db.session.get(AssinaturaCliente, acid)
        if not assi:
            return jsonify(
                {"erro": "Assinatura não encontrada"}
            ), 404
        if assi.status != "ativa":
            return jsonify(
                {"erro": "Assinatura não está ativa"}
            ), 400
        if assi.usos_restantes <= 0:
            return jsonify(
                {"erro": "Limite de usos atingido"}
            ), 400

        from datetime import date as date_type
        if assi.data_fim < date_type.today():
            assi.status = "expirada"
            db.session.commit()
            return jsonify(
                {"erro": "Assinatura expirada"}
            ), 400

        assi.usos_realizados += 1
        db.session.commit()

        return jsonify({
            "mensagem": "Uso registrado",
            "usos_restantes": assi.usos_restantes,
            "assinatura": assi.to_dict(),
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route(
    "/api/clientes/<int:id_cli>/assinaturas", methods=["GET"]
)
@api_login_required
def assinaturas_cliente(id_cli):
    """Lista assinaturas de um cliente."""
    assinaturas = AssinaturaCliente.query.filter_by(
        id_cliente=id_cli
    ).order_by(AssinaturaCliente.data_inicio.desc()).all()
    return jsonify([a.to_dict() for a in assinaturas])


# =============================================================================
# FEATURE #1 — NOTIFICAÇÕES POR EMAIL (esqueleto)
# =============================================================================


@app.route("/api/notificacoes/email", methods=["POST"])
@api_admin_required
def enviar_notificacao_email():
    """Endpoint para envio de notificações por email.

    Em produção, configurar SMTP (SendGrid, AWS SES, etc).
    Aqui simula o envio e registra no log.
    """
    try:
        dados = request.get_json(silent=True) or {}
        destinatario = dados.get("email", "").strip()
        assunto = dados.get("assunto", "").strip()
        corpo = dados.get("corpo", "").strip()

        if not destinatario or not assunto or not corpo:
            return jsonify(
                {"erro": "email, assunto e corpo obrigatórios"}
            ), 400

        # Validação de email
        if not _email_valido(destinatario):
            return jsonify(
                {"erro": "Email inválido"}
            ), 400

        # Verificar se SMTP configurado
        smtp_host = os.environ.get("SMTP_HOST")
        if smtp_host:
            import smtplib
            from email.mime.text import MIMEText
            msg = MIMEText(corpo, "html", "utf-8")
            msg["Subject"] = assunto
            msg["From"] = os.environ.get(
                "SMTP_FROM", "noreply@acaiteria.com"
            )
            msg["To"] = destinatario
            try:
                with smtplib.SMTP(
                    smtp_host,
                    int(os.environ.get("SMTP_PORT", "587")),
                ) as server:
                    server.starttls()
                    server.login(
                        os.environ.get("SMTP_USER", ""),
                        os.environ.get("SMTP_PASS", ""),
                    )
                    server.send_message(msg)
                logger.info("Email enviado para %s", destinatario)
            except Exception as exc:
                logger.error("Erro ao enviar email: %s", exc)
                return jsonify({
                    "erro": "Falha ao enviar email",
                    "detalhe": str(exc),
                }), 500
        else:
            logger.info(
                "Email simulado para %s: %s", destinatario, assunto
            )

        registrar_log(
            "enviar", "notificacao",
            None, f"Email → {destinatario}: {assunto}",
        )

        return jsonify({
            "mensagem": "Notificação enviada" if smtp_host
            else "Notificação simulada (SMTP não configurado)",
            "destinatario": destinatario,
            "assunto": assunto,
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# FEATURE #3 — NFC-e (NOTA FISCAL CONSUMIDOR SIMPLIFICADA)
# =============================================================================


@app.route("/api/vendas/<int:vid>/nfce", methods=["GET"])
@api_login_required
def gerar_nfce(vid):
    """Gera cupom fiscal simplificado (NFC-e simulada) em PDF."""
    try:
        venda = Venda.query.options(
            joinedload(Venda.itens).joinedload(ItemVenda.produto)
        ).get(vid)
        if not venda:
            return jsonify(
                {"erro": "Venda não encontrada"}
            ), 404

        from reportlab.lib.pagesizes import mm
        from reportlab.pdfgen import canvas

        buf = io.BytesIO()
        # Cupom em papel 80mm x comprimento variável
        largura = 80 * mm
        n_itens = len(venda.itens)
        altura = (180 + n_itens * 20) * mm

        c = canvas.Canvas(buf, pagesize=(largura, altura))
        y = altura - 15 * mm

        # Cabeçalho
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(
            largura / 2, y, "COMBINA ACAI LTDA"
        )
        y -= 4 * mm
        c.setFont("Helvetica", 7)
        cnpj_loja = os.environ.get(
            "CNPJ_LOJA", "00.000.000/0001-00"
        )
        c.drawCentredString(
            largura / 2, y, f"CNPJ: {cnpj_loja}"
        )
        y -= 3 * mm
        c.drawCentredString(
            largura / 2, y,
            "DOCUMENTO AUXILIAR DE VENDA - NFC-e",
        )
        y -= 5 * mm
        c.line(5 * mm, y, largura - 5 * mm, y)
        y -= 5 * mm

        # Itens
        c.setFont("Helvetica", 7)
        for item in venda.itens:
            nome = (
                item.produto.nome_produto if item.produto
                else "Produto"
            )
            linha = (
                f"{item.quantidade}x {nome[:25]} "
                f"R${float(item.subtotal):.2f}"
            )
            c.drawString(5 * mm, y, linha)
            y -= 4 * mm

        # Totais
        y -= 2 * mm
        c.line(5 * mm, y, largura - 5 * mm, y)
        y -= 5 * mm
        c.setFont("Helvetica-Bold", 9)
        c.drawString(
            5 * mm, y,
            f"TOTAL: R$ {float(venda.valor_total):.2f}",
        )
        y -= 4 * mm
        c.setFont("Helvetica", 7)
        c.drawString(
            5 * mm, y,
            f"Pagamento: {venda.forma_pagamento or '-'}",
        )
        y -= 3 * mm
        c.drawString(
            5 * mm, y,
            (
                f"Data: "
                f"{venda.data_venda.strftime('%d/%m/%Y %H:%M')}"
                if venda.data_venda
                else "Data: -"
            ),
        )
        y -= 5 * mm
        c.line(5 * mm, y, largura - 5 * mm, y)
        y -= 4 * mm
        c.setFont("Helvetica", 6)
        c.drawCentredString(
            largura / 2, y,
            f"Venda #{venda.id_venda} | Obrigado pela preferencia!",
        )

        c.save()
        buf.seek(0)

        return send_file(
            buf, mimetype="application/pdf",
            as_attachment=True,
            download_name=f"nfce_venda_{vid}.pdf",
        )
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# FEATURE #5 — PAINEL DO CLIENTE ENHANCED (favoritos + reordenar)
# =============================================================================


@app.route("/api/cliente/favoritos", methods=["GET"])
def cliente_favoritos():
    """Retorna produtos mais comprados pelo cliente logado."""
    id_cli = session.get("cliente_id")
    if not id_cli:
        return jsonify({"erro": "Não autenticado"}), 401

    try:
        # Top 5 produtos mais comprados pelo cliente
        from sqlalchemy import func
        top = (
            db.session.query(
                Produto.id_produto,
                Produto.nome_produto,
                Produto.preco,
                Produto.foto_url,
                func.sum(ItemVenda.quantidade).label("total_qtd"),
            )
            .join(ItemVenda, ItemVenda.id_produto == Produto.id_produto)
            .join(Venda, Venda.id_venda == ItemVenda.id_venda)
            .filter(Venda.id_cliente == id_cli)
            .filter(Venda.status_pagamento != "Cancelado")
            .group_by(
                Produto.id_produto, Produto.nome_produto,
                Produto.preco, Produto.foto_url,
            )
            .order_by(func.sum(ItemVenda.quantidade).desc())
            .limit(5)
            .all()
        )
        return jsonify([
            {
                "id_produto": r.id_produto,
                "nome_produto": r.nome_produto,
                "preco": float(r.preco),
                "foto_url": r.foto_url,
                "total_comprado": int(r.total_qtd),
            }
            for r in top
        ])
    except Exception as e:
        return _erro_interno(e)


@app.route("/api/cliente/reordenar/<int:vid>", methods=["POST"])
def cliente_reordenar(vid):
    """Cria nova venda replicando itens de uma venda anterior."""
    id_cli = session.get("cliente_id")
    if not id_cli:
        return jsonify({"erro": "Não autenticado"}), 401

    try:
        venda_anterior = Venda.query.options(
            joinedload(Venda.itens)
        ).get(vid)
        if not venda_anterior or venda_anterior.id_cliente != id_cli:
            return jsonify(
                {"erro": "Venda não encontrada"}
            ), 404

        # Verificar LGPD
        cliente = db.session.get(Cliente, id_cli)
        if not cliente or not cliente.consentimento_lgpd:
            return jsonify(
                {"erro": "Consentimento LGPD necessário"}
            ), 403

        # Criar nova venda com os mesmos itens
        nova = Venda(
            id_cliente=id_cli,
            valor_total=0,
            forma_pagamento="PIX",
            status_pagamento="Pendente",
            status_pedido="Recebido",
        )
        total = Decimal("0")
        for item in venda_anterior.itens:
            prod = db.session.get(Produto, item.id_produto)
            if not prod or not prod.ativo:
                continue
            preco = (
                prod.preco_promocional
                if prod.preco_promocional else prod.preco
            )
            subtotal = preco * item.quantidade
            total += subtotal
            nova.itens.append(ItemVenda(
                id_produto=item.id_produto,
                quantidade=item.quantidade,
                preco_unitario=preco,
                subtotal=subtotal,
            ))

        if not nova.itens:
            return jsonify(
                {"erro": "Nenhum produto disponível para reordenar"}
            ), 400

        nova.valor_total = total
        db.session.add(nova)
        db.session.commit()

        return jsonify({
            "mensagem": "Pedido refeito com sucesso",
            "venda": nova.to_dict(),
        }), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# FEATURE #19 — MULTI-LOJA (CRUD de lojas)
# =============================================================================


@app.route("/api/lojas", methods=["GET"])
@api_login_required
def listar_lojas():
    """Lista todas as lojas/unidades."""
    lojas = Loja.query.filter_by(ativa=True).all()
    return jsonify([lj.to_dict() for lj in lojas])


@app.route("/api/lojas", methods=["POST"])
@api_admin_required
def criar_loja():
    """Cria nova unidade/loja."""
    try:
        dados = request.get_json(silent=True) or {}
        nome = dados.get("nome", "").strip()
        if not nome:
            return jsonify({"erro": "nome obrigatório"}), 400

        loja = Loja(
            nome=nome,
            endereco=dados.get("endereco"),
            telefone=dados.get("telefone"),
            cnpj=dados.get("cnpj"),
        )
        db.session.add(loja)
        db.session.commit()

        registrar_log(
            "criar", "loja", loja.id_loja, f"Loja: {nome}"
        )
        return jsonify(loja.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/lojas/<int:lid>", methods=["PUT"])
@api_admin_required
def atualizar_loja(lid):
    """Atualiza dados de uma loja."""
    try:
        loja = db.session.get(Loja, lid)
        if not loja:
            return jsonify({"erro": "Loja não encontrada"}), 404

        dados = request.get_json(silent=True) or {}
        if "nome" in dados:
            loja.nome = dados["nome"].strip()
        if "endereco" in dados:
            loja.endereco = dados["endereco"]
        if "telefone" in dados:
            loja.telefone = dados["telefone"]
        if "cnpj" in dados:
            loja.cnpj = dados["cnpj"]
        if "ativa" in dados:
            loja.ativa = bool(dados["ativa"])

        db.session.commit()
        return jsonify(loja.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route("/api/lojas/<int:lid>", methods=["DELETE"])
@api_admin_required
def desativar_loja(lid):
    """Desativa uma loja (soft delete)."""
    loja = db.session.get(Loja, lid)
    if not loja:
        return jsonify({"erro": "Loja não encontrada"}), 404
    loja.ativa = False
    db.session.commit()
    return jsonify({"mensagem": "Loja desativada"})


# =============================================================================
# TRATAMENTO DE ERROS
# =============================================================================


@app.errorhandler(404)
def nao_encontrado(erro):
    if request.path.startswith("/api/"):
        return jsonify({"erro": "Endpoint não encontrado"}), 404
    return render_template("404.html"), 404


@app.errorhandler(500)
def erro_interno(erro):
    if request.path.startswith("/api/"):
        return jsonify({"erro": "Erro interno do servidor"}), 500
    return render_template("500.html"), 500


# =============================================================================
# SEED — Cria admin padrão e produtos iniciais se não existirem
# =============================================================================


def _seed_admin():
    """Cria usuário admin padrão caso a tabela esteja vazia."""
    if Usuario.query.first() is None:
        admin = Usuario(
            nome="Administrador",
            email=os.environ.get("ADMIN_EMAIL", "admin@acaiteria.com"),
            papel="admin",
        )
        _admin_senha = os.environ.get("ADMIN_SENHA", "")
        if not _admin_senha:
            _admin_senha = secrets.token_urlsafe(16)
            logger.warning(
                "ADMIN_SENHA não definida; gerada aleatória: %s",
                _admin_senha,
            )
        admin.set_senha(_admin_senha)
        db.session.add(admin)
        db.session.commit()
        logger.info("Admin padrão criado: %s", admin.email)


def _seed_produtos():
    """Semeia catálogo real Combina Açaí.

    Estratégia: se a tabela não possui o catálogo completo (marcador =
    pelo menos 1 produto com volume='10L' e categoria='Açaí'), ADICIONA
    os produtos que estão faltando sem deletar os existentes (evita FK
    constraint com vendas já realizadas).
    """
    marcador = Produto.query.filter_by(
        categoria="Açaí", volume="10L"
    ).first()
    if marcador:
        return  # catálogo real já presente

    # ── Açaís (baldes de 10L e 5L) ──────────────────────────
    # (nome, volume, preço/kg-copo, descrição, estoque_atual, estoque_mínimo)
    _acais = [
        # Sabores clássicos
        ("Açaí Tradicional", "10L", 15.90,
         "Açaí puro 100% natural do Pará — o original",
         8, 3),
        ("Açaí Grego", "10L", 17.90,
         "Açaí extra-cremoso estilo grego, textura aveludada",
         6, 3),
        ("Açaí Black (Premium)", "10L", 19.90,
         "Açaí premium extra-concentrado, mais escuro e encorpado",
         4, 2),
        ("Açaí Zero Açúcar", "10L", 18.90,
         "Açaí sem adição de açúcar — ideal para dietas",
         4, 2),
        # Sabores frutados
        ("Açaí com Morango", "10L", 17.90,
         "Açaí com pedaços de morango natural",
         5, 2),
        ("Açaí com Banana", "5L", 16.90,
         "Açaí cremoso com banana natural",
         5, 2),
        ("Açaí Cupuaçu", "5L", 17.90,
         "Açaí com cupuaçu da Amazônia — sabor regional",
         4, 2),
        ("Açaí com Manga", "5L", 17.90,
         "Açaí tropical com manga madura",
         3, 2),
        ("Açaí com Maracujá", "5L", 17.90,
         "Açaí com toque cítrico de maracujá",
         3, 2),
        # Sabores especiais / gourmet
        ("Açaí Trufado", "10L", 19.90,
         "Açaí com sabor trufa de chocolate belga",
         3, 2),
        ("Açaí Ninho", "10L", 18.90,
         "Açaí com leite ninho cremoso — favorito das crianças",
         5, 2),
        ("Açaí Paçoca", "10L", 18.90,
         "Açaí com paçoca triturada artesanal",
         3, 2),
        ("Açaí Nutella®", "5L", 20.90,
         "Açaí com creme de avelã estilo Nutella",
         3, 2),
        ("Açaí Doce de Leite", "5L", 18.90,
         "Açaí com doce de leite artesanal mineiro",
         3, 2),
        ("Açaí Coco", "5L", 17.90,
         "Açaí com leite de coco e coco ralado",
         3, 2),
    ]

    # ── Sorvetes (baldes de 10L e 5L) ──────────────────────
    _sorvetes = [
        ("Menta com Chocolate", "10L", 12.90,
         "Refrescante menta com lascas de chocolate",
         4, 2),
        ("Chocolate Belga", "10L", 14.90,
         "Chocolate belga premium importado",
         5, 2),
        ("Pistache", "10L", 15.90,
         "Pistache artesanal cremoso",
         3, 2),
        ("Coco", "10L", 12.90,
         "Coco natural ralado cremoso",
         4, 2),
        ("Cappuccino", "10L", 13.90,
         "Sabor cappuccino com toque de canela",
         3, 2),
        ("Doce de Leite", "10L", 12.90,
         "Doce de leite artesanal mineiro",
         5, 2),
        ("Grego Maracujá", "10L", 14.90,
         "Sorvete grego com calda de maracujá",
         3, 2),
        ("Grego Cereja", "10L", 14.90,
         "Sorvete grego com cerejas",
         3, 2),
        ("Unicórnio", "10L", 13.90,
         "Mix colorido de sabores frutados",
         3, 2),
        ("Pitaya", "10L", 14.90,
         "Pitaya rosa natural e refrescante",
         3, 2),
        ("Limão Siciliano", "10L", 11.90,
         "Limão siciliano refrescante — sorbet",
         3, 2),
        ("Morango", "10L", 12.90,
         "Morango com pedaços de fruta",
         5, 2),
        ("Flocos", "10L", 12.90,
         "Creme com flocos crocantes de chocolate",
         4, 2),
        ("Manga", "5L", 12.90,
         "Manga madura tropical",
         3, 2),
        ("Abacaxi", "5L", 11.90,
         "Abacaxi refrescante natural",
         3, 2),
        ("Banana Caramelizada", "5L", 12.90,
         "Banana com calda de caramelo artesanal",
         3, 2),
        ("Paçoca", "5L", 12.90,
         "Paçoca cremosa artesanal",
         3, 2),
        ("Chocolate Branco", "5L", 12.90,
         "Chocolate branco aveludado premium",
         3, 2),
        ("Baunilha", "5L", 11.90,
         "Baunilha clássica de Madagascar",
         4, 2),
        ("Café", "5L", 12.90,
         "Café expresso intenso",
         3, 2),
        ("Goiaba", "5L", 11.90,
         "Goiaba vermelha cascuda",
         3, 2),
        ("Creme de Cupuaçu", "5L", 13.90,
         "Cupuaçu amazônico cremoso",
         3, 2),
        ("Milho Verde", "5L", 12.90,
         "Milho verde cremoso estilo junino",
         3, 2),
        ("Algodão Doce", "5L", 12.90,
         "Sabor algodão doce colorido e divertido",
         3, 2),
    ]

    # Corrigir produtos antigos: "Sobremesa" → "Sorvete"
    Produto.query.filter(
        Produto.categoria.ilike("Sobremesa")
    ).update({"categoria": "Sorvete"}, synchronize_session=False)

    # Inserir apenas produtos que ainda não existem (por nome)
    nomes_existentes = {
        p.nome_produto
        for p in Produto.query.with_entities(Produto.nome_produto).all()
    }
    novos = 0

    for nome, vol, preco, desc, est_at, est_min in _acais:
        if nome not in nomes_existentes:
            db.session.add(Produto(
                nome_produto=nome, categoria="Açaí", volume=vol,
                preco=Decimal(str(preco)), descricao=desc,
                estoque_atual=est_at, estoque_minimo=est_min,
            ))
            novos += 1

    for nome, vol, preco, desc, est_at, est_min in _sorvetes:
        if nome not in nomes_existentes:
            db.session.add(Produto(
                nome_produto=nome, categoria="Sorvete", volume=vol,
                preco=Decimal(str(preco)), descricao=desc,
                estoque_atual=est_at, estoque_minimo=est_min,
            ))
            novos += 1

    # Atualizar produtos existentes sem volume para incluir vol/estmin
    for nome, vol, preco, desc, est_at, est_min in _acais + _sorvetes:
        if nome in nomes_existentes:
            Produto.query.filter_by(nome_produto=nome).update({
                "volume": vol,
                "estoque_minimo": est_min,
            }, synchronize_session=False)

    db.session.commit()
    if novos:
        logger.info("Catálogo semeado: %d novos produtos adicionados", novos)
    else:
        logger.info("Catálogo verificado: nenhum produto novo necessário")


def _seed_complementos():
    """Semeia complementos/toppings reais de açaiteria.

    Adiciona apenas os que não existem (por nome). Nunca deleta
    para não quebrar FKs de vendas realizadas.
    """
    _complementos = [
        # Frutas
        ("Morango", "Fruta", "porção", 3.00),
        ("Banana", "Fruta", "porção", 2.00),
        ("Manga", "Fruta", "porção", 3.00),
        ("Kiwi", "Fruta", "porção", 4.00),
        ("Uva", "Fruta", "porção", 3.50),
        ("Abacaxi", "Fruta", "porção", 2.50),
        ("Blueberry", "Fruta", "porção", 5.00),
        # Caldas
        ("Leite Condensado", "Calda", "porção", 2.00),
        ("Calda de Morango", "Calda", "porção", 2.00),
        ("Calda de Chocolate", "Calda", "porção", 2.00),
        ("Calda de Caramelo", "Calda", "porção", 2.50),
        ("Mel", "Calda", "porção", 3.00),
        ("Creme de Avelã", "Calda", "porção", 4.00),
        # Farináceos / Crocantes
        ("Granola", "Farináceo", "porção", 2.00),
        ("Paçoca Triturada", "Farináceo", "porção", 2.50),
        ("Farinha Láctea", "Farináceo", "porção", 2.00),
        ("Aveia", "Farináceo", "porção", 1.50),
        ("Sucrilhos", "Farináceo", "porção", 2.00),
        ("Granulado de Chocolate", "Farináceo", "porção", 2.00),
        ("Coco Ralado", "Farináceo", "porção", 2.00),
        # Extras
        ("Leite Ninho", "Extra", "porção", 3.00),
        ("Confete / M&M's", "Extra", "porção", 4.00),
        ("Bis Triturado", "Extra", "porção", 3.50),
        ("Castanha de Caju", "Extra", "porção", 4.00),
        ("Amendoim Torrado", "Extra", "porção", 2.50),
        ("Chantilly", "Extra", "porção", 2.50),
        ("Jujuba", "Extra", "porção", 2.00),
        ("Tapioca", "Extra", "porção", 2.50),
    ]

    nomes_existentes = {
        c.nome
        for c in Complemento.query.with_entities(Complemento.nome).all()
    }
    novos = 0

    for nome, cat, unid, preco in _complementos:
        if nome not in nomes_existentes:
            db.session.add(Complemento(
                nome=nome,
                categoria=cat,
                unidade_medida=unid,
                preco_adicional=Decimal(str(preco)),
                ativo=True,
            ))
            novos += 1

    db.session.commit()
    if novos:
        logger.info("Complementos semeados: %d novos toppings", novos)
    else:
        logger.info("Complementos verificados: nenhum novo necessário")


# =============================================================================
# CRIAR TABELAS E SEED
# (usado pelo gunicorn na nuvem — tabelas já criadas acima)
# =============================================================================

with app.app_context():
    _seed_admin()
    _seed_produtos()
    _seed_complementos()


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port)
